"""
Bloomberg Futures Strip Data Downloader
=======================================
Downloads full futures strip data (F1 to Fn) for user-specified tickers,
fields, and date range. Outputs a formatted Excel workbook with
one sheet per commodity — Price & Volume side by side.

Requirements:
    pip install blpapi pdblp openpyxl pandas
    Bloomberg Terminal must be running on same machine or accessible via SAPI.

Usage:
    python bloomberg_data_downloader.py
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os


# ─────────────────────────────────────────────
# 1. USER CONFIGURATION
# ─────────────────────────────────────────────

def get_user_inputs():
    """Collect all inputs from user interactively."""
    print("\n" + "=" * 65)
    print("   BLOOMBERG FUTURES STRIP DATA DOWNLOADER")
    print("=" * 65)

    # --- Tickers ---
    print("\nEnter commodity tickers (Bloomberg root codes).")
    print("Example: LP, LA, GC, LX, LN, HO, XB, CL, CO, NG")
    print("For NGL: BAP, CAP, DAE")
    ticker_input = input("\nTickers (comma-separated): ").strip()
    tickers = [t.strip().upper() for t in ticker_input.split(",") if t.strip()]

    # --- Asset class hint for suffix ---
    print("\nTicker suffix convention:")
    print("  1) Generic numbered (e.g. LP1 Comdty, LP2 Comdty)  [LME / CME standard]")
    print("  2) Month-Year codes (e.g. CLZ24 Comdty)             [manual / specific]")
    suffix_choice = input("Use generic numbered strip? (Y/n): ").strip().lower()
    use_generic = suffix_choice != "n"

    # --- Max contract months ---
    max_months_input = input("\nMax number of futures contracts per ticker (default 27): ").strip()
    max_months = int(max_months_input) if max_months_input else 27

    # --- Fields ---
    print("\nBloomberg fields to download.")
    print("Common: PX_LAST, PX_VOLUME, PX_OPEN, PX_HIGH, PX_LOW, OPEN_INT")
    print("Vol:    IVOL_MID, HIST_VOL_30D, HIST_VOL_60D, HIST_VOL_90D")
    fields_input = input("\nFields (comma-separated, default PX_LAST,PX_VOLUME): ").strip()
    if not fields_input:
        fields = ["PX_LAST", "PX_VOLUME"]
    else:
        fields = [f.strip().upper() for f in fields_input.split(",") if f.strip()]

    # --- Date Range ---
    start_input = input("\nStart date (YYYYMMDD, default 20050101): ").strip()
    start_date = start_input if start_input else "20050101"

    end_input = input("End date   (YYYYMMDD, default today):    ").strip()
    end_date = end_input if end_input else datetime.today().strftime("%Y%m%d")

    # --- Output file ---
    output_input = input("\nOutput filename (default bloomberg_futures_data.xlsx): ").strip()
    output_file = output_input if output_input else "bloomberg_futures_data.xlsx"

    # --- Bloomberg connection ---
    print("\nBloomberg connection:")
    print("  1) Desktop API (localhost:8194)  — default")
    print("  2) Server API (SAPI)            — enter host:port")
    conn_choice = input("Choice (1/2, default 1): ").strip()
    if conn_choice == "2":
        host = input("  SAPI host: ").strip()
        port = int(input("  SAPI port: ").strip())
    else:
        host = "localhost"
        port = 8194

    return {
        "tickers": tickers,
        "use_generic": use_generic,
        "max_months": max_months,
        "fields": fields,
        "start_date": start_date,
        "end_date": end_date,
        "output_file": output_file,
        "host": host,
        "port": port,
    }


# ─────────────────────────────────────────────
# 2. BUILD TICKER UNIVERSE
# ─────────────────────────────────────────────

def build_ticker_list(root, max_months, use_generic=True):
    """
    Build list of Bloomberg tickers for a futures strip.
    e.g. root='LP', max_months=27  →  ['LP1 Comdty', 'LP2 Comdty', ..., 'LP27 Comdty']
    """
    if use_generic:
        return [f"{root}{i} Comdty" for i in range(1, max_months + 1)]
    else:
        return [f"{root}{i} Comdty" for i in range(1, max_months + 1)]


# ─────────────────────────────────────────────
# 3. CONNECT & DOWNLOAD
# ─────────────────────────────────────────────

def connect_bloomberg(host="localhost", port=8194):
    """Establish Bloomberg connection via pdblp."""
    print(f"\nConnecting to Bloomberg at {host}:{port} ...")
    con = pdblp.BCon(debug=False, host=host, port=port, timeout=30000)
    con.start()
    print("✓ Connected successfully.\n")
    return con


def download_data(con, tickers, fields, start_date, end_date):
    """
    Download historical data for all tickers and fields.
    Returns a dict: {ticker: DataFrame with DatetimeIndex, columns = fields}
    Falls back gracefully if a ticker has no data.
    """
    data = {}
    total = len(tickers)
    for idx, tkr in enumerate(tickers, 1):
        pct = idx / total * 100
        print(f"  [{idx:3d}/{total}] ({pct:5.1f}%) Downloading {tkr} ...", end="")
        try:
            df = con.bdh(tkr, fields, start_date, end_date)
            if df.empty:
                print(" ⚠ No data returned.")
                continue
            # pdblp returns multi-level columns: (ticker, field)
            # Flatten to just field names
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(1)
            data[tkr] = df
            print(f" ✓ {len(df)} rows")
        except Exception as e:
            print(f" ✗ Error: {e}")
    return data


# ─────────────────────────────────────────────
# 4. ORGANIZE & WRITE EXCEL
# ─────────────────────────────────────────────

# Style constants
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(name="Arial", bold=True, color="1F4E79", size=9)
DATA_FONT = Font(name="Arial", size=9)
DATE_FONT = Font(name="Arial", size=9, bold=True)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="D0D0D0"),
)
NUMBER_FMT = "#,##0.00"
VOLUME_FMT = "#,##0"
DATE_FMT = "YYYY-MM-DD"


def get_field_format(field_name):
    """Return appropriate number format for a Bloomberg field."""
    field_upper = field_name.upper()
    if "VOL" in field_upper and "PX" not in field_upper:
        return "0.00%"
    if "VOLUME" in field_upper or "OPEN_INT" in field_upper:
        return VOLUME_FMT
    return NUMBER_FMT


def write_commodity_sheet(wb, root_ticker, strip_data, fields, max_months):
    """
    Write one sheet per commodity.
    Layout: Date | F1_Price | F1_Volume | F2_Price | F2_Volume | ...
    Each pair grouped under a contract header.
    """
    sheet_name = root_ticker.replace(" ", "_")[:31]
    ws = wb.create_sheet(title=sheet_name)

    # Collect all dates across all contracts
    all_dates = set()
    for tkr, df in strip_data.items():
        all_dates.update(df.index.tolist())
    all_dates = sorted(all_dates)

    if not all_dates:
        ws["A1"] = f"No data available for {root_ticker}"
        return

    # --- Row 1: Commodity title bar ---
    n_fields = len(fields)
    contracts_with_data = sorted(
        strip_data.keys(),
        key=lambda t: int("".join(filter(str.isdigit, t.split()[0].replace(root_ticker, ""))) or "0")
    )
    total_cols = 1 + len(contracts_with_data) * n_fields

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{root_ticker} — Futures Strip Data"
    title_cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    title_cell.fill = PatternFill("solid", fgColor="0D3B66")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # --- Row 2: Contract group headers (F1, F2, ...) ---
    ws.cell(row=2, column=1, value="Date")
    ws.cell(row=2, column=1).font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    col = 2
    for tkr in contracts_with_data:
        contract_num = "".join(filter(str.isdigit, tkr.split()[0].replace(root_ticker, "")))
        label = f"F{contract_num}"

        if n_fields > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + n_fields - 1)
        cell = ws.cell(row=2, column=col)
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

        for j in range(n_fields):
            if j > 0:
                ws.cell(row=2, column=col + j).fill = HEADER_FILL

        col += n_fields

    # --- Row 3: Field sub-headers (PX_LAST, PX_VOLUME, ...) ---
    ws.cell(row=3, column=1, value="")
    ws.cell(row=3, column=1).fill = SUBHEADER_FILL

    col = 2
    field_display = {
        "PX_LAST": "Price",
        "PX_VOLUME": "Volume",
        "PX_OPEN": "Open",
        "PX_HIGH": "High",
        "PX_LOW": "Low",
        "OPEN_INT": "Open Int",
        "IVOL_MID": "Impl Vol",
        "HIST_VOL_30D": "HV 30D",
        "HIST_VOL_60D": "HV 60D",
        "HIST_VOL_90D": "HV 90D",
    }
    for tkr in contracts_with_data:
        for fld in fields:
            cell = ws.cell(row=3, column=col)
            cell.value = field_display.get(fld, fld)
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            col += 1

    # --- Data rows ---
    for row_idx, dt in enumerate(all_dates, start=4):
        date_cell = ws.cell(row=row_idx, column=1)
        date_cell.value = dt
        date_cell.number_format = DATE_FMT
        date_cell.font = DATE_FONT
        date_cell.border = THIN_BORDER

        col = 2
        for tkr in contracts_with_data:
            df = strip_data[tkr]
            for fld in fields:
                cell = ws.cell(row=row_idx, column=col)
                if dt in df.index and fld in df.columns:
                    val = df.loc[dt, fld]
                    if pd.notna(val):
                        cell.value = val
                cell.font = DATA_FONT
                cell.number_format = get_field_format(fld)
                cell.border = THIN_BORDER
                col += 1

    # --- Column widths ---
    ws.column_dimensions["A"].width = 13
    for c in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    # Freeze panes: lock date column and header rows
    ws.freeze_panes = "B4"

    print(f"  ✓ Sheet '{sheet_name}' — {len(all_dates)} rows × {len(contracts_with_data)} contracts × {n_fields} fields")


def write_summary_sheet(wb, config, download_stats):
    """Write a metadata / summary sheet as the first tab."""
    ws = wb.active
    ws.title = "README"

    rows = [
        ("Bloomberg Futures Data Download", ""),
        ("", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Start Date", config["start_date"]),
        ("End Date", config["end_date"]),
        ("Fields", ", ".join(config["fields"])),
        ("Max Contracts", config["max_months"]),
        ("", ""),
        ("COMMODITY", "CONTRACTS WITH DATA"),
    ]
    for root, count in download_stats.items():
        rows.append((root, count))

    rows.extend([
        ("", ""),
        ("NOTES", ""),
        ("", "Each sheet contains one commodity's full futures strip."),
        ("", "Columns are grouped by contract: F1, F2, ... Fn."),
        ("", "Within each group: Price and Volume (or other fields) side by side."),
        ("", "Dates are aligned across all contracts for easy comparison."),
        ("", "Blank cells indicate no data for that contract on that date."),
    ])

    for r, (a, b) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)

    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0D3B66")
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55


# ─────────────────────────────────────────────
# 5. MAIN
# ─────────────────────────────────────────────

def main():
    config = get_user_inputs()

    # Connect
    con = connect_bloomberg(config["host"], config["port"])

    # Build universe & download
    download_stats = {}
    all_strip_data = {}

    for root in config["tickers"]:
        print(f"\n{'─' * 50}")
        print(f"  {root} — Building futures strip (F1 to F{config['max_months']})")
        print(f"{'─' * 50}")

        strip_tickers = build_ticker_list(root, config["max_months"], config["use_generic"])
        strip_data = download_data(con, strip_tickers, config["fields"], config["start_date"], config["end_date"])

        all_strip_data[root] = strip_data
        download_stats[root] = len(strip_data)
        print(f"\n  {root}: {len(strip_data)} contracts returned data.\n")

    # Build Excel workbook
    print("\n" + "=" * 50)
    print("  WRITING EXCEL WORKBOOK")
    print("=" * 50)

    wb = Workbook()
    write_summary_sheet(wb, config, download_stats)

    for root in config["tickers"]:
        if all_strip_data[root]:
            write_commodity_sheet(
                wb, root, all_strip_data[root],
                config["fields"], config["max_months"]
            )

    wb.save(config["output_file"])
    print(f"\n✓ Saved: {os.path.abspath(config['output_file'])}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"\nDone.\n")

    # Disconnect
    try:
        con.stop()
    except Exception:
        pass


if __name__ == "__main__":
    main()
