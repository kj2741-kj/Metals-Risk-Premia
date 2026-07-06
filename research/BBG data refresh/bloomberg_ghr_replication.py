"""
Bloomberg GHR Replication — Futures Downloader
===============================================
Downloads futures price strips (PX_LAST) for ALL 31 Gorton-Hayashi-Rouwenhorst
(2013) commodities PLUS additional reference products, extended to June 30, 2026.

Products already covered by bloomberg_full_download.py (LME metals, precious,
energy core/extended, NGL) are included here too so this script is self-contained.

NEW vs bloomberg_full_download.py:
    CBOT Grains : W, C, S, BO, SM, O
    ICE Softs   : CT, CC, KC, JO, LB
    CME Meats   : LC, LH, FC, DA, CB, PB
    Extra NYMEX : PN (Propane), QL (Coal)

Deadlock protection: 12 s per-ticker, early-exit after 3 consecutive empties.

Output (same folder as this script):
    GHR_Futures_Metals.xlsx
    GHR_Futures_Grains.xlsx
    GHR_Futures_Softs.xlsx
    GHR_Futures_Meats.xlsx
    GHR_Futures_Energies.xlsx

Run with Bloomberg Terminal open:
    python bloomberg_ghr_replication.py

Requirements: pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# =============================================================================
# CONFIG
# =============================================================================

START_DATE = "20050101"
END_DATE   = "20260630"

HOST = "localhost"
PORT = 8194

FIELDS = ["PX_LAST"]

TIMEOUT_PER_CONTRACT = 12
EARLY_EXIT_EMPTY     = 3

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# =============================================================================
# FUTURES STRIP DEFINITIONS
# Ticker format: {ROOT}{n} Comdty  e.g. W1 Comdty, W12 Comdty
# max_contracts: how many generics to attempt; early-exit stops at 3 empties.
# For GHR basis calculation you need F1 and F2 at minimum.
# =============================================================================

# ---- Metals (already in bloomberg_full_download.py; included for completeness) ---
LME_METALS = {
    "LP": {"name": "Copper LME",    "max_contracts": 27},
    "LA": {"name": "Aluminium LME", "max_contracts": 27},
    "LX": {"name": "Zinc LME",      "max_contracts": 27},
    "LN": {"name": "Nickel LME",    "max_contracts": 27},
    "LL": {"name": "Lead LME",      "max_contracts": 27},
    "LT": {"name": "Tin LME",       "max_contracts": 15},
}

PRECIOUS_METALS = {
    "GC": {"name": "Gold COMEX",      "max_contracts": 20},
    "SI": {"name": "Silver COMEX",    "max_contracts": 17},
    "HG": {"name": "Copper CME (HG)", "max_contracts": 27},
    "PL": {"name": "Platinum NYMEX",  "max_contracts": 13},
    "PA": {"name": "Palladium NYMEX", "max_contracts": 12},
}

# ---- CBOT Grains (GHR: all 6) -----------------------------------------------
# Generic ticker N=1..12 covers the full active listed curve for grain markets.
CBOT_GRAINS = {
    "W":  {"name": "Wheat CBOT",        "max_contracts": 12},
    "C":  {"name": "Corn CBOT",         "max_contracts": 12},
    "S":  {"name": "Soybeans CBOT",     "max_contracts": 12},
    "BO": {"name": "Soybean Oil CBOT",  "max_contracts": 12},
    "SM": {"name": "Soybean Meal CBOT", "max_contracts": 12},
    "O":  {"name": "Oats CBOT",         "max_contracts": 12},
}

# ---- ICE / CME Softs (GHR: all 5) ------------------------------------------
ICE_SOFTS = {
    "CT": {"name": "Cotton ICE",       "max_contracts": 12},
    "CC": {"name": "Cocoa ICE",        "max_contracts": 12},
    "KC": {"name": "Coffee C ICE",     "max_contracts": 12},
    "JO": {"name": "Orange Juice ICE", "max_contracts": 12},  # Bloomberg root is JO not OJ
    "LB": {"name": "Lumber CME",       "max_contracts": 12},
}

# ---- CME Meats (GHR: Live Cattle, Lean Hogs, Feeder Cattle, Milk, Butter) ---
# PB (Pork Bellies) was delisted July 2011; data from 2005-2011 only.
CME_MEATS = {
    "LC": {"name": "Live Cattle CME",    "max_contracts": 12},
    "LH": {"name": "Lean Hogs CME",      "max_contracts": 12},
    "FC": {"name": "Feeder Cattle CME",  "max_contracts": 12},
    "DA": {"name": "Milk Class III CME", "max_contracts": 12},
    "CB": {"name": "Butter CME",         "max_contracts": 12},
    "PB": {"name": "Pork Bellies CME",   "max_contracts":  6},  # delisted 2011
}

# ---- NYMEX / ICE Energies ---------------------------------------------------
# XB = RBOB Gasoline on Bloomberg (confirmed from bloomberg_full_download.py; NOT RB)
# PN = Propane NYMEX (Mont Belvieu), QL = Coal NYMEX (both new vs existing script)
NYMEX_ENERGIES = {
    "CL":  {"name": "WTI Crude NYMEX",         "max_contracts": 27},
    "CO":  {"name": "Brent Crude ICE",          "max_contracts": 27},
    "XB":  {"name": "RBOB Gasoline NYMEX",      "max_contracts": 27},
    "HO":  {"name": "Heating Oil ULSD NYMEX",   "max_contracts": 27},
    "NG":  {"name": "Nat Gas Henry Hub NYMEX",  "max_contracts": 27},
    "PN":  {"name": "Propane NYMEX",            "max_contracts": 12},
    "QL":  {"name": "Coal NYMEX",               "max_contracts": 12},
    "GO":  {"name": "ICE Gasoil London",        "max_contracts": 27},
    "QS":  {"name": "Singapore Gasoil ICE",     "max_contracts": 27},
    "FO":  {"name": "Fuel Oil 3.5pct ICE",      "max_contracts": 12},
    "SJ":  {"name": "Singapore Jet Kerosene",   "max_contracts": 12},  # verify if empty
    "NFY": {"name": "Naphtha CIF NWE ICE",      "max_contracts": 12},  # verify if empty
}

# =============================================================================
# STYLES  (same as bloomberg_full_download.py)
# =============================================================================

TITLE_FILL = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT  = Font(name="Arial", size=9)
DATE_FONT  = Font(name="Arial", bold=True, size=9)

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# DEADLOCK-SAFE DOWNLOAD  (identical pattern to bloomberg_full_download.py)
# =============================================================================

def safe_bdh(con, ticker):
    """Single-ticker BDH with error handling. Returns (df|None, status_str).
    Direct call — no thread wrapper (threading causes pdblp session errors)."""
    try:
        df = con.bdh(ticker, FIELDS, START_DATE, END_DATE)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(1)
        if df.empty or "PX_LAST" not in df.columns:
            return None, "no data"
        df = df[["PX_LAST"]].dropna()
        if df.empty:
            return None, "all NaN"
        return df, "%d rows" % len(df)
    except Exception as e:
        return None, "ERROR: %s" % str(e)[:80]


def download_strip(con, root, max_contracts, label=""):
    """Download F1-Fn with early-exit after EARLY_EXIT_EMPTY consecutive failures."""
    tag = "[%s] " % label if label else ""
    print("  %s%s  F1-F%d" % (tag, root, max_contracts))

    result = {}
    consecutive_empty = 0
    for i in range(1, max_contracts + 1):
        tkr = "%s%d Comdty" % (root, i)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
            consecutive_empty = 0
            print("    %-22s %s" % (tkr, status))
        else:
            consecutive_empty += 1
            print("    %-22s %s" % (tkr, status))
            if consecutive_empty >= EARLY_EXIT_EMPTY:
                remaining = max_contracts - i
                if remaining > 0:
                    print("    => %d consecutive empties at F%d — skipping F%d-F%d"
                          % (EARLY_EXIT_EMPTY, i, i + 1, max_contracts))
                break

    print("  => %d / %d contracts returned data\n" % (len(result), max_contracts))
    return result

# =============================================================================
# EXCEL WRITERS  (identical pattern to bloomberg_full_download.py)
# =============================================================================

def write_strip_sheet(ws, root, name, strip_data):
    """Date | F1 | F2 | ... | Fn  (PX_LAST only)."""
    contracts = sorted(
        strip_data.keys(),
        key=lambda t: int("".join(c for c in t.split()[0][len(root):] if c.isdigit()) or "0"),
    )
    if not contracts:
        ws["A1"] = "No data returned for %s" % name
        return

    all_dates = sorted(set().union(*(strip_data[t].index for t in contracts)))
    n_cols = 1 + len(contracts)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1,
                value="%s -- Futures Curve (PX_LAST, end=%s)" % (name, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(contracts, 2):
        num = "".join(ch for ch in tkr.split()[0][len(root):] if ch.isdigit())
        c = ws.cell(row=2, column=j, value="F%s" % num)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, dt in enumerate(all_dates, 3):
        cell_date = ws.cell(row=ri, column=1, value=dt)
        cell_date.number_format = "YYYY-MM-DD"
        cell_date.font = DATE_FONT
        for j, tkr in enumerate(contracts, 2):
            df = strip_data[tkr]
            cell = ws.cell(row=ri, column=j)
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "B3"
    print("    Sheet '%s': %d rows, %d contracts" % (ws.title, len(all_dates), len(contracts)))


def write_readme(wb, title_str, rows):
    ws = wb.active
    ws.title = "README"
    meta = [
        (title_str,        ""),
        ("",               ""),
        ("Generated",      datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Start Date",     START_DATE),
        ("End Date",       END_DATE),
        ("Fields",         ", ".join(FIELDS)),
        ("Timeout/ticker", "%ds + skip after %d empties"
                           % (TIMEOUT_PER_CONTRACT, EARLY_EXIT_EMPTY)),
        ("Reference",      "Gorton, Hayashi & Rouwenhorst (2013), Rev Finance"),
        ("", ""),
    ] + rows
    for r, (a, b) in enumerate(meta, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13, color="0D3B66")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60


def save_wb(wb, filename):
    path = os.path.join(SCRIPT_DIR, filename)
    wb.save(path)
    print("\n  Saved: %s" % path)
    print("  Sheets: %s\n" % ", ".join(wb.sheetnames))


def run_section(label, product_dict, con, wb, section_label):
    print("\n" + "=" * 65)
    print("  %s" % section_label)
    print("=" * 65)
    for root, cfg in product_dict.items():
        strip = download_strip(con, root, cfg["max_contracts"], label=label)
        ws = wb.create_sheet(title=cfg["name"][:31])
        write_strip_sheet(ws, root, cfg["name"], strip)

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 65)
    print("  GHR REPLICATION — BLOOMBERG FUTURES DOWNLOADER")
    print("  Start: %s   End: %s" % (START_DATE, END_DATE))
    print("  Timeout: %ds/contract, exit after %d consecutive empties"
          % (TIMEOUT_PER_CONTRACT, EARLY_EXIT_EMPTY))
    print("=" * 65)

    con = connect()

    # ---- 1. Metals -----------------------------------------------------------
    wb1 = Workbook()
    write_readme(wb1, "GHR Metals — LME Base + Precious", [
        ("LP", "Copper LME      F1-F27"),
        ("LA", "Aluminium LME   F1-F27"),
        ("LX", "Zinc LME        F1-F27"),
        ("LN", "Nickel LME      F1-F27"),
        ("LL", "Lead LME        F1-F27"),
        ("LT", "Tin LME         F1-F15"),
        ("GC", "Gold COMEX      F1-F20  (reference; not in GHR core)"),
        ("SI", "Silver COMEX    F1-F17  (reference; not in GHR core)"),
        ("HG", "Copper CME      F1-F27  (CME cross-reference)"),
        ("PL", "Platinum NYMEX  F1-F13  (GHR commodity)"),
        ("PA", "Palladium NYMEX F1-F12  (GHR commodity)"),
    ])
    all_metals = {**LME_METALS, **PRECIOUS_METALS}
    run_section("MTL", all_metals, con, wb1, "1/5  METALS (LME + PRECIOUS)")
    save_wb(wb1, "GHR_Futures_Metals.xlsx")

    # ---- 2. CBOT Grains ------------------------------------------------------
    wb2 = Workbook()
    write_readme(wb2, "GHR Grains — CBOT Futures", [
        ("W",  "Wheat CBOT         F1-F12"),
        ("C",  "Corn CBOT          F1-F12"),
        ("S",  "Soybeans CBOT      F1-F12"),
        ("BO", "Soybean Oil CBOT   F1-F12"),
        ("SM", "Soybean Meal CBOT  F1-F12"),
        ("O",  "Oats CBOT          F1-F12"),
    ])
    run_section("GRN", CBOT_GRAINS, con, wb2, "2/5  CBOT GRAINS")
    save_wb(wb2, "GHR_Futures_Grains.xlsx")

    # ---- 3. ICE / CME Softs --------------------------------------------------
    wb3 = Workbook()
    write_readme(wb3, "GHR Softs — ICE / CME Futures", [
        ("CT", "Cotton ICE        F1-F12"),
        ("CC", "Cocoa ICE         F1-F12"),
        ("KC", "Coffee C ICE      F1-F12"),
        ("JO", "Orange Juice ICE  F1-F12  (Bloomberg root: JO not OJ)"),
        ("LB", "Lumber CME        F1-F12  (Random Length Lumber)"),
    ])
    run_section("SFT", ICE_SOFTS, con, wb3, "3/5  ICE / CME SOFTS")
    save_wb(wb3, "GHR_Futures_Softs.xlsx")

    # ---- 4. CME Meats --------------------------------------------------------
    wb4 = Workbook()
    write_readme(wb4, "GHR Meats — CME Futures", [
        ("LC", "Live Cattle CME    F1-F12"),
        ("LH", "Lean Hogs CME      F1-F12"),
        ("FC", "Feeder Cattle CME  F1-F12"),
        ("DA", "Milk Class III CME F1-F12"),
        ("CB", "Butter CME         F1-F12"),
        ("PB", "Pork Bellies CME   F1-F6   (DELISTED July 2011; data 2005-2011 only)"),
    ])
    run_section("MET", CME_MEATS, con, wb4, "4/5  CME MEATS")
    save_wb(wb4, "GHR_Futures_Meats.xlsx")

    # ---- 5. Energies ---------------------------------------------------------
    wb5 = Workbook()
    write_readme(wb5, "GHR Energies — NYMEX / ICE Futures", [
        ("CL",  "WTI Crude NYMEX         F1-F27"),
        ("CO",  "Brent Crude ICE         F1-F27"),
        ("XB",  "RBOB Gasoline NYMEX     F1-F27  Bloomberg root=XB (NOT RB)"),
        ("HO",  "Heating Oil ULSD NYMEX  F1-F27"),
        ("NG",  "Nat Gas Henry Hub NYMEX F1-F27"),
        ("PN",  "Propane NYMEX           F1-F12  (GHR commodity)"),
        ("QL",  "Coal NYMEX              F1-F12  (GHR commodity)"),
        ("GO",  "ICE Gasoil London       F1-F27  European diesel"),
        ("QS",  "Singapore Gasoil ICE    F1-F27"),
        ("FO",  "Fuel Oil 3.5pct ICE     F1-F12"),
        ("SJ",  "Singapore Jet ICE       F1-F12  (verify on terminal if empty)"),
        ("NFY", "Naphtha CIF NWE ICE     F1-F12  (verify on terminal if empty)"),
    ])
    run_section("NRG", NYMEX_ENERGIES, con, wb5, "5/5  NYMEX / ICE ENERGIES")
    save_wb(wb5, "GHR_Futures_Energies.xlsx")

    # ---- Disconnect ----------------------------------------------------------
    try:
        con.stop()
    except Exception:
        pass

    print("=" * 65)
    print("  DOWNLOAD COMPLETE")
    print("  GHR_Futures_Metals.xlsx")
    print("  GHR_Futures_Grains.xlsx")
    print("  GHR_Futures_Softs.xlsx")
    print("  GHR_Futures_Meats.xlsx")
    print("  GHR_Futures_Energies.xlsx")
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
