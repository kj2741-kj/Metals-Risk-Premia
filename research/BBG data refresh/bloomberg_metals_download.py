"""
Bloomberg Metals Data Downloader — PRE-CONFIGURED
===================================================
Pre-filled for the Metals Risk Premia project.
Just run it with Bloomberg Terminal open.

    python bloomberg_metals_download.py

Requirements:
    pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

# ═══════════════════════════════════════════════
# CONFIG — EDIT THESE AS NEEDED
# ═══════════════════════════════════════════════

CONFIG = {
    # ── Date Range ──
    "start_date": "20050101",
    "end_date": datetime.today().strftime("%Y%m%d"),

    # ── Bloomberg Connection ──
    "host": "localhost",
    "port": 8194,

    # ── Downloads to run (each becomes a separate Excel file) ──
    # Set any to False to skip
    "download_lme_benchmarks": True,   # ★ LME official: Cash, 3M, 15M, 27M (most liquid)
    "download_futures_strip": True,    # Generic monthly strip LP1-LP27 (synthetic rolls)
    "download_cash_carry": True,       # Carry: Cash-3M spreads, full term structure
    "download_vol_surface": True,
    "download_fundamentals": True,
    "download_cross_metal": True,
}

# ── 1A. LME OFFICIAL BENCHMARK INSTRUMENTS ──
# These are the ACTUAL traded instruments on LME — the 3M forward
# is the primary liquidity point, NOT the generic monthly strip.
#
# LME STRUCTURE:
#   Cash (spot)   → highest LME turnover after 3M
#   Tom (T+1)     → next-day settlement
#   3-Month (3M)  → ★ BENCHMARK — most liquid instrument ★
#   15-Month      → medium-term reference
#   27-Month      → long-dated reference
#   Cash-3M Spread → direct carry/backwardation signal
#
# The generic strip (LP1, LP2...) are Bloomberg's synthetic
# monthly roll constructs — useful for term structure analysis
# but NOT the actual LME trading instruments.

LME_BENCHMARKS = {
    "Copper": {
        "LMCADS03 Comdty":  "Cu Cash (Spot)",
        "LMCADT03 Comdty":  "Cu Tom (T+1)",
        "LMCADY03 Comdty":  "Cu 3M ★",          # Primary benchmark
        "LMCA15M3 Comdty":  "Cu 15M",
        "LMCA27M3 Comdty":  "Cu 27M",
        "LMCASP03 Comdty":  "Cu Cash-3M Spread",
    },
    "Aluminium": {
        "LMAHDS03 Comdty":  "Al Cash (Spot)",
        "LMAHDT03 Comdty":  "Al Tom (T+1)",
        "LMAHDY03 Comdty":  "Al 3M ★",
        "LMAH15M3 Comdty":  "Al 15M",
        "LMAH27M3 Comdty":  "Al 27M",
        "LMAHSP03 Comdty":  "Al Cash-3M Spread",
    },
    "Zinc": {
        "LMZSDS03 Comdty":  "Zn Cash (Spot)",
        "LMZSDT03 Comdty":  "Zn Tom (T+1)",
        "LMZSDY03 Comdty":  "Zn 3M ★",
        "LMZS15M3 Comdty":  "Zn 15M",
        "LMZS27M3 Comdty":  "Zn 27M",
        "LMZSSP03 Comdty":  "Zn Cash-3M Spread",
    },
    "Nickel": {
        "LMNIDS03 Comdty":  "Ni Cash (Spot)",
        "LMNIDT03 Comdty":  "Ni Tom (T+1)",
        "LMNIDY03 Comdty":  "Ni 3M ★",
        "LMNI15M3 Comdty":  "Ni 15M",
        "LMNI27M3 Comdty":  "Ni 27M",
        "LMNISP03 Comdty":  "Ni Cash-3M Spread",
    },
    "Lead": {
        "LMPBDS03 Comdty":  "Pb Cash (Spot)",
        "LMPBDY03 Comdty":  "Pb 3M ★",
        "LMPB15M3 Comdty":  "Pb 15M",
        "LMPBSP03 Comdty":  "Pb Cash-3M Spread",
    },
    "Tin": {
        "LMSNDS03 Comdty":  "Sn Cash (Spot)",
        "LMSNDY03 Comdty":  "Sn 3M ★",
        "LMSN15M3 Comdty":  "Sn 15M",
        "LMSNSP03 Comdty":  "Sn Cash-3M Spread",
    },
}
LME_BENCHMARK_FIELDS = ["PX_LAST", "PX_VOLUME", "OPEN_INT"]

# ── 1B. GENERIC MONTHLY STRIP (Bloomberg synthetic rolls) ──
# Useful for building continuous curves and momentum signals,
# but remember these are LESS liquid than the 3M benchmark.
FUTURES_STRIP = {
    "LP":  {"name": "Copper (LME)",      "max_months": 27, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "LA":  {"name": "Aluminium (LME)",    "max_months": 27, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "GC":  {"name": "Gold (COMEX)",       "max_months": 12, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "LX":  {"name": "Zinc (LME)",         "max_months": 27, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "LN":  {"name": "Nickel (LME)",       "max_months": 27, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "SI":  {"name": "Silver (COMEX)",     "max_months": 12, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
    "HG":  {"name": "Copper (COMEX)",     "max_months": 12, "fields": ["PX_LAST", "PX_VOLUME", "OPEN_INT"]},
}

# ── 2. CASH / CARRY / TERM STRUCTURE ──
# Dedicated carry sheet: Cash, 3M, 15M, 27M, spreads
# (Separated from benchmarks for cleaner carry signal construction)
CASH_CARRY_TICKERS = {
    # Copper — full term structure
    "LMCADS03 Comdty":  "Cu Cash",
    "LMCADY03 Comdty":  "Cu 3M",
    "LMCA15M3 Comdty":  "Cu 15M",
    "LMCA27M3 Comdty":  "Cu 27M",
    "LMCASP03 Comdty":  "Cu Cash-3M Spread",
    # Aluminium
    "LMAHDS03 Comdty":  "Al Cash",
    "LMAHDY03 Comdty":  "Al 3M",
    "LMAH15M3 Comdty":  "Al 15M",
    "LMAH27M3 Comdty":  "Al 27M",
    "LMAHSP03 Comdty":  "Al Cash-3M Spread",
    # Zinc
    "LMZSDS03 Comdty":  "Zn Cash",
    "LMZSDY03 Comdty":  "Zn 3M",
    "LMZS15M3 Comdty":  "Zn 15M",
    "LMZS27M3 Comdty":  "Zn 27M",
    "LMZSSP03 Comdty":  "Zn Cash-3M Spread",
    # Nickel
    "LMNIDS03 Comdty":  "Ni Cash",
    "LMNIDY03 Comdty":  "Ni 3M",
    "LMNI15M3 Comdty":  "Ni 15M",
    "LMNI27M3 Comdty":  "Ni 27M",
    "LMNISP03 Comdty":  "Ni Cash-3M Spread",
    # Gold — uses COMEX structure (no LME gold)
    "GOLDLNPM Index":   "Gold London PM Fix",
    "XAU Curncy":       "Gold Spot (XAU)",
    "GC1 Comdty":       "Gold F1 (COMEX)",
    "GC4 Comdty":       "Gold F4 (COMEX ~3M eq)",
    "GC12 Comdty":      "Gold F12 (COMEX ~1Y eq)",
    # Silver
    "SI1 Comdty":       "Silver F1 (COMEX)",
    "SI4 Comdty":       "Silver F4 (COMEX ~3M eq)",
}
CASH_CARRY_FIELDS = ["PX_LAST"]

# ── 3. VOLATILITY SURFACE ──
# Implied vol, realized vol, skew metrics
VOL_TICKERS = {
    # Copper
    "LP1 Comdty":   "Copper F1",
    "LP3 Comdty":   "Copper F3",
    "LP6 Comdty":   "Copper F6",
    "LP12 Comdty":  "Copper F12",
    # Aluminium
    "LA1 Comdty":   "Aluminium F1",
    "LA3 Comdty":   "Aluminium F3",
    "LA6 Comdty":   "Aluminium F6",
    "LA12 Comdty":  "Aluminium F12",
    # Gold
    "GC1 Comdty":   "Gold F1",
    "GC3 Comdty":   "Gold F3",
    "GC6 Comdty":   "Gold F6",
    "GC12 Comdty":  "Gold F12",
}
VOL_FIELDS = [
    "PX_LAST",
    "HIST_PUT_IMP_VOL",    # ATM implied vol
    "30DAY_IMPVOL_100.0%", # 30-day ATM IV
    "VOLATILITY_30D",      # 30-day realized vol
    "VOLATILITY_60D",      # 60-day realized vol
    "VOLATILITY_90D",      # 90-day realized vol
]

# ── 4. FUNDAMENTALS / MACRO ──
FUNDAMENTAL_TICKERS = {
    # LME Inventories
    "LCSNSTOT Index":   "LME Copper Inventory (MT)",
    "LASSNTOT Index":   "LME Aluminium Inventory (MT)",
    "LMZSNTOT Index":   "LME Zinc Inventory (MT)",
    "LMNITON Index":    "LME Nickel Inventory (MT)",
    # Cancelled Warrants
    "LCSNCANC Index":   "Copper Cancelled Warrants (%)",
    # Macro
    "DXY Curncy":       "US Dollar Index",
    "GTII10 Govt":      "US 10Y TIPS Yield",
    "USGG10YR Index":   "US 10Y Treasury Yield",
    "CPMINDX Index":    "China Mfg PMI",
    "MXWO Index":       "MSCI World Index",
    "CRY Index":        "CRB Commodity Index",
    # Energy cross-refs
    "CL1 Comdty":       "WTI Crude F1",
    "CO1 Comdty":       "Brent Crude F1",
}
FUNDAMENTAL_FIELDS = ["PX_LAST"]

# ── 5. CROSS-METAL RATIOS (will be computed, but pull raw data) ──
CROSS_METAL_TICKERS = {
    "LP1 Comdty":   "Copper F1",
    "LA1 Comdty":   "Aluminium F1",
    "GC1 Comdty":   "Gold F1",
    "SI1 Comdty":   "Silver F1",
    "LX1 Comdty":   "Zinc F1",
    "LN1 Comdty":   "Nickel F1",
    "HG1 Comdty":   "COMEX Copper F1",
}
CROSS_METAL_FIELDS = ["PX_LAST"]


# ═══════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════

TITLE_FILL = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT = Font(name="Arial", bold=True, color="1F4E79", size=9)
DATA_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", size=9, bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="C0C0C0"))

FIELD_DISPLAY = {
    "PX_LAST": "Price", "PX_VOLUME": "Volume", "PX_OPEN": "Open",
    "PX_HIGH": "High", "PX_LOW": "Low", "OPEN_INT": "Open Int",
    "HIST_PUT_IMP_VOL": "ATM IV", "30DAY_IMPVOL_100.0%": "30D IV",
    "VOLATILITY_30D": "RV 30D", "VOLATILITY_60D": "RV 60D",
    "VOLATILITY_90D": "RV 90D",
}

def fmt_field(f):
    return FIELD_DISPLAY.get(f, f.replace("_", " ").title())

def num_fmt(f):
    fu = f.upper()
    if "VOL" in fu and "PX" not in fu and "VOLUME" not in fu:
        return "0.00%"
    if "VOLUME" in fu or "OPEN_INT" in fu:
        return "#,##0"
    return "#,##0.00"


# ═══════════════════════════════════════════════
# DOWNLOAD ENGINE
# ═══════════════════════════════════════════════

def connect(host, port):
    print(f"\n  Connecting to Bloomberg ({host}:{port}) ...")
    con = pdblp.BCon(debug=False, host=host, port=port, timeout=30000)
    con.start()
    print("  ✓ Connected.\n")
    return con


def bulk_download(con, tickers, fields, start, end, label=""):
    """Download multiple tickers, return {ticker: DataFrame}."""
    data = {}
    total = len(tickers)
    for i, tkr in enumerate(tickers, 1):
        tag = f"[{label}]" if label else ""
        print(f"    {tag} ({i}/{total}) {tkr} ...", end="", flush=True)
        try:
            df = con.bdh(tkr, fields, start, end)
            if df.empty:
                print(" ⚠ empty")
                continue
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(1)
            data[tkr] = df
            print(f" ✓ {len(df)} rows")
        except Exception as e:
            print(f" ✗ {e}")
    return data


# ═══════════════════════════════════════════════
# SHEET WRITERS
# ═══════════════════════════════════════════════

def write_strip_sheet(wb, root, metal_name, strip_data, fields):
    """One sheet per metal: Date | F1 Price | F1 Vol | F1 OI | F2 Price | ..."""
    ws = wb.create_sheet(title=f"{root} Strip")

    contracts = sorted(
        strip_data.keys(),
        key=lambda t: int("".join(filter(str.isdigit, t.split()[0].replace(root, ""))) or "0")
    )
    all_dates = sorted(set().union(*(df.index for df in strip_data.values())))
    if not all_dates:
        ws["A1"] = "No data"
        return

    nf = len(fields)
    total_cols = 1 + len(contracts) * nf

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
    c = ws.cell(row=1, column=1, value=f"{metal_name} — Futures Strip (F1–F{len(contracts)})")
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    # Row 2: Contract headers
    ws.cell(row=2, column=1, value="Date").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    col = 2
    for tkr in contracts:
        num = "".join(filter(str.isdigit, tkr.split()[0].replace(root, "")))
        if nf > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + nf - 1)
        c = ws.cell(row=2, column=col, value=f"F{num}")
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal="center")
        for j in range(nf):
            ws.cell(row=2, column=col + j).fill = HEADER_FILL
        col += nf

    # Row 3: Field sub-headers
    ws.cell(row=3, column=1).fill = SUB_FILL
    col = 2
    for tkr in contracts:
        for f in fields:
            c = ws.cell(row=3, column=col, value=fmt_field(f))
            c.font = SUB_FONT; c.fill = SUB_FILL; c.alignment = Alignment(horizontal="center")
            col += 1

    # Data rows
    for ri, dt in enumerate(all_dates, 4):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = BOLD_FONT
        col = 2
        for tkr in contracts:
            df = strip_data[tkr]
            for f in fields:
                cell = ws.cell(row=ri, column=col)
                if dt in df.index and f in df.columns:
                    v = df.loc[dt, f]
                    if pd.notna(v):
                        cell.value = v
                cell.font = DATA_FONT
                cell.number_format = num_fmt(f)
                col += 1

    # Widths & freeze
    ws.column_dimensions["A"].width = 13
    for c in range(2, min(total_cols + 1, 1000)):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "B4"
    print(f"    ✓ Sheet: {ws.title}")


def write_flat_sheet(wb, sheet_name, ticker_map, data, fields):
    """Write a flat sheet: Date | Ticker1_Field1 | Ticker1_Field2 | ... """
    ws = wb.create_sheet(title=sheet_name[:31])

    ordered_tickers = [t for t in ticker_map if t in data]
    all_dates = sorted(set().union(*(data[t].index for t in ordered_tickers))) if ordered_tickers else []

    if not all_dates:
        ws["A1"] = "No data available"
        return

    nf = len(fields)

    # Row 1: Title
    total_cols = 1 + len(ordered_tickers) * nf
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
    c = ws.cell(row=1, column=1, value=sheet_name)
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = Alignment(horizontal="center")

    # Row 2: Ticker name headers
    ws.cell(row=2, column=1, value="Date").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    col = 2
    for tkr in ordered_tickers:
        label = ticker_map[tkr]
        span = nf
        if span > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + span - 1)
        c = ws.cell(row=2, column=col, value=label)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal="center")
        for j in range(span):
            ws.cell(row=2, column=col + j).fill = HEADER_FILL
        col += span

    # Row 3: Field sub-headers
    ws.cell(row=3, column=1).fill = SUB_FILL
    col = 2
    for tkr in ordered_tickers:
        for f in fields:
            c = ws.cell(row=3, column=col, value=fmt_field(f))
            c.font = SUB_FONT; c.fill = SUB_FILL; c.alignment = Alignment(horizontal="center")
            col += 1

    # Data
    for ri, dt in enumerate(all_dates, 4):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = BOLD_FONT
        col = 2
        for tkr in ordered_tickers:
            df = data[tkr]
            for f in fields:
                cell = ws.cell(row=ri, column=col)
                if dt in df.index and f in df.columns:
                    v = df.loc[dt, f]
                    if pd.notna(v):
                        cell.value = v
                cell.font = DATA_FONT
                cell.number_format = num_fmt(f)
                col += 1

    ws.column_dimensions["A"].width = 13
    for c in range(2, min(total_cols + 1, 500)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B4"
    print(f"    ✓ Sheet: {ws.title}")


def write_readme(wb, config):
    ws = wb.active
    ws.title = "README"
    info = [
        ("Metals Risk Premia — Bloomberg Data", ""),
        ("", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Date Range", f"{config['start_date']} → {config['end_date']}"),
        ("", ""),
        ("SHEET GUIDE", ""),
        ("LME [Metal]", "★ Official LME instruments: Cash, 3M (benchmark), 15M, 27M, Cash-3M spread"),
        ("[ROOT] Strip", "Generic monthly strip (F1–Fn) — Bloomberg synthetic rolls, less liquid than 3M"),
        ("Cash & Carry", "Full term structure for carry signal: Cash, 3M, 15M, 27M + Gold/Silver COMEX"),
        ("Vol Surface", "Implied vol, realized vol (30/60/90D) for VRP construction"),
        ("Fundamentals", "LME inventories, cancelled warrants, DXY, rates, PMI, etc."),
        ("Cross-Metal", "Front-month prices for ratio/spread construction (Cu/Au, Cu/Al, etc.)"),
        ("", ""),
        ("LME MARKET STRUCTURE", ""),
        ("", "LME trades on a daily prompt-date system, NOT monthly contracts like CME."),
        ("", "The 3-Month forward is THE benchmark — highest volume and open interest."),
        ("", "Cash-3M spread is the primary carry/backwardation signal."),
        ("", "Generic strip (LP1, LP2...) are Bloomberg synthetic monthly constructs."),
        ("", "Gold & Silver trade on COMEX (CME) — standard monthly expiry structure."),
        ("", ""),
        ("UNITS", ""),
        ("", "LME base metals: $/metric ton (MT)"),
        ("", "Gold: $/troy oz | Silver: $/troy oz"),
        ("", "COMEX Copper (HG): $/lb"),
    ]
    for r, (a, b) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0D3B66")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


# ═══════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════

def main():
    print("\n" + "=" * 60)
    print("  METALS RISK PREMIA — BLOOMBERG DATA DOWNLOADER")
    print("=" * 60)

    start = CONFIG["start_date"]
    end = CONFIG["end_date"]
    con = connect(CONFIG["host"], CONFIG["port"])
    wb = Workbook()

    # ── 0. LME BENCHMARKS (Cash, 3M, 15M, 27M — actual traded instruments) ──
    if CONFIG["download_lme_benchmarks"]:
        print("\n━━ LME BENCHMARK INSTRUMENTS (Cash / 3M / 15M / 27M) ━━")
        for metal, tkr_map in LME_BENCHMARKS.items():
            print(f"\n  {metal}:")
            tickers = list(tkr_map.keys())
            data = bulk_download(con, tickers, LME_BENCHMARK_FIELDS, start, end, label=metal[:2])
            if data:
                write_flat_sheet(wb, f"LME {metal}", tkr_map, data, LME_BENCHMARK_FIELDS)

    # ── 1. GENERIC FUTURES STRIP (LP1-LP27 synthetic monthly rolls) ──
    if CONFIG["download_futures_strip"]:
        print("\n━━ GENERIC FUTURES STRIP (Monthly Rolls) ━━")
        for root, spec in FUTURES_STRIP.items():
            print(f"\n  {spec['name']} ({root}1–{root}{spec['max_months']}):")
            tickers = [f"{root}{i} Comdty" for i in range(1, spec["max_months"] + 1)]
            data = bulk_download(con, tickers, spec["fields"], start, end, label=root)
            if data:
                write_strip_sheet(wb, root, spec["name"], data, spec["fields"])

    # ── 2. CASH & CARRY ──
    if CONFIG["download_cash_carry"]:
        print("\n━━ CASH & CARRY DATA ━━")
        tickers = list(CASH_CARRY_TICKERS.keys())
        data = bulk_download(con, tickers, CASH_CARRY_FIELDS, start, end, label="CARRY")
        if data:
            write_flat_sheet(wb, "Cash & Carry", CASH_CARRY_TICKERS, data, CASH_CARRY_FIELDS)

    # ── 3. VOL SURFACE ──
    if CONFIG["download_vol_surface"]:
        print("\n━━ VOLATILITY SURFACE DATA ━━")
        tickers = list(VOL_TICKERS.keys())
        data = bulk_download(con, tickers, VOL_FIELDS, start, end, label="VOL")
        if data:
            write_flat_sheet(wb, "Vol Surface", VOL_TICKERS, data, VOL_FIELDS)

    # ── 4. FUNDAMENTALS ──
    if CONFIG["download_fundamentals"]:
        print("\n━━ FUNDAMENTALS & MACRO ━━")
        tickers = list(FUNDAMENTAL_TICKERS.keys())
        data = bulk_download(con, tickers, FUNDAMENTAL_FIELDS, start, end, label="FUND")
        if data:
            write_flat_sheet(wb, "Fundamentals", FUNDAMENTAL_TICKERS, data, FUNDAMENTAL_FIELDS)

    # ── 5. CROSS-METAL ──
    if CONFIG["download_cross_metal"]:
        print("\n━━ CROSS-METAL RATIO DATA ━━")
        tickers = list(CROSS_METAL_TICKERS.keys())
        data = bulk_download(con, tickers, CROSS_METAL_FIELDS, start, end, label="XMETAL")
        if data:
            write_flat_sheet(wb, "Cross-Metal", CROSS_METAL_TICKERS, data, CROSS_METAL_FIELDS)

    # ── SAVE ──
    write_readme(wb, CONFIG)
    # Move README to first position
    wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)

    output = "metals_risk_premia_data.xlsx"
    wb.save(output)

    print("\n" + "=" * 60)
    print(f"  ✓ SAVED: {os.path.abspath(output)}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print("=" * 60 + "\n")

    try:
        con.stop()
    except Exception:
        pass


if __name__ == "__main__":
    main()
