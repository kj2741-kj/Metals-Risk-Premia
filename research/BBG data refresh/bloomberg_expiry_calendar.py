"""
Bloomberg Expiry Calendar Downloader - Metals Risk Premia
=========================================================
Fetches static reference data (expiry/delivery dates) for all futures contracts
across LME metals, precious metals, energy, and NGL products.

Fields retrieved per contract (via Bloomberg ref() - no date range needed):
  LAST_TRADEABLE_DT  : Last day the contract can be traded
  FUT_NOTICE_FIRST   : First notice date (physical delivery warning)
  FUT_DLV_DT_LAST    : Last delivery date
  FUT_CONT_SIZE      : Contract lot size
  CRNCY              : Settlement currency
  LONG_COMP_NAME     : Full Bloomberg contract name

Output: Expiry_Calendars_Updated.xlsx
  - One sheet per product category
  - All active contracts (F1-Fn) listed with expiry details

Run with Bloomberg Terminal open:
    python bloomberg_expiry_calendar.py

Requirements:
    pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

# =============================================================================
# CONFIG
# =============================================================================

HOST = "localhost"
PORT = 8194

TIMEOUT_PER_CONTRACT = 12   # seconds hard cap per contract ref() call
EARLY_EXIT_EMPTY     = 3    # stop strip after N consecutive no-data contracts

OUTPUT_FILE = "Expiry_Calendars_Updated.xlsx"

# Reference fields to pull (static, no date range)
REF_FIELDS = [
    "LAST_TRADEABLE_DT",   # last trading day
    "FUT_NOTICE_FIRST",    # first notice date (for physical contracts)
    "FUT_DLV_DT_LAST",     # last delivery date
    "FUT_CONT_SIZE",       # contract lot size
    "CRNCY",               # settlement currency
    "LONG_COMP_NAME",      # full contract description
]

# =============================================================================
# PRODUCT UNIVERSE (same as bloomberg_full_download.py)
# =============================================================================

ALL_PRODUCTS = {
    # ---- LME Base Metals ----
    "Copper LME":     {"root": "LP",  "max_contracts": 27, "category": "LME Base Metals"},
    "Aluminium LME":  {"root": "LA",  "max_contracts": 27, "category": "LME Base Metals"},
    "Zinc LME":       {"root": "LX",  "max_contracts": 27, "category": "LME Base Metals"},
    "Nickel LME":     {"root": "LN",  "max_contracts": 27, "category": "LME Base Metals"},
    "Lead LME":       {"root": "LL",  "max_contracts": 27, "category": "LME Base Metals"},
    "Tin LME":        {"root": "LT",  "max_contracts": 15, "category": "LME Base Metals"},
    # ---- Precious Metals ----
    "Gold COMEX":     {"root": "GC",  "max_contracts": 20, "category": "Precious Metals"},
    "Silver COMEX":   {"root": "SI",  "max_contracts": 17, "category": "Precious Metals"},
    "Copper CME HG":  {"root": "HG",  "max_contracts": 27, "category": "Precious Metals"},
    "Platinum NYMEX": {"root": "PL",  "max_contracts": 13, "category": "Precious Metals"},
    "Palladium NYMEX":{"root": "PA",  "max_contracts": 12, "category": "Precious Metals"},
    # ---- Energy Core ----
    "WTI Crude":      {"root": "CL",  "max_contracts": 27, "category": "Energy Core"},
    "Brent Crude":    {"root": "CO",  "max_contracts": 27, "category": "Energy Core"},
    "RBOB Gasoline":  {"root": "XB",  "max_contracts": 27, "category": "Energy Core"},
    "Heating Oil":    {"root": "HO",  "max_contracts": 27, "category": "Energy Core"},
    "Nat Gas HH":     {"root": "NG",  "max_contracts": 27, "category": "Energy Core"},
    "Singapore Gasoil":{"root":"QS",  "max_contracts": 27, "category": "Energy Core"},
    # ---- Energy Extended ----
    "ICE Gasoil":     {"root": "GO",  "max_contracts": 27, "category": "Energy Extended"},
    "Fuel Oil Barges":{"root": "FO",  "max_contracts": 12, "category": "Energy Extended"},
    "Singapore Jet":  {"root": "SJ",  "max_contracts": 12, "category": "Energy Extended"},
    "Naphtha CIF NWE":{"root": "NFY", "max_contracts": 12, "category": "Energy Extended"},
    # ---- NGL Products ----
    "Propane Argus C3":   {"root": "CAP", "max_contracts": 27, "category": "NGL Products"},
    "Butane Argus C4":    {"root": "BAP", "max_contracts": 27, "category": "NGL Products"},
    "Ethane Argus":       {"root": "DAE", "max_contracts": 27, "category": "NGL Products"},
    "Isobutane Argus":    {"root": "IBD", "max_contracts": 27, "category": "NGL Products"},
    "Propane C3 CIF ARA": {"root": "PCW", "max_contracts": 27, "category": "NGL Products"},
    "Propylene Argus":    {"root": "PGP", "max_contracts": 27, "category": "NGL Products"},
}

# =============================================================================
# STYLES
# =============================================================================

TITLE_FILL = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL   = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT   = Font(name="Arial", bold=True, color="1F4E79", size=9)
DATA_FONT  = Font(name="Arial", size=9)
DATE_FONT  = Font(name="Arial", bold=True, size=9)
ALT_FILL   = PatternFill("solid", fgColor="F5F9FC")

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# SAFE REF DATA DOWNLOAD
# =============================================================================

def safe_ref(con, ticker, fields):
    """
    Fetch Bloomberg reference (static) data for a single ticker.
    Returns dict {field: value} or None on timeout/error.
    Never blocks longer than TIMEOUT_PER_CONTRACT seconds.
    """
    def _call():
        return con.ref(ticker, fields)

    with ThreadPoolExecutor(max_workers=1) as ex:
        fut = ex.submit(_call)
        try:
            df = fut.result(timeout=TIMEOUT_PER_CONTRACT)
            if df is None or df.empty:
                return None, "no data"
            # pdblp ref() returns DataFrame with columns: security, field, value
            result = {}
            for _, row in df.iterrows():
                fld = row.get("field", "")
                val = row.get("value", None)
                if fld:
                    result[fld] = val
            if not result:
                return None, "empty response"
            return result, "OK"
        except FuturesTimeoutError:
            return None, "TIMEOUT (%ds)" % TIMEOUT_PER_CONTRACT
        except Exception as e:
            return None, "ERROR: %s" % str(e)[:80]


def get_strip_expiry(con, root, max_contracts, product_name):
    """
    Fetch expiry reference data for F1-Fn of a single root.
    Returns list of dicts, one per contract that returned data.
    Applies early-exit after EARLY_EXIT_EMPTY consecutive failures.
    """
    rows = []
    consecutive_empty = 0
    print("  %-22s (root=%s, max F%d)" % (product_name, root, max_contracts))

    for i in range(1, max_contracts + 1):
        tkr = "%s%d Comdty" % (root, i)
        ref_data, status = safe_ref(con, tkr, REF_FIELDS)

        if ref_data is not None:
            row = {
                "Product":           product_name,
                "Root":              root,
                "Contract":          "F%d" % i,
                "Ticker":            tkr,
                "Last Trade Date":   ref_data.get("LAST_TRADEABLE_DT", ""),
                "First Notice Date": ref_data.get("FUT_NOTICE_FIRST", ""),
                "Last Delivery Date":ref_data.get("FUT_DLV_DT_LAST", ""),
                "Contract Size":     ref_data.get("FUT_CONT_SIZE", ""),
                "Currency":          ref_data.get("CRNCY", ""),
                "Name":              ref_data.get("LONG_COMP_NAME", ""),
            }
            rows.append(row)
            consecutive_empty = 0
            ltd = row["Last Trade Date"]
            print("    %-20s  last trade: %s  (%s)" % (tkr, ltd, row["Currency"]))
        else:
            consecutive_empty += 1
            print("    %-20s  %s" % (tkr, status))
            if consecutive_empty >= EARLY_EXIT_EMPTY:
                remaining = max_contracts - i
                if remaining > 0:
                    print("    => %d consecutive empties at F%d — stopping (skipping F%d-F%d)"
                          % (EARLY_EXIT_EMPTY, i, i + 1, max_contracts))
                break

    print("  => %d contracts with expiry data\n" % len(rows))
    return rows

# =============================================================================
# EXCEL WRITER
# =============================================================================

COLUMNS = [
    "Product", "Root", "Contract", "Ticker",
    "Last Trade Date", "First Notice Date", "Last Delivery Date",
    "Contract Size", "Currency", "Name",
]

COL_WIDTHS = [22, 7, 10, 18, 17, 18, 19, 14, 10, 40]

DATE_COLS = {"Last Trade Date", "First Notice Date", "Last Delivery Date"}


def write_category_sheet(wb, category_name, rows):
    """Write one sheet for a product category."""
    ws = wb.create_sheet(title=category_name[:31])

    if not rows:
        ws["A1"] = "No data returned for %s" % category_name
        return

    n_cols = len(COLUMNS)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="%s -- Expiry Calendar (as of %s)"
                % (category_name, datetime.now().strftime("%Y-%m-%d")))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Header row
    for j, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(row=2, column=j, value=col_name)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    # Data rows — alternating fill for readability
    prev_product = None
    fill_flag = False
    for ri, row_dict in enumerate(rows, 3):
        product = row_dict.get("Product", "")
        if product != prev_product:
            fill_flag = not fill_flag
            prev_product = product
        row_fill = SUB_FILL if fill_flag else ALT_FILL

        for j, col_name in enumerate(COLUMNS, 1):
            val = row_dict.get(col_name, "")
            cell = ws.cell(row=ri, column=j)

            # Parse date objects returned by Bloomberg
            if col_name in DATE_COLS and val not in ("", None):
                try:
                    if hasattr(val, "strftime"):
                        cell.value = val
                        cell.number_format = "YYYY-MM-DD"
                    else:
                        dt = pd.to_datetime(str(val), errors="coerce")
                        if pd.notna(dt):
                            cell.value = dt.to_pydatetime()
                            cell.number_format = "YYYY-MM-DD"
                        else:
                            cell.value = str(val)
                except Exception:
                    cell.value = str(val)
            else:
                cell.value = val if val not in (None, "") else ""

            cell.font = DATA_FONT
            cell.fill = row_fill
            if col_name in ("Product", "Root"):
                cell.font = Font(name="Arial", size=9, bold=True)

    # Column widths + freeze
    for j, width in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = "A3"
    print("    Sheet '%s': %d contract rows" % (ws.title, len(rows)))


def write_summary_sheet(wb, all_rows):
    """First sheet: summary count per product."""
    ws = wb.create_sheet(title="Summary")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    c = ws.cell(row=1, column=1, value="Expiry Calendar Summary -- Generated %s"
                % datetime.now().strftime("%Y-%m-%d %H:%M"))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for j, h in enumerate(["Product", "Root", "Contracts", "Earliest Expiry"], 1):
        c = ws.cell(row=2, column=j, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    # Group by product
    from collections import defaultdict
    grouped = defaultdict(list)
    for row in all_rows:
        grouped[row["Product"]].append(row)

    for ri, (product, prows) in enumerate(grouped.items(), 3):
        root = prows[0]["Root"]
        count = len(prows)
        # Earliest last trade date
        dates = []
        for r in prows:
            d = r.get("Last Trade Date", "")
            if d and d not in ("", None):
                try:
                    dates.append(pd.to_datetime(str(d), errors="coerce"))
                except Exception:
                    pass
        earliest = min((d for d in dates if pd.notna(d)), default=None)

        ws.cell(row=ri, column=1, value=product).font = DATA_FONT
        ws.cell(row=ri, column=2, value=root).font   = DATA_FONT
        ws.cell(row=ri, column=3, value=count).font  = DATA_FONT
        cell = ws.cell(row=ri, column=4)
        if earliest is not None:
            cell.value = earliest.to_pydatetime()
            cell.number_format = "YYYY-MM-DD"
        cell.font = DATA_FONT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.freeze_panes = "A3"

# =============================================================================
# MAIN
# =============================================================================

def main():
    print("\n" + "=" * 65)
    print("  METALS RISK PREMIA - BLOOMBERG EXPIRY CALENDAR DOWNLOADER")
    print("  Fields: %s" % ", ".join(REF_FIELDS))
    print("  Timeout: %ds/contract, exit after %d consecutive empties"
          % (TIMEOUT_PER_CONTRACT, EARLY_EXIT_EMPTY))
    print("=" * 65)

    con = connect()

    # Group products by category for sheet organisation
    from collections import defaultdict
    category_rows = defaultdict(list)
    all_rows = []

    for product_name, cfg in ALL_PRODUCTS.items():
        root     = cfg["root"]
        max_c    = cfg["max_contracts"]
        category = cfg["category"]
        rows = get_strip_expiry(con, root, max_c, product_name)
        category_rows[category].extend(rows)
        all_rows.extend(rows)

    # Disconnect
    try:
        con.stop()
    except Exception:
        pass

    # Build workbook
    print("\n" + "=" * 65)
    print("  WRITING EXCEL WORKBOOK")
    print("=" * 65)

    wb = Workbook()

    # Summary sheet first
    write_summary_sheet(wb, all_rows)

    # One sheet per category
    for category in ["LME Base Metals", "Precious Metals",
                     "Energy Core", "Energy Extended", "NGL Products"]:
        rows = category_rows.get(category, [])
        write_category_sheet(wb, category, rows)

    # Move Summary to front
    wb.move_sheet("Summary", offset=-len(wb.sheetnames) + 1)

    # Save
    folder = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(folder, OUTPUT_FILE)
    wb.save(out_path)

    print("\n" + "=" * 65)
    print("  EXPIRY CALENDAR COMPLETE")
    print("  Saved: %s" % out_path)
    print("  Sheets: %s" % ", ".join(wb.sheetnames))
    print("  Total contracts with data: %d" % len(all_rows))
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
