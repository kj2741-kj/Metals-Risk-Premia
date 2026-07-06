"""
Bloomberg Metals Data Downloader — PRE-CONFIGURED
===================================================
Pre-filled for the Metals Risk Premia project.
Just run it with Bloomberg Terminal open.

    python bloomberg_metals_download.py

Requirements:
    pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

# ═══════════════════════════════════════════════
# CONFIG — EDIT THESE AS NEEDED
# ═══════════════════════════════════════════════

CONFIG = {
    # ── Date Range ──
    "start_date": "20050101",
    "end_date": datetime.today().strftime("%Y%m%d"),

    # ── Bloomberg Connection ──
    "host": "localhost",
    "port": 8194,

    # ── Downloads to run ──
    # Set any to False to skip
    "download_lme_benchmarks": True,   # ★ LME official: Cash, Tom, 3M, 15M, 27M
    "download_cash_carry": True,       # Carry: Cash, Tom, 3M + CME equivalents (Price/Vol/OI)
    "download_fundamentals": True,     # Inventories, macro, cancelled warrants
}

# ═══════════════════════════════════════════════
# TICKER DEFINITIONS
# ═══════════════════════════════════════════════

# ── 1. LME OFFICIAL BENCHMARK INSTRUMENTS ──
# CORRECTED TICKER CONVENTION (verified on Bloomberg Terminal):
#
#   Cash (Spot)    →  LM[XX]DY Comdty        (e.g. LMCADY)
#   Tom Next (T+1) →  LM[XX]DTTT Comdty      (e.g. LMCADTTT)
#   3-Month (3M)   →  LM[XX]DS03 Comdty      (e.g. LMCADS03)  ★ BENCHMARK
#   Cash-3M Spread →  LM[XX]Y03 Comdty       (e.g. LMCAY03)
#
# Where [XX] = metal code: CA=Copper, AH=Aluminium, ZS=Zinc,
#                           NI=Nickel, PB=Lead, SN=Tin

LME_BENCHMARKS = {
    "Copper": {
        "LMCADY Comdty":    "Cu Cash (Spot)",
        "LMCADTTT Comdty":  "Cu Tom Next (T+1)",
        "LMCADS03 Comdty":  "Cu 3M Forward ★",
        "LMCAY03 Comdty":   "Cu Cash-3M Spread",
    },
    "Aluminium": {
        "LMAHDY Comdty":    "Al Cash (Spot)",
        "LMAHDTTT Comdty":  "Al Tom Next (T+1)",
        "LMAHDS03 Comdty":  "Al 3M Forward ★",
        "LMAHY03 Comdty":   "Al Cash-3M Spread",
    },
    "Zinc": {
        "LMZSDY Comdty":    "Zn Cash (Spot)",
        "LMZSDTTT Comdty":  "Zn Tom Next (T+1)",
        "LMZSDS03 Comdty":  "Zn 3M Forward ★",
        "LMZSY03 Comdty":   "Zn Cash-3M Spread",
    },
    "Nickel": {
        "LMNIDY Comdty":    "Ni Cash (Spot)",
        "LMNIDTTT Comdty":  "Ni Tom Next (T+1)",
        "LMNIDS03 Comdty":  "Ni 3M Forward ★",
        "LMNIY03 Comdty":   "Ni Cash-3M Spread",
    },
    "Lead": {
        "LMPBDY Comdty":    "Pb Cash (Spot)",
        "LMPBDTTT Comdty":  "Pb Tom Next (T+1)",
        "LMPBDS03 Comdty":  "Pb 3M Forward ★",
        "LMPBY03 Comdty":   "Pb Cash-3M Spread",
    },
    "Tin": {
        "LMSNDY Comdty":    "Sn Cash (Spot)",
        "LMSNDTTT Comdty":  "Sn Tom Next (T+1)",
        "LMSNDS03 Comdty":  "Sn 3M Forward ★",
        "LMSNY03 Comdty":   "Sn Cash-3M Spread",
    },
}
LME_BENCHMARK_FIELDS = ["PX_LAST", "PX_VOLUME", "OPEN_INT"]

# ── 2. CASH & CARRY / TERM STRUCTURE ──
# Corrected LME tickers + CME equivalents
# All with Price, Volume, Open Interest
CASH_CARRY_TICKERS = {
    # ── LME Copper ──
    "LMCADY Comdty":    "Cu Cash",
    "LMCADTTT Comdty":  "Cu Tom",
    "LMCADS03 Comdty":  "Cu 3M",
    "LMCAY03 Comdty":   "Cu Cash-3M Spread",
    # ── LME Aluminium ──
    "LMAHDY Comdty":    "Al Cash",
    "LMAHDTTT Comdty":  "Al Tom",
    "LMAHDS03 Comdty":  "Al 3M",
    "LMAHY03 Comdty":   "Al Cash-3M Spread",
    # ── LME Zinc ──
    "LMZSDY Comdty":    "Zn Cash",
    "LMZSDTTT Comdty":  "Zn Tom",
    "LMZSDS03 Comdty":  "Zn 3M",
    "LMZSY03 Comdty":   "Zn Cash-3M Spread",
    # ── LME Nickel ──
    "LMNIDY Comdty":    "Ni Cash",
    "LMNIDTTT Comdty":  "Ni Tom",
    "LMNIDS03 Comdty":  "Ni 3M",
    "LMNIY03 Comdty":   "Ni Cash-3M Spread",
    # ── LME Lead ──
    "LMPBDY Comdty":    "Pb Cash",
    "LMPBDTTT Comdty":  "Pb Tom",
    "LMPBDS03 Comdty":  "Pb 3M",
    "LMPBY03 Comdty":   "Pb Cash-3M Spread",
    # ── LME Tin ──
    "LMSNDY Comdty":    "Sn Cash",
    "LMSNDTTT Comdty":  "Sn Tom",
    "LMSNDS03 Comdty":  "Sn 3M",
    "LMSNY03 Comdty":   "Sn Cash-3M Spread",
    # ── COMEX Copper (CME) ──
    "HG1 Comdty":       "COMEX Cu F1",
    "HG2 Comdty":       "COMEX Cu F2",
    "HG3 Comdty":       "COMEX Cu F3 (~3M eq)",
    # ── Gold (COMEX) ──
    "GOLDLNPM Index":   "Gold London PM Fix",
    "XAU Curncy":       "Gold Spot (XAU)",
    "GC1 Comdty":       "Gold F1",
    "GC2 Comdty":       "Gold F2",
    "GC4 Comdty":       "Gold F4 (~3M eq)",
    "GC6 Comdty":       "Gold F6 (~6M eq)",
    "GC12 Comdty":      "Gold F12 (~1Y eq)",
    # ── Silver (COMEX) ──
    "SI1 Comdty":       "Silver F1",
    "SI2 Comdty":       "Silver F2",
    "SI4 Comdty":       "Silver F4 (~3M eq)",
    # ── Platinum (NYMEX) ──
    "PL1 Comdty":       "Platinum F1",
    "PL2 Comdty":       "Platinum F2",
    "PL3 Comdty":       "Platinum F3 (~3M eq)",
    # ── Palladium (NYMEX) ──
    "PA1 Comdty":       "Palladium F1",
    "PA2 Comdty":       "Palladium F2",
    "PA3 Comdty":       "Palladium F3 (~3M eq)",
    # ── Brent Crude (ICE) ──
    "CO1 Comdty":       "Brent F1",
    "CO2 Comdty":       "Brent F2",
    "CO4 Comdty":       "Brent F4 (~3M eq)",
}
CASH_CARRY_FIELDS = ["PX_LAST", "PX_VOLUME", "OPEN_INT"]

# ── 3. FUNDAMENTALS / MACRO ──
FUNDAMENTAL_TICKERS = {
    # LME Inventories
    "LCSNSTOT Index":   "LME Copper Inventory (MT)",
    "LASSNTOT Index":   "LME Aluminium Inventory (MT)",
    "LMZSNTOT Index":   "LME Zinc Inventory (MT)",
    "LMNITON Index":    "LME Nickel Inventory (MT)",
    "LMPBTOT Index":    "LME Lead Inventory (MT)",
    "LMSNTOT Index":    "LME Tin Inventory (MT)",
    # Cancelled Warrants (% of total inventory)
    "LCSNCANC Index":   "Cu Cancelled Warrants (%)",
    "LASCANWP Index":   "Al Cancelled Warrants (%)",
    "LMZSCANW Index":   "Zn Cancelled Warrants (%)",
    "LMNICANW Index":   "Ni Cancelled Warrants (%)",
    # Macro
    "DXY Curncy":       "US Dollar Index",
    "GTII10 Govt":      "US 10Y TIPS Yield",
    "USGG10YR Index":   "US 10Y Treasury Yield",
    "CPMINDX Index":    "China Mfg PMI",
    "MXWO Index":       "MSCI World Index",
    "CRY Index":        "CRB Commodity Index",
    # Energy cross-refs
    "CL1 Comdty":       "WTI Crude F1",
    "CO1 Comdty":       "Brent Crude F1",
}
FUNDAMENTAL_FIELDS = ["PX_LAST"]


# ═══════════════════════════════════════════════
# STYLING
# ═══════════════════════════════════════════════

TITLE_FILL = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
SUB_FILL = PatternFill("solid", fgColor="D6E4F0")
SUB_FONT = Font(name="Arial", bold=True, color="1F4E79", size=9)
DATA_FONT = Font(name="Arial", size=9)
BOLD_FONT = Font(name="Arial", size=9, bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="C0C0C0"))

FIELD_DISPLAY = {
    "PX_LAST": "Price", "PX_VOLUME": "Volume", "PX_OPEN": "Open",
    "PX_HIGH": "High", "PX_LOW": "Low", "OPEN_INT": "Open Int",
}

def fmt_field(f):
    return FIELD_DISPLAY.get(f, f.replace("_", " ").title())

def num_fmt(f):
    fu = f.upper()
    if "VOLUME" in fu or "OPEN_INT" in fu:
        return "#,##0"
    return "#,##0.00"


# ═══════════════════════════════════════════════
# DOWNLOAD ENGINE
# ═══════════════════════════════════════════════

def connect(host, port):
    print(f"\n  Connecting to Bloomberg ({host}:{port}) ...")
    con = pdblp.BCon(debug=False, host=host, port=port, timeout=30000)
    con.start()
    print("  ✓ Connected.\n")
    return con


def bulk_download(con, tickers, fields, start, end, label=""):
    """
    Download multiple tickers, return {ticker: DataFrame}.
    
    ROBUST FALLBACK: If requesting multiple fields returns empty
    (common on LME — some tickers don't support PX_VOLUME or OPEN_INT),
    retries field-by-field and merges whatever is available.
    This ensures PX_LAST is captured even when other fields fail.
    """
    data = {}
    total = len(tickers)
    for i, tkr in enumerate(tickers, 1):
        tag = f"[{label}]" if label else ""
        print(f"    {tag} ({i}/{total}) {tkr} ...", end="", flush=True)
        try:
            # First attempt: all fields at once (fastest)
            df = con.bdh(tkr, fields, start, end)
            if isinstance(df.columns, pd.MultiIndex):
                df.columns = df.columns.get_level_values(1)

            if not df.empty and len(df.columns) > 0:
                data[tkr] = df
                print(f" ✓ {len(df)} rows × {len(df.columns)} fields")
                continue

            # Fallback: try each field individually and merge
            # This handles LME tickers where PX_VOLUME/OPEN_INT
            # don't exist — requesting them together kills the whole call
            if len(fields) > 1:
                print(" ⚠ combined empty, trying field-by-field ...", end="", flush=True)
                parts = []
                found_fields = []
                for fld in fields:
                    try:
                        df_single = con.bdh(tkr, fld, start, end)
                        if isinstance(df_single.columns, pd.MultiIndex):
                            df_single.columns = df_single.columns.get_level_values(1)
                        if not df_single.empty:
                            parts.append(df_single)
                            found_fields.append(fld)
                    except Exception:
                        pass

                if parts:
                    merged = pd.concat(parts, axis=1)
                    # Remove any duplicate columns
                    merged = merged.loc[:, ~merged.columns.duplicated()]
                    data[tkr] = merged
                    print(f" ✓ {len(merged)} rows ({', '.join(found_fields)})")
                else:
                    print(" ⚠ no fields returned")
            else:
                print(" ⚠ empty")

        except Exception as e:
            print(f" ✗ {e}")
            # Even on exception, try PX_LAST alone as last resort
            if "PX_LAST" in fields and len(fields) > 1:
                try:
                    df_px = con.bdh(tkr, "PX_LAST", start, end)
                    if isinstance(df_px.columns, pd.MultiIndex):
                        df_px.columns = df_px.columns.get_level_values(1)
                    if not df_px.empty:
                        data[tkr] = df_px
                        print(f"      ↳ recovered PX_LAST: {len(df_px)} rows")
                except Exception:
                    pass
    return data


# ═══════════════════════════════════════════════
# SHEET WRITERS
# ═══════════════════════════════════════════════

def write_flat_sheet(wb, sheet_name, ticker_map, data, fields):
    """
    Write a flat sheet with grouped columns:
    Date | Ticker1 Price | Ticker1 Volume | Ticker1 OI | Ticker2 Price | ...
    """
    ws = wb.create_sheet(title=sheet_name[:31])

    ordered_tickers = [t for t in ticker_map if t in data]
    all_dates = sorted(set().union(*(data[t].index for t in ordered_tickers))) if ordered_tickers else []

    if not all_dates:
        ws["A1"] = "No data available"
        return

    nf = len(fields)
    total_cols = 1 + len(ordered_tickers) * nf

    # Row 1: Title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=min(total_cols, 50))
    c = ws.cell(row=1, column=1, value=sheet_name)
    c.font = TITLE_FONT; c.fill = TITLE_FILL; c.alignment = Alignment(horizontal="center")

    # Row 2: Ticker group headers
    ws.cell(row=2, column=1, value="Date").font = HEADER_FONT
    ws.cell(row=2, column=1).fill = HEADER_FILL
    col = 2
    for tkr in ordered_tickers:
        label = ticker_map[tkr]
        if nf > 1:
            ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + nf - 1)
        c = ws.cell(row=2, column=col, value=label)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal="center")
        for j in range(nf):
            ws.cell(row=2, column=col + j).fill = HEADER_FILL
        col += nf

    # Row 3: Field sub-headers (Price | Volume | Open Int)
    ws.cell(row=3, column=1).fill = SUB_FILL
    col = 2
    for tkr in ordered_tickers:
        for f in fields:
            c = ws.cell(row=3, column=col, value=fmt_field(f))
            c.font = SUB_FONT; c.fill = SUB_FILL; c.alignment = Alignment(horizontal="center")
            col += 1

    # Data rows
    for ri, dt in enumerate(all_dates, 4):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = BOLD_FONT
        col = 2
        for tkr in ordered_tickers:
            df = data[tkr]
            for f in fields:
                cell = ws.cell(row=ri, column=col)
                if dt in df.index and f in df.columns:
                    v = df.loc[dt, f]
                    if pd.notna(v):
                        cell.value = v
                cell.font = DATA_FONT
                cell.number_format = num_fmt(f)
                col += 1

    # Column widths & freeze
    ws.column_dimensions["A"].width = 13
    for c in range(2, min(total_cols + 1, 500)):
        ws.column_dimensions[get_column_letter(c)].width = 14
    ws.freeze_panes = "B4"
    print(f"    ✓ Sheet: {ws.title}")


def write_readme(wb, config):
    ws = wb.active
    ws.title = "README"
    info = [
        ("Metals Risk Premia — Bloomberg Data", ""),
        ("", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Date Range", f"{config['start_date']} → {config['end_date']}"),
        ("", ""),
        ("SHEET GUIDE", ""),
        ("LME [Metal]", "★ Official LME instruments: Cash, Tom, 3M (benchmark), 15M, 27M — Price/Volume/OI"),
        ("Cash & Carry", "Full term structure: LME Cash/Tom/3M + CME equivalents — Price/Volume/OI"),
        ("Fundamentals", "LME inventories, cancelled warrants, DXY, rates, PMI, macro"),
        ("", ""),
        ("LME TICKER CONVENTION (verified)", ""),
        ("", "Cash (Spot):     LM[XX]DY Comdty      e.g. LMCADY Comdty"),
        ("", "Tom Next (T+1):  LM[XX]DTTT Comdty    e.g. LMCADTTT Comdty"),
        ("", "3M Forward:      LM[XX]DS03 Comdty    e.g. LMCADS03 Comdty  ★ benchmark"),
        ("", "Cash-3M Spread:  LM[XX]Y03 Comdty     e.g. LMCAY03 Comdty"),
        ("", "Metal codes: CA=Copper, AH=Aluminium, ZS=Zinc, NI=Nickel, PB=Lead, SN=Tin"),
        ("", ""),
        ("FIELDS ON ALL SHEETS", ""),
        ("Price", "PX_LAST — daily settlement / closing price"),
        ("Volume", "PX_VOLUME — daily traded volume (contracts)"),
        ("Open Int", "OPEN_INT — open interest (outstanding contracts)"),
        ("", ""),
        ("UNITS", ""),
        ("", "LME base metals (Cu, Al, Zn, Ni, Pb, Sn): $/metric ton"),
        ("", "Gold, Platinum, Palladium: $/troy oz"),
        ("", "Silver: ¢/troy oz"),
        ("", "COMEX Copper (HG): ¢/lb"),
        ("", "Brent Crude: $/bbl"),
    ]
    for r, (a, b) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=14, color="0D3B66")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70


# ═══════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════

def main():
    print("\n" + "=" * 60)
    print("  METALS RISK PREMIA — BLOOMBERG DATA DOWNLOADER")
    print("=" * 60)

    start = CONFIG["start_date"]
    end = CONFIG["end_date"]
    con = connect(CONFIG["host"], CONFIG["port"])
    wb = Workbook()

    # ── 1. LME BENCHMARKS (Cash, Tom, 3M, 15M, 27M) ──
    if CONFIG["download_lme_benchmarks"]:
        print("\n━━ LME BENCHMARK INSTRUMENTS (Cash / Tom / 3M / 15M / 27M) ━━")
        for metal, tkr_map in LME_BENCHMARKS.items():
            print(f"\n  {metal}:")
            tickers = list(tkr_map.keys())
            data = bulk_download(con, tickers, LME_BENCHMARK_FIELDS, start, end, label=metal[:2])
            if data:
                write_flat_sheet(wb, f"LME {metal}", tkr_map, data, LME_BENCHMARK_FIELDS)

    # ── 2. CASH & CARRY (Price / Volume / Open Interest) ──
    if CONFIG["download_cash_carry"]:
        print("\n━━ CASH & CARRY DATA (Price / Volume / OI) ━━")
        tickers = list(CASH_CARRY_TICKERS.keys())
        data = bulk_download(con, tickers, CASH_CARRY_FIELDS, start, end, label="CARRY")
        if data:
            write_flat_sheet(wb, "Cash & Carry", CASH_CARRY_TICKERS, data, CASH_CARRY_FIELDS)

    # ── 3. FUNDAMENTALS ──
    if CONFIG["download_fundamentals"]:
        print("\n━━ FUNDAMENTALS & MACRO ━━")
        tickers = list(FUNDAMENTAL_TICKERS.keys())
        data = bulk_download(con, tickers, FUNDAMENTAL_FIELDS, start, end, label="FUND")
        if data:
            write_flat_sheet(wb, "Fundamentals", FUNDAMENTAL_TICKERS, data, FUNDAMENTAL_FIELDS)

    # ── SAVE ──
    write_readme(wb, CONFIG)
    wb.move_sheet("README", offset=-len(wb.sheetnames) + 1)

    output = "metals_risk_premia_data.xlsx"
    wb.save(output)

    print("\n" + "=" * 60)
    print(f"  ✓ SAVED: {os.path.abspath(output)}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print("=" * 60 + "\n")

    try:
        con.stop()
    except Exception:
        pass


if __name__ == "__main__":
    main()
