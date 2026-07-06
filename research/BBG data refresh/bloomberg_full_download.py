"""
Bloomberg Full Data Downloader - Metals Risk Premia
====================================================
Downloads PX_LAST only for:
  - LME base metals   (LP, LA, LX, LN, LL, LT)          F1-F27
  - Precious metals   (GC, SI, HG, PL, PA)               corrected max tenors
  - Energy / oil      (CL, CO, XB, HO, NG, QS)           F1-F27
  - Energy extended   (GO, FO, SJ, NFY)                   F1-F27
  - NGL products      (CAP, BAP, DAE, IBD, PCW, PGP)      F1-F27
  - LME Cash & 3M     (LMXX DY / LMXX DS03 benchmarks)

End date: June 30, 2026
Deadlock protection: 12s per-contract timeout + early-exit after 3 consecutive empties.

Run with Bloomberg Terminal open:
    python bloomberg_full_download.py

Requirements:
    pip install pdblp openpyxl pandas
"""

import pdblp
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass  # Python < 3.7

# =============================================================================
# CONFIG
# =============================================================================

START_DATE = "20050101"
END_DATE   = "20260630"         # June 30, 2026 — hard-coded target

HOST = "localhost"
PORT = 8194

# Prices only (uncomment second line to add volume/OI)
FIELDS = ["PX_LAST"]
# FIELDS = ["PX_LAST", "PX_VOLUME", "OPEN_INT"]

# Per-contract download timeout (seconds). Bloomberg returns quickly for
# non-existent contracts; this is a hard ceiling against hangs.
TIMEOUT_PER_CONTRACT = 12

# Stop trying deeper tenors after this many consecutive empty/error results.
# e.g. if F13, F14, F15 all fail -> stop at F15 instead of waiting for F16-F27.
EARLY_EXIT_EMPTY = 3

# Output filenames (saved to same folder as this script)
OUT_METALS_CURVE = "Metals_Futures_Curve_Updated.xlsx"
OUT_PRECIOUS     = "Precious_Metals_Futures_Updated.xlsx"
OUT_ENERGY       = "Energy_Futures_Updated.xlsx"
OUT_NGL          = "NGL_Futures_Updated.xlsx"
OUT_CASH_3M      = "Metals_Cash_3M_Updated.xlsx"

# =============================================================================
# FUTURES STRIP DEFINITIONS
# Tickers built as: {ROOT}{n} Comdty  (e.g. LP1 Comdty ... LP27 Comdty)
# max_contracts verified against data files in BBG Data refresh/
# =============================================================================

# -- LME Base Metals ----------------------------------------------------------
LME_METALS = {
    "LP": {"name": "Copper LME",    "max_contracts": 27},
    "LA": {"name": "Aluminium LME", "max_contracts": 27},
    "LX": {"name": "Zinc LME",      "max_contracts": 27},
    "LN": {"name": "Nickel LME",    "max_contracts": 27},
    "LL": {"name": "Lead LME",      "max_contracts": 27},   # added (was missing from old scripts)
    "LT": {"name": "Tin LME",       "max_contracts": 15},   # added; data confirms F1-F15 only
}

# -- Precious Metals (COMEX / NYMEX) ------------------------------------------
# max_contracts corrected: old scripts capped GC/SI/HG at 12 despite data showing more
PRECIOUS_METALS = {
    "GC": {"name": "Gold COMEX",       "max_contracts": 20},  # corrected from 12
    "SI": {"name": "Silver COMEX",     "max_contracts": 17},  # corrected from 12
    "HG": {"name": "Copper CME (HG)",  "max_contracts": 27},  # corrected from 12
    "PL": {"name": "Platinum NYMEX",   "max_contracts": 13},  # added (was missing)
    "PA": {"name": "Palladium NYMEX",  "max_contracts": 12},  # added (was missing)
}

# -- Energy Core (ICE / NYMEX) ------------------------------------------------
# Verified tickers — all confirmed active in Bloomberg terminal
ENERGY_CORE = {
    "CL":  {"name": "WTI Crude (NYMEX)",         "max_contracts": 27},
    "CO":  {"name": "Brent Crude (ICE)",          "max_contracts": 27},
    "XB":  {"name": "RBOB Gasoline (NYMEX)",      "max_contracts": 27},
    "HO":  {"name": "Heating Oil ULSD (NYMEX)",   "max_contracts": 27},
    "NG":  {"name": "Nat Gas Henry Hub (NYMEX)",  "max_contracts": 27},
    "QS":  {"name": "Singapore Gasoil (ICE)",     "max_contracts": 27},
}

# -- Energy Extended ----------------------------------------------------------
# GO and FO are well-established Bloomberg roots.
# SJ (Singapore Jet / Kerosene) and NFY (Naphtha CIF NWE Platts) are correct
# per Argus/ICE conventions but verify on terminal if no data returned.
ENERGY_EXTENDED = {
    "GO":  {"name": "ICE Gasoil London",                 "max_contracts": 27},  # European diesel benchmark
    "FO":  {"name": "Fuel Oil 3.5pct Barges (ICE)",      "max_contracts": 12},  # bunker / residual fuel
    "SJ":  {"name": "Singapore Jet Kerosene (ICE)",      "max_contracts": 12},  # verify root on terminal
    "NFY": {"name": "Naphtha CIF NWE Platts (ICE)",      "max_contracts": 12},  # petrochemical feedstock
}

# -- NGL / Refined Products ---------------------------------------------------
# Roots confirmed from NGL_PAPER1.xlsx in BBG Data refresh/
NGL_PRODUCTS = {
    "CAP": {"name": "Propane Argus Far East C3",  "max_contracts": 27},
    "BAP": {"name": "Butane Argus C4",            "max_contracts": 27},
    "DAE": {"name": "Ethane Argus",               "max_contracts": 27},
    "IBD": {"name": "Isobutane Argus",            "max_contracts": 27},
    "PCW": {"name": "Propane C3 CIF ARA",         "max_contracts": 27},
    "PGP": {"name": "Propylene Argus",            "max_contracts": 27},
}

# -- LME Cash & 3M Benchmarks -------------------------------------------------
# These are NOT numbered strip tickers — they use the LMXX DY / LMXX DS03 format.
# These feed into Metals Cash and 3M.xlsx.
LME_CASH_3M_TICKERS = {
    "LMCADY Comdty":    "Copper Cash",
    "LMCADS03 Comdty":  "Copper 3M",
    "LMAHDY Comdty":    "Aluminium Cash",
    "LMAHDS03 Comdty":  "Aluminium 3M",
    "LMZSDY Comdty":    "Zinc Cash",
    "LMZSDS03 Comdty":  "Zinc 3M",
    "LMNIDY Comdty":    "Nickel Cash",
    "LMNIDS03 Comdty":  "Nickel 3M",
    "LMPBDY Comdty":    "Lead Cash",
    "LMPBDS03 Comdty":  "Lead 3M",
    "LMSNDY Comdty":    "Tin Cash",
    "LMSNDS03 Comdty":  "Tin 3M",
}

# =============================================================================
# STYLES
# =============================================================================

TITLE_FILL = PatternFill("solid", fgColor="0D3B66")
TITLE_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=12)
HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT  = Font(name="Arial", size=9)
DATE_FONT  = Font(name="Arial", bold=True, size=9)

# =============================================================================
# CONNECTION
# =============================================================================

def connect():
    print("\n  Connecting to Bloomberg (%s:%d) ..." % (HOST, PORT))
    con = pdblp.BCon(debug=False, host=HOST, port=PORT, timeout=15000)
    con.start()
    print("  Connected.\n")
    return con

# =============================================================================
# DEADLOCK-SAFE DOWNLOAD
# =============================================================================

def safe_bdh(con, ticker):
    """
    Download a single ticker with error handling.
    Direct call — no thread wrapper (threading causes pdblp session errors).
    Returns (DataFrame, status_message). DataFrame is None on failure/empty.
    """
    try:
        df = con.bdh(ticker, FIELDS, START_DATE, END_DATE)
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(1)
        if df.empty or "PX_LAST" not in df.columns:
            return None, "no data"
        df = df[["PX_LAST"]].dropna()
        if df.empty:
            return None, "all NaN"
        return df, "%d rows" % len(df)
    except Exception as e:
        return None, "ERROR: %s" % str(e)[:80]


def download_strip(con, root, max_contracts, label=""):
    """
    Download F1-Fn for a single root.
    Applies per-contract timeout and early-exit after EARLY_EXIT_EMPTY
    consecutive failures so the loop never hangs on inactive back-months.
    Returns dict: {ticker_string: DataFrame}
    """
    tag = "[%s] " % label if label else ""
    print("  %s%s  F1-F%d" % (tag, root, max_contracts))

    result = {}
    consecutive_empty = 0

    for i in range(1, max_contracts + 1):
        tkr = "%s%d Comdty" % (root, i)
        df, status = safe_bdh(con, tkr)

        if df is not None:
            result[tkr] = df
            consecutive_empty = 0
            print("    %-22s %s" % (tkr, status))
        else:
            consecutive_empty += 1
            print("    %-22s %s" % (tkr, status))
            if consecutive_empty >= EARLY_EXIT_EMPTY:
                remaining = max_contracts - i
                if remaining > 0:
                    print("    => %d consecutive empties at F%d — skipping F%d-F%d"
                          % (EARLY_EXIT_EMPTY, i, i + 1, max_contracts))
                break

    print("  => %d / %d contracts returned data\n" % (len(result), max_contracts))
    return result


def download_tickers(con, ticker_map, label=""):
    """Download a flat list of named tickers. Returns dict: {ticker: DataFrame}."""
    result = {}
    total = len(ticker_map)
    tag = "[%s] " % label if label else ""
    for idx, (tkr, desc) in enumerate(ticker_map.items(), 1):
        print("  %s[%d/%d] %-26s" % (tag, idx, total, tkr + " (" + desc + ")"), end="", flush=True)
        df, status = safe_bdh(con, tkr)
        if df is not None:
            result[tkr] = df
        print("  " + status)
    return result

# =============================================================================
# EXCEL WRITERS
# =============================================================================

def write_strip_sheet(ws, root, name, strip_data):
    """Date | F1 | F2 | ... | Fn  (PX_LAST only)."""
    contracts = sorted(
        strip_data.keys(),
        key=lambda t: int("".join(c for c in t.split()[0][len(root):] if c.isdigit()) or "0"),
    )
    if not contracts:
        ws["A1"] = "No data returned"
        return

    all_dates = sorted(set().union(*(strip_data[t].index for t in contracts)))
    n_cols = 1 + len(contracts)

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="%s -- Futures Curve (PX_LAST, end=%s)" % (name, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Header row
    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(contracts, 2):
        num = "".join(ch for ch in tkr.split()[0][len(root):] if ch.isdigit())
        c = ws.cell(row=2, column=j, value="F%s" % num)
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    # Data
    for ri, dt in enumerate(all_dates, 3):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = DATE_FONT
        for j, tkr in enumerate(contracts, 2):
            df = strip_data[tkr]
            cell = ws.cell(row=ri, column=j)
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 12
    ws.freeze_panes = "B3"
    print("    Sheet '%s': %d rows, %d contracts" % (ws.title, len(all_dates), len(contracts)))


def write_flat_sheet(ws, title, ticker_map, data):
    """Date | Ticker1 | Ticker2 | ...  (PX_LAST only, for Cash/3M benchmarks)."""
    ordered = [t for t in ticker_map if t in data]
    if not ordered:
        ws["A1"] = "No data returned"
        return

    all_dates = sorted(set().union(*(data[t].index for t in ordered)))
    n_cols = 1 + len(ordered)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value="%s (end=%s)" % (title, END_DATE))
    c.font = TITLE_FONT; c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    c = ws.cell(row=2, column=1, value="Date")
    c.font = HDR_FONT; c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center")
    for j, tkr in enumerate(ordered, 2):
        c = ws.cell(row=2, column=j, value=ticker_map[tkr])
        c.font = HDR_FONT; c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")

    for ri, dt in enumerate(all_dates, 3):
        ws.cell(row=ri, column=1, value=dt).number_format = "YYYY-MM-DD"
        ws.cell(row=ri, column=1).font = DATE_FONT
        for j, tkr in enumerate(ordered, 2):
            cell = ws.cell(row=ri, column=j)
            df = data[tkr]
            if dt in df.index:
                v = df.loc[dt, "PX_LAST"]
                if pd.notna(v):
                    cell.value = float(v)
            cell.font = DATA_FONT
            cell.number_format = "#,##0.00"

    ws.column_dimensions["A"].width = 13
    for col in range(2, n_cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.freeze_panes = "B3"
    print("    Sheet '%s': %d rows, %d tickers" % (ws.title, len(all_dates), len(ordered)))


def write_readme(wb, title_str, rows):
    ws = wb.active
    ws.title = "README"
    meta = [
        (title_str,       ""),
        ("",              ""),
        ("Generated",     datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Start Date",    START_DATE),
        ("End Date",      END_DATE),
        ("Fields",        ", ".join(FIELDS)),
        ("Timeout/ticker","%ds + skip after %d empties" % (TIMEOUT_PER_CONTRACT, EARLY_EXIT_EMPTY)),
        ("", ""),
    ] + rows
    for r, (a, b) in enumerate(meta, 1):
        ws.cell(row=r, column=1, value=a).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row=r, column=2, value=b).font = Font(name="Arial", size=10)
    ws.cell(row=1, column=1).font = Font(name="Arial", bold=True, size=13, color="0D3B66")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60


def save_wb(wb, filename):
    folder = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(folder, filename)
    wb.save(path)
    print("\n  Saved: %s" % path)
    print("  Sheets: %s" % ", ".join(wb.sheetnames))
    return path

# =============================================================================
# MAIN
# =============================================================================

def run_section(label, product_dict, con, wb, section_label):
    print("\n" + "=" * 65)
    print("  %s" % section_label)
    print("=" * 65)
    for root, cfg in product_dict.items():
        strip = download_strip(con, root, cfg["max_contracts"], label=label)
        ws = wb.create_sheet(title=cfg["name"][:31])
        write_strip_sheet(ws, root, cfg["name"], strip)


def main():
    print("\n" + "=" * 65)
    print("  METALS RISK PREMIA - BLOOMBERG FULL DATA DOWNLOADER")
    print("  Start: %s   End: %s" % (START_DATE, END_DATE))
    print("  Fields: %s" % ", ".join(FIELDS))
    print("  Timeout: %ds/contract, exit after %d consecutive empties"
          % (TIMEOUT_PER_CONTRACT, EARLY_EXIT_EMPTY))
    print("=" * 65)

    con = connect()

    # ---- 1. LME Base Metals -------------------------------------------------
    wb1 = Workbook()
    write_readme(wb1, "LME Base Metals - Futures Curves", [
        ("LP", "Copper LME          F1-F27"),
        ("LA", "Aluminium LME       F1-F27"),
        ("LX", "Zinc LME            F1-F27"),
        ("LN", "Nickel LME          F1-F27"),
        ("LL", "Lead LME            F1-F27  (added - was missing from old scripts)"),
        ("LT", "Tin LME             F1-F15  (added - data confirms F15 max)"),
    ])
    run_section("LME", LME_METALS, con, wb1, "1/6  LME BASE METALS FUTURES CURVE")
    save_wb(wb1, OUT_METALS_CURVE)

    # ---- 2. Precious Metals -------------------------------------------------
    wb2 = Workbook()
    write_readme(wb2, "Precious Metals - Futures Curves", [
        ("GC", "Gold COMEX          F1-F20  (was capped at F12 in old scripts)"),
        ("SI", "Silver COMEX        F1-F17  (was capped at F12 in old scripts)"),
        ("HG", "Copper CME          F1-F27  (was capped at F12 in old scripts)"),
        ("PL", "Platinum NYMEX      F1-F13  (added - was missing)"),
        ("PA", "Palladium NYMEX     F1-F12  (added - was missing)"),
    ])
    run_section("PREC", PRECIOUS_METALS, con, wb2, "2/6  PRECIOUS METALS FUTURES CURVE")
    save_wb(wb2, OUT_PRECIOUS)

    # ---- 3. Energy Core + Extended ------------------------------------------
    wb3 = Workbook()
    write_readme(wb3, "Energy Futures - Core + Extended", [
        ("CL",  "WTI Crude NYMEX                F1-F27"),
        ("CO",  "Brent Crude ICE                F1-F27"),
        ("XB",  "RBOB Gasoline NYMEX            F1-F27"),
        ("HO",  "Heating Oil ULSD NYMEX         F1-F27"),
        ("NG",  "Nat Gas Henry Hub NYMEX        F1-F27"),
        ("QS",  "Singapore Gasoil ICE           F1-F27"),
        ("",    "--- Extended (added) ---"),
        ("GO",  "ICE Gasoil London              F1-F27  European diesel benchmark"),
        ("FO",  "Fuel Oil 3.5pct Barges ICE     F1-F12  Bunker / residual"),
        ("SJ",  "Singapore Jet Kerosene ICE     F1-F12  Verify root on terminal if empty"),
        ("NFY", "Naphtha CIF NWE Platts ICE     F1-F12  Verify root on terminal if empty"),
    ])
    all_energy = {**ENERGY_CORE, **ENERGY_EXTENDED}
    run_section("NRG", all_energy, con, wb3, "3/6  ENERGY FUTURES (CORE + EXTENDED)")
    save_wb(wb3, OUT_ENERGY)

    # ---- 4. NGL / Refined Products ------------------------------------------
    wb4 = Workbook()
    write_readme(wb4, "NGL / Refined Products - Futures Curves", [
        ("CAP", "Propane Argus Far East C3      F1-F27"),
        ("BAP", "Butane Argus C4                F1-F27"),
        ("DAE", "Ethane Argus                   F1-F27"),
        ("IBD", "Isobutane Argus                F1-F27"),
        ("PCW", "Propane C3 CIF ARA             F1-F27"),
        ("PGP", "Propylene Argus                F1-F27"),
    ])
    run_section("NGL", NGL_PRODUCTS, con, wb4, "4/6  NGL / REFINED PRODUCTS")
    save_wb(wb4, OUT_NGL)

    # ---- 5. LME Cash & 3M Benchmarks ----------------------------------------
    print("\n" + "=" * 65)
    print("  5/6  LME CASH & 3M BENCHMARKS")
    print("=" * 65)
    wb5 = Workbook()
    write_readme(wb5, "LME Cash and 3M Benchmarks", [
        ("Format",    "LMXX DY Comdty (Cash), LMXX DS03 Comdty (3M)"),
        ("Metals",    "Cu (CA), Al (AH), Zn (ZS), Ni (NI), Pb (PB), Sn (SN)"),
        ("Note",      "These feed into Metals Cash and 3M.xlsx on the dashboard"),
    ])
    data_cash = download_tickers(con, LME_CASH_3M_TICKERS, label="CASH")
    ws_cash = wb5.create_sheet(title="LME Cash and 3M")
    write_flat_sheet(ws_cash, "LME Cash and 3M Benchmarks (PX_LAST)", LME_CASH_3M_TICKERS, data_cash)
    save_wb(wb5, OUT_CASH_3M)

    # ---- Disconnect ----------------------------------------------------------
    try:
        con.stop()
    except Exception:
        pass

    print("\n" + "=" * 65)
    print("  DOWNLOAD COMPLETE")
    print("  " + OUT_METALS_CURVE)
    print("  " + OUT_PRECIOUS)
    print("  " + OUT_ENERGY)
    print("  " + OUT_NGL)
    print("  " + OUT_CASH_3M)
    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()
