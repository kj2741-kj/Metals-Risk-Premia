"""
value_signals.py
================
Value signals for LME Copper — two variants:

V1 — Long-Run MA Reversion (F1 through F15 as reference contract):
    deviation_t = (Fk_t - rolling_mean(Fk, N)) / rolling_mean(Fk, N)
    +1  if deviation < -threshold   (below fair value -> long)
    -1  if deviation > +threshold   (above fair value -> short)
     0  if within band              (near fair value -> flat)

V2 — Baz-Granger N-year Return Reversal (on F1_raw):
    reversal_t = F1_raw[t-N] - F1_raw[t]
    +1  if reversal > 0  (price has fallen over N years -> contrarian long)
    -1  if reversal < 0  (price has risen over N years -> contrarian short)
    No threshold — always in market.

Signal / PnL separation (no leakage):
    Signal   : computed from curve prices (Fk or F1_raw), no future data
    Position : Lag-1 = signal[t-1], Same-Day = signal[t]
    PnL      : Position[t] x delta_F1_continuous[t]

Outputs -> value_output/
"""

from __future__ import annotations

import io
import os
import sys
import warnings

import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
warnings.filterwarnings("ignore")

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "value_output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

CURVE_PATH   = os.path.join(BASE_DIR, "Metals Futures Curve.csv")
F1_CONT_PATH = os.path.join(BASE_DIR, "LME_Copper_Rolling_F1_v2.csv")

# ── Parameters ─────────────────────────────────────────────────────────────────
V1_CONTRACTS  = list(range(1, 16))                          # F1 through F15
V1_LOOKBACKS  = [252, 756, 1260, 1764, 2520]               # 1yr, 3yr, 5yr, 7yr, 10yr
V1_THRESHOLD  = 0.10                                        # ±10% default

V2_LOOKBACKS  = [252, 756, 1260, 1764, 2520]

LOOKBACK_LABELS = {
    252:  "1yr",
    756:  "3yr",
    1260: "5yr",
    1764: "7yr",
    2520: "10yr",
}

# ── Excel theme ────────────────────────────────────────────────────────────────
COPPER     = "B87333"
CHARCOAL   = "1E1E1E"
DARK_BG    = "161616"
PANEL_BG   = "222222"
MUTED_GOLD = "C9A84C"
LIGHT_TEXT = "D4CFC8"
DIM_TEXT   = "7A7068"

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(h, bold=False, size=10, mono=False):
    return Font(color=h, bold=bold, size=size, name="Consolas" if mono else "Calibri")

HDR_FILL  = _fill(COPPER)
HDR_FONT  = _font("FFFFFF", bold=True, size=10)
ALT_FILL  = _fill("1A1A1A")
BASE_FILL = _fill(DARK_BG)
CENTER    = Alignment(horizontal="center", vertical="center")
LEFT      = Alignment(horizontal="left",   vertical="center")
THIN      = Border(
    bottom=Side(style="thin", color="2A2A2A"),
    right=Side(style="thin",  color="2A2A2A"),
)


# ── Data loading ───────────────────────────────────────────────────────────────

def load_curve_prices() -> pd.DataFrame:
    """Load Metals Futures Curve -> Copper sheet -> F1-F15 prices."""
    try:
        xls = pd.ExcelFile(CURVE_PATH)
        cu_sheets = [s for s in xls.sheet_names if "copper" in s.lower()]
        sheet = cu_sheets[0] if cu_sheets else xls.sheet_names[0]
        df_raw = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=5)
        header_rows = []
        for i in range(min(4, len(df_raw))):
            row_vals = [str(v).strip().lower() for v in df_raw.iloc[i].values if pd.notna(v)]
            if any(kw in " ".join(row_vals) for kw in ["date", "f1", "f2", "price"]):
                header_rows.append(i)
        df = pd.read_excel(xls, sheet_name=sheet,
                           header=header_rows if header_rows else [0, 1])
        if isinstance(df.columns, pd.MultiIndex):
            new_cols = []
            for ct in df.columns:
                parts = [str(p).strip() for p in ct
                         if pd.notna(p) and "Unnamed" not in str(p) and str(p).strip()]
                new_cols.append("_".join(parts) if parts else str(ct))
            df.columns = new_cols
        else:
            df.columns = [str(c).strip() for c in df.columns]
        date_col = [c for c in df.columns if "date" in c.lower()]
        if date_col:
            df = df.rename(columns={date_col[0]: "Date"})
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df = df.dropna(subset=["Date"]).set_index("Date").sort_index()
        df.index = df.index.normalize()
        prices = {}
        for k in range(1, 16):
            for col in df.columns:
                cl = col.lower().replace(" ", "_")
                if cl.startswith(f"f{k}_") and "price" in cl:
                    prices[f"F{k}"] = pd.to_numeric(df[col], errors="coerce")
                    break
        result = pd.DataFrame(prices, index=df.index)
        print(f"  Loaded curve: {len(result)} rows, contracts: {list(result.columns)}")
        return result
    except Exception as e:
        print(f"  [ERROR] load_curve_prices: {e}")
        return pd.DataFrame()


def load_f1_continuous() -> pd.DataFrame:
    """Load LME_Copper_Rolling_F1_v2.csv -> F1_raw + F1_continuous."""
    df = pd.read_csv(F1_CONT_PATH, parse_dates=["Date"]).set_index("Date")
    df.index = df.index.normalize()
    return df[["F1_raw", "F1_continuous"]].sort_index()


# ── Signal computation ─────────────────────────────────────────────────────────

def compute_v1_deviation(curve_prices: pd.DataFrame, k: int, N: int) -> pd.Series:
    """
    V1: (Fk - MA_N) / MA_N  -- positive = above long-run mean (expensive).
    Signal = -1 when positive > threshold, +1 when negative > threshold, 0 within band.
    """
    col = f"F{k}"
    if col not in curve_prices.columns:
        return pd.Series(dtype=float)
    price = curve_prices[col].dropna()
    if len(price) < N // 2:
        return pd.Series(dtype=float)
    ma  = price.rolling(N, min_periods=max(N // 2, 60)).mean()
    dev = (price - ma) / ma.replace(0, np.nan)
    return dev.replace([np.inf, -np.inf], np.nan).dropna()


def compute_v2_reversal(f1r: pd.Series, N: int) -> pd.Series:
    """
    V2 (Baz-Granger): F1_raw[t-N] - F1_raw[t].
    Positive = price has fallen over N periods = contrarian long.
    """
    rev = f1r.shift(N) - f1r
    return rev.replace([np.inf, -np.inf], np.nan).dropna()


# ── Tradebook builder ──────────────────────────────────────────────────────────

def _compute_metrics(pos_np: np.ndarray, delta: np.ndarray,
                     f1c_prev: np.ndarray) -> dict:
    """Compute performance metrics for a position array."""
    gross_pnl = pos_np * delta
    with np.errstate(invalid="ignore", divide="ignore"):
        gross_ret = np.where(f1c_prev > 0, gross_pnl / f1c_prev, np.nan)
    active = gross_ret[(pos_np != 0) & np.isfinite(gross_ret)]
    n = len(active)
    if n < 20:
        return {}
    ann_r  = float(np.mean(active) * 252 * 100)
    ann_sd = float(np.std(active, ddof=1) * np.sqrt(252) * 100)
    sharpe = ann_r / ann_sd if ann_sd > 0 else np.nan
    down   = active[active < 0]
    srt_d  = float(np.std(down, ddof=1) * np.sqrt(252) * 100) if len(down) > 1 else np.nan
    sortino = ann_r / srt_d if srt_d and srt_d > 0 else np.nan
    cum_r   = pd.Series(np.where(np.isfinite(gross_ret), gross_ret, 0.0)).cumsum() * 100
    mdd     = float((cum_r - cum_r.cummax()).min())
    calmar  = ann_r / abs(mdd) if mdd != 0 else np.nan
    pnl_act = gross_pnl[(pos_np != 0)]
    wins    = pnl_act[pnl_act > 0]
    losses  = pnl_act[pnl_act < 0]
    T = len(pos_np)
    n_long  = int((pos_np > 0).sum())
    n_flat  = int((pos_np == 0).sum())
    n_short = int((pos_np < 0).sum())
    return {
        "sharpe": sharpe, "sortino": sortino,
        "ann_ret_pct": ann_r, "ann_std_pct": ann_sd,
        "mdd_pct": mdd, "calmar": calmar,
        "hit_rate": float((active > 0).mean()) * 100,
        "profit_factor": abs(wins.sum() / losses.sum()) if len(losses) > 0 else np.nan,
        "total_pnl_usdmt": float(np.nansum(gross_pnl)),
        "n": n,
        "pct_long":  n_long  / T * 100,
        "pct_flat":  n_flat  / T * 100,
        "pct_short": n_short / T * 100,
        "n_long": n_long, "n_flat": n_flat, "n_short": n_short,
    }


def build_value_tradebook(
    signal_raw: pd.Series,
    f1r: pd.Series,
    f1c: pd.Series,
    variant: str,
    threshold: float = 0.10,
) -> dict:
    """
    Build tradebook for a value signal.
    Returns: { tb: DataFrame, met_lag: dict, met_same: dict }
    """
    idx = signal_raw.index.intersection(f1c.index).intersection(f1r.index)
    sig = signal_raw.reindex(idx).dropna()
    idx = sig.index
    f1c_a = f1c.reindex(idx)
    f1r_a = f1r.reindex(idx)

    # Binary position from raw signal
    if variant == "v1":
        # V1: deviation negative -> price cheap -> LONG; deviation positive -> expensive -> SHORT
        sig_bin = np.where(sig.values < -threshold,  1.0,
                  np.where(sig.values >  threshold, -1.0, 0.0))
    else:
        # V2: reversal positive (price fallen) -> LONG; negative -> SHORT
        sig_bin = np.sign(sig.values).astype(float)

    T = len(sig_bin)
    pos_lag  = np.empty(T); pos_lag[0] = 0.0
    pos_lag[1:]  = np.where(np.isfinite(sig_bin[:-1]),  sig_bin[:-1],  0.0)
    pos_same = np.where(np.isfinite(sig_bin), sig_bin, 0.0)

    delta    = f1c_a.diff().values.astype(float)
    f1c_prev = f1c_a.shift(1).values.astype(float)

    met_lag  = _compute_metrics(pos_lag,  delta, f1c_prev)
    met_same = _compute_metrics(pos_same, delta, f1c_prev)

    # Signal state labels
    def _state(pos): return ["LONG" if p > 0 else ("SHORT" if p < 0 else "FLAT") for p in pos]

    tb = pd.DataFrame({
        "Date":              idx,
        "F1_raw":            f1r_a.values.round(2),
        "F1_continuous":     f1c_a.values.round(2),
        "Signal_Raw":        sig.values.round(6),
        "Position_Lag1":     pos_lag.astype(int),
        "State_Lag1":        _state(pos_lag),
        "Daily_PnL_Lag1":   (pos_lag  * delta).round(2),
        "Cum_PnL_Lag1":     np.nancumsum(pos_lag  * delta).round(2),
        "Position_SameDay":  pos_same.astype(int),
        "State_SameDay":     _state(pos_same),
        "Daily_PnL_SameDay": (pos_same * delta).round(2),
        "Cum_PnL_SameDay":  np.nancumsum(pos_same * delta).round(2),
    })
    return {"tb": tb, "met_lag": met_lag, "met_same": met_same}


# ── Excel output ───────────────────────────────────────────────────────────────

def _safe_sheet(name: str) -> str:
    for ch in r"/\?*[]:' ":
        name = name.replace(ch, "-")
    return name[:31]


def _write_sheet(ws, result: dict, title: str):
    """Write one strategy sheet with metrics block + tradebook."""
    met_l = result["met_lag"]
    met_s = result["met_same"]
    tb    = result["tb"]
    ws.sheet_view.showGridLines = False

    def _w(r, c, v, fill=None, font=None, align=None, border=None):
        cell = ws.cell(row=r, column=c, value=v)
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        if align:  cell.alignment = align
        if border: cell.border    = border

    # Title
    ws.row_dimensions[1].height = 20
    _w(1, 1, f"VALUE SIGNAL - {title}", fill=_fill(COPPER),
       font=_font("FFFFFF", bold=True, size=12), align=LEFT)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    _w(2, 1, "PERFORMANCE SUMMARY (% of notional unless noted)",
       fill=_fill(PANEL_BG), font=_font(MUTED_GOLD, bold=True), align=LEFT)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    def _na(v, fmt=".3f"):
        if v is None or (isinstance(v, float) and np.isnan(v)):
            return "-"
        try:
            return f"{v:{fmt}}"
        except Exception:
            return str(v)

    metrics = [
        ("Sharpe",          _na(met_l.get("sharpe")),             _na(met_s.get("sharpe"))),
        ("Sortino",         _na(met_l.get("sortino")),            _na(met_s.get("sortino"))),
        ("Ann Return %",    _na(met_l.get("ann_ret_pct")),        _na(met_s.get("ann_ret_pct"))),
        ("Ann Std Dev %",   _na(met_l.get("ann_std_pct")),        _na(met_s.get("ann_std_pct"))),
        ("Max Drawdown %",  _na(met_l.get("mdd_pct")),           _na(met_s.get("mdd_pct"))),
        ("Calmar",          _na(met_l.get("calmar")),             _na(met_s.get("calmar"))),
        ("Hit Rate %",      _na(met_l.get("hit_rate")),           _na(met_s.get("hit_rate"))),
        ("Profit Factor",   _na(met_l.get("profit_factor")),      _na(met_s.get("profit_factor"))),
        ("Total PnL $/MT",  _na(met_l.get("total_pnl_usdmt"), ",.2f"), _na(met_s.get("total_pnl_usdmt"), ",.2f")),
        ("Active Days",     _na(met_l.get("n"), ".0f"),           _na(met_s.get("n"), ".0f")),
        ("% Days Long",     _na(met_l.get("pct_long"), ".1f"),   _na(met_s.get("pct_long"), ".1f")),
        ("% Days Flat",     _na(met_l.get("pct_flat"), ".1f"),   _na(met_s.get("pct_flat"), ".1f")),
        ("% Days Short",    _na(met_l.get("pct_short"), ".1f"),  _na(met_s.get("pct_short"), ".1f")),
    ]

    hrow = 3
    for j, lbl in enumerate(["Metric", "Lag-1", "Same-Day"], 1):
        _w(hrow, j, lbl, fill=_fill(PANEL_BG),
           font=_font(COPPER, bold=True), align=CENTER)

    for i, (lbl, vl, vs) in enumerate(metrics, hrow + 1):
        bg = BASE_FILL if i % 2 == 0 else ALT_FILL
        _w(i, 1, lbl, fill=bg, font=_font(DIM_TEXT),                       align=LEFT)
        _w(i, 2, vl,  fill=bg, font=_font(LIGHT_TEXT, mono=True),          align=CENTER)
        _w(i, 3, vs,  fill=bg, font=_font(LIGHT_TEXT, mono=True),          align=CENTER)

    # Tradebook
    tb_row = hrow + len(metrics) + 2
    for j, col in enumerate(tb.columns, 1):
        _w(tb_row, j, col, fill=_fill(PANEL_BG),
           font=_font(COPPER, bold=True), align=CENTER)
        ws.column_dimensions[get_column_letter(j)].width = max(13, len(col) + 2)

    for ri, row in enumerate(tb.itertuples(index=False), tb_row + 1):
        bg = BASE_FILL if ri % 2 == 0 else ALT_FILL
        for j, val in enumerate(row, 1):
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%Y-%m-%d")
            _w(ri, j, val, fill=bg, font=_font(LIGHT_TEXT, mono=True),
               align=CENTER, border=THIN)

    ws.column_dimensions["A"].width = 13


def save_variant_excel(sheets: list[tuple[str, str, dict]], filepath: str):
    """Save list of (sheet_title, pretty_label, result) to .xlsx."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, title, result in sheets:
        ws = wb.create_sheet(title=_safe_sheet(sheet_name))
        _write_sheet(ws, result, title)
    wb.save(filepath)
    print(f"  Saved: {os.path.basename(filepath)}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    curve_prices = load_curve_prices()
    f1_df        = load_f1_continuous()
    f1r = f1_df["F1_raw"]
    f1c = f1_df["F1_continuous"]

    if curve_prices.empty:
        print("ERROR: Could not load curve prices from Metals Futures Curve.csv")
        return

    summary_rows = []

    # ── V1: MA Reversion (F1-F15 x 5 lookbacks) ──────────────────────────────
    print("\n=== V1: Long-Run MA Reversion (F1-F15) ===")
    for k in V1_CONTRACTS:
        col = f"F{k}"
        if col not in curve_prices.columns:
            print(f"  {col}: not in curve data, skipping")
            continue

        sheets = []
        for N in V1_LOOKBACKS:
            lbl = LOOKBACK_LABELS[N]
            dev = compute_v1_deviation(curve_prices, k, N)
            if dev.dropna().shape[0] < 200:
                continue
            res = build_value_tradebook(dev, f1r, f1c, variant="v1",
                                        threshold=V1_THRESHOLD)
            if not res["met_lag"]:
                continue
            title = f"{col} | {lbl} MA | threshold=10pct"
            sheets.append((f"{col}_{lbl}", title, res))
            ml = res["met_lag"]; ms = res["met_same"]
            print(f"  {col} {lbl:>4}: Sharpe lag={ml.get('sharpe', float('nan')):.3f}"
                  f"  same={ms.get('sharpe', float('nan')):.3f}"
                  f"  flat={ml.get('pct_flat', float('nan')):.1f}%")
            summary_rows.append({
                "Variant": "V1", "Contract": col, "Lookback": lbl,
                "Sharpe_Lag1":   round(ml.get("sharpe", np.nan), 4),
                "Sharpe_Same":   round(ms.get("sharpe", np.nan), 4),
                "AnnRet_Lag1":   round(ml.get("ann_ret_pct", np.nan), 2),
                "AnnRet_Same":   round(ms.get("ann_ret_pct", np.nan), 2),
                "MaxDD_Lag1":    round(ml.get("mdd_pct", np.nan), 2),
                "PctLong_Lag1":  round(ml.get("pct_long", np.nan), 1),
                "PctFlat_Lag1":  round(ml.get("pct_flat", np.nan), 1),
                "PctShort_Lag1": round(ml.get("pct_short", np.nan), 1),
            })

        if sheets:
            fpath = os.path.join(OUTPUT_DIR, f"Value_V1_{col}.xlsx")
            save_variant_excel(sheets, fpath)

    # ── V2: Baz-Granger N-year Reversal ───────────────────────────────────────
    print("\n=== V2: Baz-Granger N-year Reversal (F1_raw) ===")
    sheets_v2 = []
    for N in V2_LOOKBACKS:
        lbl = LOOKBACK_LABELS[N]
        rev = compute_v2_reversal(f1r, N)
        if rev.dropna().shape[0] < 200:
            continue
        res = build_value_tradebook(rev, f1r, f1c, variant="v2")
        if not res["met_lag"]:
            continue
        title = f"F1_raw | {lbl} reversal (Baz-Granger)"
        sheets_v2.append((f"V2_{lbl}", title, res))
        ml = res["met_lag"]; ms = res["met_same"]
        print(f"  {lbl:>4}: Sharpe lag={ml.get('sharpe', float('nan')):.3f}"
              f"  same={ms.get('sharpe', float('nan')):.3f}")
        summary_rows.append({
            "Variant": "V2", "Contract": "F1_raw", "Lookback": lbl,
            "Sharpe_Lag1":   round(ml.get("sharpe", np.nan), 4),
            "Sharpe_Same":   round(ms.get("sharpe", np.nan), 4),
            "AnnRet_Lag1":   round(ml.get("ann_ret_pct", np.nan), 2),
            "AnnRet_Same":   round(ms.get("ann_ret_pct", np.nan), 2),
            "MaxDD_Lag1":    round(ml.get("mdd_pct", np.nan), 2),
            "PctLong_Lag1":  round(ml.get("pct_long", np.nan), 1),
            "PctFlat_Lag1":  round(ml.get("pct_flat", np.nan), 1),
            "PctShort_Lag1": round(ml.get("pct_short", np.nan), 1),
        })

    if sheets_v2:
        fpath_v2 = os.path.join(OUTPUT_DIR, "Value_V2_BazGranger_Reversal.xlsx")
        save_variant_excel(sheets_v2, fpath_v2)

    # ── Summary ────────────────────────────────────────────────────────────────
    if summary_rows:
        summary_df = pd.DataFrame(summary_rows)
        out_path   = os.path.join(OUTPUT_DIR, "Value_All_Summary.csv")
        summary_df.to_csv(out_path, index=False)
        print(f"\nSummary -> {out_path}")

        print("\nTop 15 by Sharpe (Lag-1):")
        top = summary_df.nlargest(15, "Sharpe_Lag1")[
            ["Variant", "Contract", "Lookback",
             "Sharpe_Lag1", "Sharpe_Same", "PctFlat_Lag1"]]
        print(top.to_string(index=False))


if __name__ == "__main__":
    main()
