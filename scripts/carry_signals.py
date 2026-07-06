"""
carry_signals.py — LME Copper Carry Signal Generator
=====================================================
Produces 4 variant .xlsx files in carry_output/

V1: Roll Yield (sign of raw spread / F1)
    - (F1-F2)/F1
    - (F1-F3)/F1
    - (Cash-3M)/Cash

V2: Annualized Roll Yield  (Baz-Granger style)
    - (F1-F2)/F1 x 12   (~1 month between contracts)
    - (F1-F3)/F1 x 6    (~2 months)

V3: Long-Tenor Slope  (sign of Fj - Fk)
    - F3-F15, F4-F16, F5-F17, F6-F18, F7-F19,
      F8-F20, F9-F21, F10-F22, F11-F23, F12-F24

V4: Normalized Z-score of (F1-F2) carry
    - 252-day rolling z-score

Signal always computed from raw price series (no leakage).
PnL always from F1_continuous (return-based roll-adjusted).
Position conventions: Lag-1  AND  Same-Day (both shown in every sheet).
"""

import io
import os
import sys
import warnings
from pathlib import Path

import numpy as np
import pandas as pd

warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl not installed. Run: pip install openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
DATA_DIR = os.path.join(_REPO_ROOT, "data")
OUTPUTS_DIR = os.path.join(_REPO_ROOT, "outputs")

OUT_DIR = Path(OUTPUTS_DIR) / "carry_output"
OUT_DIR.mkdir(exist_ok=True)

FUTURES_FILE = Path(DATA_DIR) / "Metals Futures Curve.csv"
CASH_FILE    = Path(DATA_DIR) / "Metals Cash and 3M.xlsx"
F1_CONT_FILE = Path(DATA_DIR) / "LME_Copper_Rolling_F1_v2.csv"

# Excel style constants (copper / charcoal theme)
CLR_BG_DARK  = "0E0E0E"
CLR_BG_MID   = "161616"
CLR_COPPER   = "B87333"
CLR_SLATE    = "2B3A47"
CLR_AMBER    = "D4873A"
CLR_GREEN    = "5BAD72"
CLR_RED      = "B85450"
CLR_HEADER   = "1A1A1A"
CLR_TEXT     = "E8E0D5"
CLR_DIM      = "8A8A8A"

# ─────────────────────────────────────────────────────────────────────────────
# DATA LOADERS
# ─────────────────────────────────────────────────────────────────────────────

def load_futures_curve() -> pd.DataFrame:
    """Return DataFrame of F1-F27 prices for Copper LME, Date index."""
    xls = pd.ExcelFile(FUTURES_FILE)
    df  = pd.read_excel(xls, sheet_name="Copper LME", header=[0, 1, 2])

    new_cols = []
    for c in df.columns:
        parts = [str(p).strip() for p in c
                 if "Unnamed" not in str(p) and str(p).strip()]
        new_cols.append("_".join(parts) if parts else str(c))
    df.columns = new_cols

    date_col = [c for c in df.columns if "date" in c.lower()][0]
    df = df.rename(columns={date_col: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"]).set_index("Date").sort_index()
    df.index = df.index.normalize()

    prices = {}
    for col in df.columns:
        cl = col.lower()
        for i in range(1, 28):
            if f"f{i}_price" in cl:
                prices[f"F{i}"] = pd.to_numeric(df[col], errors="coerce")
                break

    return pd.DataFrame(prices, index=df.index)


def load_cash_3m() -> pd.DataFrame:
    """Return DataFrame with cash_price and 3m_price columns, Date index."""
    xls = pd.ExcelFile(CASH_FILE)
    df  = pd.read_excel(xls, sheet_name="LME Copper", header=[0, 1, 2])

    new_cols = []
    for c in df.columns:
        parts = [str(p).strip() for p in c
                 if "Unnamed" not in str(p) and str(p).strip()]
        new_cols.append("_".join(parts) if parts else str(c))
    df.columns = new_cols

    date_col = [c for c in df.columns if "date" in c.lower()][0]
    df = df.rename(columns={date_col: "Date"})
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df = df.dropna(subset=["Date"]).set_index("Date").sort_index()
    df.index = df.index.normalize()

    result = pd.DataFrame(index=df.index)
    for col in df.columns:
        cl = col.lower()
        if "cash" in cl and "price" in cl and "spread" not in cl and "cash_price" not in result.columns:
            result["cash_price"] = pd.to_numeric(df[col], errors="coerce")
        elif "3m" in cl and "price" in cl and "spread" not in cl and "3m_price" not in result.columns:
            result["3m_price"]   = pd.to_numeric(df[col], errors="coerce")
        elif "spread" in cl and "price" in cl and "spread_price" not in result.columns:
            result["spread_price"] = pd.to_numeric(df[col], errors="coerce")

    return result


def load_f1_continuous() -> pd.DataFrame:
    """Return DataFrame with F1_raw and F1_continuous, Date index."""
    df = pd.read_csv(F1_CONT_FILE, parse_dates=["Date"]).set_index("Date")
    df.index = df.index.normalize()
    return df[["F1_raw", "F1_continuous"]].sort_index()


# ─────────────────────────────────────────────────────────────────────────────
# PERFORMANCE METRICS
# ─────────────────────────────────────────────────────────────────────────────

def _compute_metrics(daily_pnl: pd.Series, position: pd.Series,
                     f1c: pd.Series) -> dict:
    """All metrics in % terms (daily_ret = daily_pnl / F1_cont_prev)."""
    f1_prev   = f1c.shift(1)
    daily_ret = (daily_pnl / f1_prev).replace([np.inf, -np.inf], np.nan)

    active    = daily_ret[position != 0].dropna()
    pnl_act   = daily_pnl[position != 0].dropna()
    n         = len(active)

    if n < 20:
        return {"n": n, "sharpe": np.nan, "sortino": np.nan,
                "ann_ret_pct": np.nan, "ann_std_pct": np.nan,
                "mdd_pct": np.nan, "calmar": np.nan,
                "hit_rate": np.nan, "profit_factor": np.nan,
                "total_pnl": np.nan}

    ann_r  = float(active.mean() * 252 * 100)
    ann_sd = float(active.std()  * np.sqrt(252) * 100)
    sharpe = ann_r / ann_sd if ann_sd > 0 else np.nan

    down    = active[active < 0]
    srt_d   = float(down.std() * np.sqrt(252) * 100) if len(down) > 1 else np.nan
    sortino = ann_r / srt_d if (srt_d and srt_d > 0) else np.nan

    cum_r   = daily_ret.fillna(0).cumsum() * 100
    mdd     = float((cum_r - cum_r.cummax()).min())
    calmar  = ann_r / abs(mdd) if mdd != 0 else np.nan

    wins    = pnl_act[pnl_act > 0]
    losses  = pnl_act[pnl_act < 0]
    hit     = float((active > 0).mean()) * 100
    pf      = abs(wins.sum() / losses.sum()) if len(losses) > 0 else np.nan
    tot_pnl = float(pnl_act.sum())

    return {"n": n, "sharpe": sharpe, "sortino": sortino,
            "ann_ret_pct": ann_r, "ann_std_pct": ann_sd,
            "mdd_pct": mdd, "calmar": calmar,
            "hit_rate": hit, "profit_factor": pf,
            "total_pnl": tot_pnl}


# ─────────────────────────────────────────────────────────────────────────────
# TRADEBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def build_carry_tradebook(carry_raw: pd.Series, f1r: pd.Series,
                          f1c: pd.Series) -> dict:
    """
    Parameters
    ----------
    carry_raw : raw carry value series (pre-sign), Date index
    f1r       : F1_raw, Date index
    f1c       : F1_continuous, Date index

    Returns
    -------
    dict with keys:
        tb       : DataFrame (full tradebook, both timing conventions)
        met_lag  : performance metrics dict for Lag-1
        met_same : performance metrics dict for Same-Day
    """
    # Align to intersection of all three series
    idx = f1r.index.intersection(f1c.index).intersection(carry_raw.dropna().index)
    f1r_  = f1r.reindex(idx)
    f1c_  = f1c.reindex(idx)
    raw_  = carry_raw.reindex(idx)

    sig   = np.sign(raw_.values).astype(float)
    T     = len(sig)

    # Lag-1 position
    pos_lag = np.empty(T)
    pos_lag[0] = 0.0
    pos_lag[1:] = np.where(np.isfinite(sig[:-1]), sig[:-1], 0.0)

    # Same-Day position
    pos_same = np.where(np.isfinite(sig), sig, 0.0).astype(float)

    delta = np.diff(f1c_.values, prepend=np.nan)

    pnl_lag  = pd.Series(pos_lag  * delta, index=idx)
    pnl_same = pd.Series(pos_same * delta, index=idx)

    pos_lag_s  = pd.Series(pos_lag,  index=idx)
    pos_same_s = pd.Series(pos_same, index=idx)

    met_lag  = _compute_metrics(pnl_lag,  pos_lag_s,  f1c_)
    met_same = _compute_metrics(pnl_same, pos_same_s, f1c_)

    tb = pd.DataFrame({
        "F1_raw":           f1r_.values,
        "F1_continuous":    f1c_.values,
        "Carry_Raw":        raw_.values,
        "Signal":           sig,
        "Position_Lag1":    pos_lag,
        "Daily_PnL_Lag1":   pnl_lag.values,
        "Cum_PnL_Lag1":     pnl_lag.cumsum().values,
        "Position_SameDay": pos_same,
        "Daily_PnL_SameDay": pnl_same.values,
        "Cum_PnL_SameDay":  pnl_same.cumsum().values,
    }, index=idx)

    return {"tb": tb, "met_lag": met_lag, "met_same": met_same}


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITER
# ─────────────────────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color=CLR_TEXT, size=10) -> Font:
    return Font(name="Calibri", bold=bold, color=color, size=size)


def _write_carry_sheet(ws, sheet_label: str, result: dict,
                       variant_desc: str, sub_desc: str) -> None:
    """Write one sub-variant to a worksheet."""
    tb       = result["tb"]
    met_lag  = result["met_lag"]
    met_same = result["met_same"]

    METRIC_LABELS = [
        ("Annualized Return %",    "ann_ret_pct",   ".2f"),
        ("Annualized Std Dev %",   "ann_std_pct",   ".2f"),
        ("Sharpe Ratio",           "sharpe",        ".2f"),
        ("Sortino Ratio",          "sortino",        ".2f"),
        ("Max Drawdown %",         "mdd_pct",       ".2f"),
        ("Calmar Ratio",           "calmar",        ".2f"),
        ("Hit Rate %",             "hit_rate",      ".2f"),
        ("Profit Factor",          "profit_factor", ".2f"),
        ("Total PnL ($/MT)",       "total_pnl",     ",.2f"),
        ("Active Days",            "n",             ",.0f"),
    ]

    ws.sheet_view.showGridLines = False

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = f"LME Copper — {variant_desc} | {sub_desc}"
    ws["A1"].font  = Font(name="Calibri", bold=True, size=13, color=CLR_COPPER)
    ws["A1"].fill  = _fill(CLR_BG_DARK)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 24

    # ── Performance header ────────────────────────────────────────────────────
    ws["A3"] = "Metric"
    ws["B3"] = "Lag-1 (Next-Day Entry)"
    ws["D3"] = "Same-Day Entry"
    for cell_addr, label in [("A3", "Metric"), ("B3", "Lag-1 (Next-Day Entry)"),
                              ("D3", "Same-Day Entry")]:
        c = ws[cell_addr]
        c.font  = _font(bold=True, color=CLR_COPPER, size=10)
        c.fill  = _fill(CLR_HEADER)
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 18

    # ── Metric rows ───────────────────────────────────────────────────────────
    for row_i, (label, key, fmt) in enumerate(METRIC_LABELS, start=4):
        r = row_i

        c_label = ws.cell(r, 1, value=label)
        c_label.font  = _font(color=CLR_DIM)
        c_label.fill  = _fill(CLR_BG_MID)
        c_label.alignment = Alignment(horizontal="left")

        for col_offset, met in [(2, met_lag), (4, met_same)]:
            v = met.get(key, np.nan)
            try:
                if np.isnan(float(v)):
                    display = "—"
                else:
                    display = f"{v:{fmt}}"
            except Exception:
                display = str(v)
            c = ws.cell(r, col_offset, value=display)
            c.font  = _font(bold=(key == "sharpe"), color=CLR_TEXT)
            c.fill  = _fill(CLR_BG_MID)
            c.alignment = Alignment(horizontal="right")

        # Spacer col between lag and same-day
        ws.cell(r, 3).fill = _fill(CLR_BG_DARK)

        ws.row_dimensions[r].height = 16

    # ── Column headers for tradebook ──────────────────────────────────────────
    HDR_ROW = 4 + len(METRIC_LABELS) + 2
    tb_cols = ["Date", "F1_raw", "F1_continuous", "Carry_Raw", "Signal",
               "Position_Lag1", "Daily_PnL_Lag1", "Cum_PnL_Lag1",
               "Position_SameDay", "Daily_PnL_SameDay", "Cum_PnL_SameDay"]

    for ci, h in enumerate(tb_cols, start=1):
        c = ws.cell(HDR_ROW, ci, value=h)
        c.font      = _font(bold=True, color=CLR_COPPER)
        c.fill      = _fill(CLR_HEADER)
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[HDR_ROW].height = 18

    # ── Tradebook rows ────────────────────────────────────────────────────────
    for di, (dt, row_vals) in enumerate(tb.iterrows()):
        r = HDR_ROW + 1 + di
        ws.cell(r, 1, value=dt.strftime("%Y-%m-%d")).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, value=round(float(row_vals["F1_raw"]), 2) if not np.isnan(row_vals["F1_raw"]) else None)
        ws.cell(r, 3, value=round(float(row_vals["F1_continuous"]), 2) if not np.isnan(row_vals["F1_continuous"]) else None)
        ws.cell(r, 4, value=round(float(row_vals["Carry_Raw"]), 6) if not np.isnan(row_vals["Carry_Raw"]) else None)
        ws.cell(r, 5, value=int(row_vals["Signal"]) if np.isfinite(row_vals["Signal"]) else None)

        for ci_offset, col_name in enumerate(
            ["Position_Lag1", "Daily_PnL_Lag1", "Cum_PnL_Lag1",
             "Position_SameDay", "Daily_PnL_SameDay", "Cum_PnL_SameDay"], start=6
        ):
            v = row_vals[col_name]
            ws.cell(r, ci_offset, value=round(float(v), 4) if not np.isnan(v) else None)

        bg = CLR_BG_DARK if di % 2 == 0 else CLR_BG_MID
        for ci in range(1, len(tb_cols) + 1):
            cell = ws.cell(r, ci)
            cell.fill = _fill(bg)
            if cell.font is None:
                cell.font = _font()

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [12, 10, 14, 14, 8, 14, 16, 16, 14, 16, 16]
    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # wider metric label column
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 4   # spacer
    ws.column_dimensions["D"].width = 22


_INVALID_XL_CHARS = str.maketrans({
    "/": "-", "\\": "-", "?": "", "*": "", "[": "(", "]": ")",
    ":": " ", "'": "",
})


def _safe_sheet_title(s: str) -> str:
    return s.translate(_INVALID_XL_CHARS)[:31]


def save_variant_excel(sheet_results: list, filepath: Path,
                       variant_desc: str) -> None:
    """
    sheet_results : list of (sheet_label, sub_desc, result_dict)
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_label, sub_desc, result in sheet_results:
        ws = wb.create_sheet(title=_safe_sheet_title(sheet_label))
        ws.sheet_properties.tabColor = CLR_COPPER
        _write_carry_sheet(ws, sheet_label, result, variant_desc, sub_desc)

    wb.save(filepath)
    print(f"  Saved: {filepath}")


# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY TABLE BUILDER
# ─────────────────────────────────────────────────────────────────────────────

def _met_row(variant: str, sub: str, timing: str, m: dict) -> dict:
    def _f(v, fmt=".2f"):
        try:
            return f"{v:{fmt}}" if not np.isnan(float(v)) else "—"
        except Exception:
            return "—"
    return {
        "Variant":    variant,
        "Sub-Variant": sub,
        "Timing":     timing,
        "Sharpe":     _f(m.get("sharpe", np.nan)),
        "Sortino":    _f(m.get("sortino", np.nan)),
        "Ann Ret %":  _f(m.get("ann_ret_pct", np.nan)),
        "Std Dev %":  _f(m.get("ann_std_pct", np.nan)),
        "Max DD %":   _f(m.get("mdd_pct", np.nan)),
        "Calmar":     _f(m.get("calmar", np.nan)),
        "Hit Rate %": _f(m.get("hit_rate", np.nan)),
        "PF":         _f(m.get("profit_factor", np.nan)),
        "Total PnL":  _f(m.get("total_pnl", np.nan), ",.2f"),
        "Active Days": _f(m.get("n", np.nan), ",.0f"),
    }


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("Loading data...")
    curves = load_futures_curve()   # DataFrame: F1-F27, Date index
    cash   = load_cash_3m()         # DataFrame: cash_price, 3m_price, Date index
    f1_df  = load_f1_continuous()   # DataFrame: F1_raw, F1_continuous, Date index

    f1r = f1_df["F1_raw"]
    f1c = f1_df["F1_continuous"]

    # Use futures curve F1 for signals where curves.F1 available
    # (they're the same source; f1r from the CSV is authoritative for dates)
    # For carry we derive signal from curves directly
    F = curves  # F["F1"], F["F2"], ... F["F27"]

    summary_rows = []

    # ─────────────────────────────────────────────────────────────────────────
    # V1: ROLL YIELD  — sign((Fi - Fj) / Fi)
    # ─────────────────────────────────────────────────────────────────────────
    print("\nVariant 1: Roll Yield ...")
    v1_sub = [
        ("(F1-F2)/F1",    "(F1 - F2) / F1",    (F["F1"] - F["F2"]) / F["F1"]),
        ("(F1-F3)/F1",    "(F1 - F3) / F1",    (F["F1"] - F["F3"]) / F["F1"]),
        ("(Cash-3M)/Cash","(Cash - 3M) / Cash", (cash["cash_price"] - cash["3m_price"]) / cash["cash_price"]),
    ]

    v1_sheets = []
    for sheet_lbl, sub_desc, carry_raw in v1_sub:
        print(f"  {sub_desc}")
        res = build_carry_tradebook(carry_raw, f1r, f1c)
        v1_sheets.append((sheet_lbl, sub_desc, res))
        summary_rows.append(_met_row("V1 Roll Yield", sub_desc, "Lag-1",    res["met_lag"]))
        summary_rows.append(_met_row("V1 Roll Yield", sub_desc, "Same-Day", res["met_same"]))

    save_variant_excel(v1_sheets, OUT_DIR / "Carry_V1_RollYield.xlsx",
                       "V1 - Roll Yield  sign[(Fi - Fj) / Fi]")

    # ─────────────────────────────────────────────────────────────────────────
    # V2: ANNUALIZED ROLL YIELD
    # ─────────────────────────────────────────────────────────────────────────
    print("\nVariant 2: Annualized Roll Yield ...")
    # Approximate: 12 months/year → F1-F2 annualization factor 12
    #                                F1-F3 annualization factor 6
    v2_sub = [
        ("(F1-F2)-F1 x12",  "(F1-F2)/F1 x 12  [~1-month annualized]",
         (F["F1"] - F["F2"]) / F["F1"] * 12),
        ("(F1-F3)-F1 x6",   "(F1-F3)/F1 x 6   [~2-month annualized]",
         (F["F1"] - F["F3"]) / F["F1"] * 6),
    ]

    v2_sheets = []
    for sheet_lbl, sub_desc, carry_raw in v2_sub:
        print(f"  {sub_desc}")
        res = build_carry_tradebook(carry_raw, f1r, f1c)
        v2_sheets.append((sheet_lbl, sub_desc, res))
        summary_rows.append(_met_row("V2 Annualized", sub_desc, "Lag-1",    res["met_lag"]))
        summary_rows.append(_met_row("V2 Annualized", sub_desc, "Same-Day", res["met_same"]))

    save_variant_excel(v2_sheets, OUT_DIR / "Carry_V2_Annualized.xlsx",
                       "V2 - Annualized Roll Yield  (Baz-Granger style)")

    # ─────────────────────────────────────────────────────────────────────────
    # V3: LONG-TENOR SLOPE  — sign(Fj - Fk)
    # Pairs: (3,15), (4,16), ..., (12,24)
    # ─────────────────────────────────────────────────────────────────────────
    print("\nVariant 3: Long-Tenor Slope ...")
    v3_sub = []
    for j in range(3, 13):          # j = 3..12, k = j+12
        k = j + 12
        sheet_lbl = f"F{j}-F{k}"
        sub_desc  = f"F{j} - F{k}  (near minus far, +1=backwardation)"
        carry_raw = F[f"F{j}"] - F[f"F{k}"]
        print(f"  {sheet_lbl}")
        res = build_carry_tradebook(carry_raw, f1r, f1c)
        v3_sub.append((sheet_lbl, sub_desc, res))
        summary_rows.append(_met_row("V3 Long Slope", sheet_lbl, "Lag-1",    res["met_lag"]))
        summary_rows.append(_met_row("V3 Long Slope", sheet_lbl, "Same-Day", res["met_same"]))

    save_variant_excel(v3_sub, OUT_DIR / "Carry_V3_LongSlope.xlsx",
                       "V3 - Long-Tenor Slope  sign(Fj - Fk)")

    # ─────────────────────────────────────────────────────────────────────────
    # V4: Z-SCORE OF CARRY
    # ─────────────────────────────────────────────────────────────────────────
    print("\nVariant 4: Z-score of Carry ...")
    raw_carry   = (F["F1"] - F["F2"]) / F["F1"]
    roll_mean   = raw_carry.rolling(252).mean()
    roll_std    = raw_carry.rolling(252).std()
    z_carry     = (raw_carry - roll_mean) / roll_std.replace(0, np.nan)

    sub_desc_v4 = "Z-score of (F1-F2)/F1  [252-day rolling]"
    print(f"  {sub_desc_v4}")
    res_v4 = build_carry_tradebook(z_carry, f1r, f1c)

    v4_sheets = [("Zscore_252d", sub_desc_v4, res_v4)]
    save_variant_excel(v4_sheets, OUT_DIR / "Carry_V4_Zscore.xlsx",
                       "V4 - Z-score Carry  (252-day rolling)")

    summary_rows.append(_met_row("V4 Z-score", sub_desc_v4, "Lag-1",    res_v4["met_lag"]))
    summary_rows.append(_met_row("V4 Z-score", sub_desc_v4, "Same-Day", res_v4["met_same"]))

    # ─────────────────────────────────────────────────────────────────────────
    # SAVE SUMMARY CSV
    # ─────────────────────────────────────────────────────────────────────────
    summary_df = pd.DataFrame(summary_rows)
    summary_path = OUT_DIR / "Carry_All_Variants_Summary.csv"
    summary_df.to_csv(summary_path, index=False)
    print(f"\n  Summary CSV: {summary_path}")

    # ─────────────────────────────────────────────────────────────────────────
    # PRINT SUMMARY TABLE
    # ─────────────────────────────────────────────────────────────────────────
    print("\n" + "=" * 120)
    print("  CARRY SIGNALS — ALL VARIANTS SUMMARY")
    print("=" * 120)
    print(f"  {'Variant':<16} {'Sub-Variant':<34} {'Timing':<12} "
          f"{'Sharpe':>7} {'Sortino':>8} {'AnnRet%':>8} {'StdDev%':>8} "
          f"{'MaxDD%':>8} {'Calmar':>7} {'HitRt%':>7} {'PF':>6} {'TotalPnL':>12}")
    print("-" * 120)

    prev_v = ""
    for row in summary_rows:
        v = row["Variant"]
        sep = "  " + "-" * 118 + "\n" if v != prev_v and prev_v != "" else ""
        print(sep, end="")
        print(f"  {v:<16} {row['Sub-Variant']:<34} {row['Timing']:<12} "
              f"{row['Sharpe']:>7} {row['Sortino']:>8} {row['Ann Ret %']:>8} "
              f"{row['Std Dev %']:>8} {row['Max DD %']:>8} {row['Calmar']:>7} "
              f"{row['Hit Rate %']:>7} {row['PF']:>6} {row['Total PnL']:>12}")
        prev_v = v

    print("=" * 120)
    print(f"\nAll files saved to {OUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
