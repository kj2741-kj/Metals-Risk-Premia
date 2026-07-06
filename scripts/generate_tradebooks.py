"""
generate_tradebooks.py  — v2: TC=5bps base case + Anchors IS/OOS
=================================================================
Folder: tradebooks/  (same structure as before)
New files added:
  momentum/IS/  MOM_IS_Anchors_EW_Lag1.xlsx
  momentum/OOS/ MOM_OOS_Anchors_ISopt_WalkForward.xlsx

TC incorporated in EVERY tradebook:
  TC_Cost(t) = |Position(t) − Position(t−1)| × (5/10,000/2) × F1_Continuous(t)
  Net_PnL(t) = Gross_PnL(t) − TC_Cost(t)
  Net_Return(t) = Net_PnL(t) / F1_Continuous(t−1)

  5 bps = the per-flip half-spread.
  Full flip +1→−1 at $9,000/MT costs: |2| × 0.025% × $9,000 = $4.50/MT
  Entry 0→+1 at $9,000/MT costs:      |1| × 0.025% × $9,000 = $2.25/MT
  The /2 reflects that each directional move has a half-spread cost per side.
  This matches exactly the TC formula used in the dashboard (tc_bps/10000/2).

All IS tradebooks show BOTH gross and net-5bps performance side-by-side.
"""

import os, sys, io, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

import numpy as np
import pandas as pd
from scipy.optimize import minimize
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
DATA_DIR = os.path.join(_REPO_ROOT, "data")
OUTPUTS_DIR = os.path.join(_REPO_ROOT, "outputs")
BASE    = DATA_DIR
TBDIR   = os.path.join(OUTPUTS_DIR, "tradebooks")
TC_BASE = 5   # bps — base case TC applied in every tradebook

for path in [
    TBDIR,
    f"{TBDIR}/momentum/IS", f"{TBDIR}/momentum/OOS",
    f"{TBDIR}/carry/IS",
    f"{TBDIR}/value/IS",    f"{TBDIR}/value/OOS",
    f"{TBDIR}/portfolio/IS", f"{TBDIR}/portfolio/OOS",
]:
    os.makedirs(path, exist_ok=True)
print("Folders OK.")

# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading F1…")
f1_df = pd.read_csv(f"{BASE}/LME_Copper_Rolling_F1_v2.csv",
                    parse_dates=["Date"]).set_index("Date")
f1_df.index = f1_df.index.normalize()
f1r = f1_df["F1_raw"].sort_index()
f1c = f1_df["F1_continuous"].sort_index()
print(f"  F1: {len(f1r)} rows  {f1r.index[0].date()} → {f1r.index[-1].date()}")

print("Loading curve…")
xls   = pd.ExcelFile(f"{BASE}/Metals Futures Curve.csv")
cu_sh = next(s for s in xls.sheet_names if "copper" in s.lower() and "lme" in s.lower())
df_raw = pd.read_excel(xls, sheet_name=cu_sh, header=None, nrows=5)
hr = [i for i in range(min(4, len(df_raw)))
      if any(kw in " ".join(str(v).lower() for v in df_raw.iloc[i].values if pd.notna(v))
             for kw in ["date","f1","f2","price"])]
df_c = pd.read_excel(xls, sheet_name=cu_sh, header=hr if hr else [0,1])
if isinstance(df_c.columns, pd.MultiIndex):
    df_c.columns = ["_".join(str(p) for p in t if pd.notna(p)
                              and "Unnamed" not in str(p) and str(p).strip()) or str(t)
                    for t in df_c.columns]
df_c.columns = [str(c).strip() for c in df_c.columns]
dc = [c for c in df_c.columns if "date" in c.lower()]
if dc: df_c = df_c.rename(columns={dc[0]: "Date"})
df_c["Date"] = pd.to_datetime(df_c["Date"], errors="coerce")
df_c = df_c.dropna(subset=["Date"]).set_index("Date").sort_index()
df_c.index = df_c.index.normalize()
crv = {}
for col in df_c.columns:
    cl = col.lower().replace(" ","_")
    for i in range(1, 27):
        if f"f{i}_" in cl and "price" in cl:
            crv[f"F{i}"] = pd.to_numeric(df_c[col], errors="coerce")
            break
crv = pd.DataFrame(crv, index=df_c.index)
print(f"  Curve: {list(crv.columns)[:8]}…  rows={len(crv)}")

print("Loading Cash&3M…")
cash_cu = pd.DataFrame()
try:
    xls_cash = pd.ExcelFile(f"{BASE}/Metals Cash and 3M.xlsx")
    cu_cs = next((s for s in xls_cash.sheet_names if "copper" in s.lower()), None)
    if cu_cs:
        df_cash = pd.read_excel(xls_cash, sheet_name=cu_cs, header=[0,1,2])
        nc = []
        for c in df_cash.columns:
            parts = [str(p).strip() for p in c if str(p).strip() and "Unnamed" not in str(p)]
            nc.append("_".join(parts) if parts else str(c))
        df_cash.columns = nc
        dc2 = [c for c in df_cash.columns if "date" in c.lower()]
        if dc2: df_cash = df_cash.rename(columns={dc2[0]: "Date"})
        df_cash["Date"] = pd.to_datetime(df_cash["Date"], errors="coerce")
        df_cash = df_cash.dropna(subset=["Date"]).set_index("Date").sort_index()
        df_cash.index = df_cash.index.normalize()
        result = pd.DataFrame(index=df_cash.index)
        for col in df_cash.columns:
            cl = col.lower()
            if "spread" not in cl and "cash" in cl and "price" in cl and "cash_price" not in result.columns:
                result["cash_price"] = pd.to_numeric(df_cash[col], errors="coerce")
            elif "spread" not in cl and "3m" in cl and "price" in cl and "3m_price" not in result.columns:
                result["3m_price"] = pd.to_numeric(df_cash[col], errors="coerce")
        cash_cu = result
        print(f"  Cash/3M: {len(cash_cu)} rows")
except Exception as e:
    print(f"  Cash/3M load failed: {e}")

IS_W, OOS_W = 1260, 252

# ── Style helpers ─────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", fgColor="1A1A2E")
D1_FILL   = PatternFill("solid", fgColor="0F0F1A")
D2_FILL   = PatternFill("solid", fgColor="141428")
POS_FILL  = PatternFill("solid", fgColor="003300")
NEG_FILL  = PatternFill("solid", fgColor="330000")
DSC_FILL  = PatternFill("solid", fgColor="0D0D1F")
HDR_FONT  = Font(name="Calibri", bold=True, color="B87333", size=10)
DAT_FONT  = Font(name="Calibri", color="D4D4D4", size=9)
DSC_FONT  = Font(name="Calibri", color="C8B89A", size=10)
TTL_FONT  = Font(name="Calibri", bold=True, color="E8D5B0", size=13)
KEY_FONT  = Font(name="Calibri", bold=True, color="B87333", size=10)
TC_FONT   = Font(name="Calibri", bold=True, color="64B5F6", size=9)   # blue for net-TC columns

def _sw(ws): ws.sheet_view.showGridLines = False
def _cw(ws, j, w): ws.column_dimensions[get_column_letter(j)].width = w
def _hdr(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
def _dat(ws, r, c, v, alt=False, hl=None):
    cell = ws.cell(row=r, column=c, value=v)
    fill = D2_FILL if alt else D1_FILL
    if hl == "pos": fill = POS_FILL
    elif hl == "neg": fill = NEG_FILL
    cell.font = DAT_FONT; cell.fill = fill
    cell.alignment = Alignment(horizontal="center")
def _dat_tc(ws, r, c, v, alt=False):   # blue tint for TC/Net columns
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = TC_FONT
    cell.fill = D2_FILL if alt else D1_FILL
    cell.alignment = Alignment(horizontal="center")
def _desc(ws, r, label, value=""):
    c1 = ws.cell(row=r, column=1, value=label); c1.font = KEY_FONT; c1.fill = DSC_FILL
    c2 = ws.cell(row=r, column=2, value=value); c2.font = DSC_FONT; c2.fill = DSC_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
def _title(ws, r, txt):
    c = ws.cell(row=r, column=1, value=txt); c.font = TTL_FONT; c.fill = DSC_FILL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)

TC_DESC_ROWS = [
    ("", ""),
    ("─── TRANSACTION COST (5 bps base case) ───────────────────────────────────────────", ""),
    ("TC formula",
     "TC_Cost(t) = |Position(t) − Position(t−1)| × (5 / 10,000 / 2) × F1_Continuous(t)"),
    ("Net_PnL(t)", "Gross_PnL(t) − TC_Cost(t)  [USD/MT]"),
    ("Net_Return(t)", "Net_PnL(t) / F1_Continuous(t−1)"),
    ("TC rate",
     "5 bps = 0.05% round-trip per full flip. "
     "Full flip +1→−1 at $9,000/MT = $4.50/MT. Entry 0→+1 at $9,000/MT = $2.25/MT. "
     "The /2 reflects a half-spread charge per position unit moved."),
    ("Convention",
     "Matches exactly the dashboard's TC formula: pnl -= chg × (tc_bps/10000/2) × F1_price. "
     "TC columns are shown in BLUE in the Tradebook sheet."),
    ("Net Sharpe",
     "Sharpe computed on Net_Return on active days (|Position|>0). "
     "Same formula as Gross: mean/std × √252."),
    ("", ""),
    ("─── WARMUP / TIME-SERIES INTEGRITY ───────────────────────────────────────────────", ""),
    ("No row deletion",
     "Every row from the data start date is retained; the date index is never truncated. "
     "Date, F1_Continuous and Price_Change are populated from day 1."),
    ("Warmup blanking",
     "Indicator, Signal, Position, PnL and TC cells are left BLANK (empty, not 0) until "
     "their rolling/EWMA window is fully initialised. A blank means 'undefined' (no trade); "
     "a 0 means a genuine flat position inside an initialised signal (e.g. a value threshold band)."),
    ("Metrics on warmup",
     "Warmup rows contribute nothing: Sharpe uses active days only, annualised return drops "
     "undefined days, and CumPnL begins at the first initialised day."),
]

# ── Metrics helper ────────────────────────────────────────────────────────────
def _tail(pos_raw: pd.Series, f1c_s: pd.Series, tc_bps: int = 5):
    """Core PnL block. pos_raw carries NaN during warmup (undefined) and real 0
    for genuine flat days. Returns blanked (NaN-in-warmup) standard columns."""
    idx  = pos_raw.index.intersection(f1c_s.index)
    pos  = pos_raw.reindex(idx)
    warm = pos.isna()
    f    = f1c_s.reindex(idx)
    posf = pos.fillna(0.0)
    gp   = (posf * f.diff()).where(~warm)
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp / f.shift(1)).replace([np.inf, -np.inf], np.nan)
    chg = posf.diff().abs()
    if len(posf): chg.iloc[0] = abs(posf.iloc[0])
    tc  = (chg * (tc_bps / 10000.0 / 2.0) * f).where(~warm)
    net = gp - tc
    with np.errstate(invalid="ignore", divide="ignore"):
        nr = (net / f.shift(1)).replace([np.inf, -np.inf], np.nan)
    return dict(idx=idx, pos=pos, warm=warm, f=f, pc=f.diff(),
                gp=gp, gr=gr, gcum=gp.cumsum(),
                tc=tc, net=net, nr=nr, ncum=net.cumsum())

def metrics_both(pos_raw: pd.Series, f1c_s: pd.Series, tc_bps: int = 5):
    t    = _tail(pos_raw, f1c_s, tc_bps)
    pos, warm, f, idx = t["pos"], t["warm"], t["f"], t["idx"]
    defined = ~warm
    p    = pos.fillna(0.0)

    def _stats(pnl_s, ret):
        cum = pnl_s.cumsum()
        act = ret[(p != 0) & defined].dropna()
        sh  = float(act.mean() / act.std(ddof=1) * np.sqrt(252)) if len(act) > 20 else np.nan
        ar  = float(ret.dropna().mean() * 252 * 100) if ret.notna().any() else np.nan
        dd  = float((cum - cum.cummax()).min())
        return sh, ar, dd, float(pnl_s.sum()), cum, ret

    gsh, gar, gdd, gtp, gcum, gret = _stats(t["gp"], t["gr"])
    nsh, nar, ndd, ntp, ncum, nret = _stats(t["net"], t["nr"])
    n_def    = int(defined.sum())
    n_active = int(((p != 0) & defined).sum()); n_flat = int(((p == 0) & defined).sum())
    n_warm   = int(warm.sum())
    return dict(
        gsh=gsh, gar=gar, gdd=gdd, gtp=gtp, gcum=gcum, gret=gret,
        nsh=nsh, nar=nar, ndd=ndd, ntp=ntp, ncum=ncum, nret=nret,
        tc_cost=t["tc"], gross_pnl=t["gp"], net_pnl=t["net"],
        n_active=n_active, n_flat=n_flat, n_warmup=n_warm,
        n_long=int(((p > 0) & defined).sum()), n_short=int(((p < 0) & defined).sum()),
        pct_active=100.0*n_active/n_def if n_def else 0.0,
        pct_flat=100.0*n_flat/n_def if n_def else 0.0,
        start=idx[0].date(), end=idx[-1].date(), n_days=len(idx),
        pos=p, f=f,
    )

def write_perf_both(ws, m: dict, label: str, extra=None):
    _sw(ws); _cw(ws,1,38); _cw(ws,2,20); _cw(ws,3,20)
    ws.cell(row=1, column=1, value=f"Performance — {label}").font = TTL_FONT
    ws.cell(row=1, column=1).fill = DSC_FILL
    ws.merge_cells("A1:C1")
    for j, h in enumerate(["Metric","Gross (0 bps)","Net (5 bps TC)"], start=1):
        _hdr(ws, 2, j, h)
    rows = [
        ("Date Range",        f"{m['start']} → {m['end']}", ""),
        ("Total Days",        m['n_days'], ""),
        ("Active Days",       m['n_active'], ""),
        ("Long Days",         m['n_long'], ""),
        ("Short Days",        m['n_short'], ""),
        ("Flat Days (initialised)", m['n_flat'], ""),
        ("Warmup Days (blank)", m.get('n_warmup', 0), ""),
        ("% Active (of initialised)", f"{m['pct_active']:.1f}%", ""),
        ("", "", ""),
        ("Sharpe (annlzd, active days)", f"{m['gsh']:.4f}" if not np.isnan(m['gsh']) else "N/A",
                                         f"{m['nsh']:.4f}" if not np.isnan(m['nsh']) else "N/A"),
        ("Ann Return %",      f"{m['gar']:.2f}%", f"{m['nar']:.2f}%"),
        ("Max Drawdown USD/MT", f"${m['gdd']:.1f}", f"${m['ndd']:.1f}"),
        ("Total PnL USD/MT",  f"${m['gtp']:.1f}", f"${m['ntp']:.1f}"),
        ("", "", ""),
        ("Sharpe formula",    "mean(ret_active)/std(ret_active)×√252", "same"),
        ("Ann Return formula","mean(ret_all_days)×252×100", "same"),
        ("TC formula",        "N/A",
         "TC_Cost(t)=|ΔPos|×(5/10000/2)×F1_Cont(t)"),
    ]
    if extra:
        rows += [("","","")] + extra
    for i, row_data in enumerate(rows, start=3):
        alt = (i % 2 == 0)
        for j2, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=j2, value=val)
            c.font = KEY_FONT if j2 == 1 else DSC_FONT
            c.fill = D2_FILL if alt else D1_FILL

def write_tb_sheet(ws, data: pd.DataFrame, tc_start_col: int = None):
    """Write tradebook. tc_start_col: column index where TC/Net columns begin (shown in blue)."""
    _sw(ws)
    ws.freeze_panes = ws.cell(row=2, column=1)
    cols = list(data.columns)
    for j, col in enumerate(cols, start=1):
        _hdr(ws, 1, j, col)
        _cw(ws, j, max(len(col)+2, 12))
    _cw(ws, 1, 13)
    for i, (date, row) in enumerate(data.iterrows(), start=2):
        alt = (i % 2 == 0)
        ws.cell(row=i, column=1, value=date.date() if hasattr(date,'date') else date).font = DAT_FONT
        ws.cell(row=i, column=1).fill = D2_FILL if alt else D1_FILL
        for j, col in enumerate(cols[1:], start=2):
            v = row[col]
            if isinstance(v, float) and not np.isnan(v):
                v = round(v, 6)
            hl = None
            if col == "Position":
                hl = "pos" if isinstance(v,(int,float)) and v > 0 else ("neg" if isinstance(v,(int,float)) and v < 0 else None)
            in_tc = tc_start_col is not None and j >= tc_start_col
            if in_tc:
                _dat_tc(ws, i, j, v if not (isinstance(v,float) and np.isnan(v)) else None, alt=alt)
            else:
                _dat(ws, i, j, v if not (isinstance(v,float) and np.isnan(v)) else None, alt=alt, hl=hl)

def _tc_cols(pos: pd.Series, f1c_s: pd.Series, gross_pnl: pd.Series, tc_bps=5):
    """Return (tc_cost, net_pnl, net_ret, net_cum) on gross_pnl's index.
    Warmup rows (Position NaN) are left blank; for fully-defined OOS positions
    nothing is blanked, so OOS output is unchanged."""
    idx  = gross_pnl.index
    pos_i = pos.reindex(idx)
    warm = pos_i.isna()
    p    = pos_i.fillna(0.0)
    f    = f1c_s.reindex(idx)
    chg  = p.diff().abs()
    if len(p): chg.iloc[0] = abs(p.iloc[0])
    tc   = (chg * (tc_bps / 10000.0 / 2.0) * f).where(~warm)
    net  = gross_pnl - tc
    with np.errstate(invalid="ignore", divide="ignore"):
        nr  = (net / f.shift(1)).replace([np.inf,-np.inf], np.nan)
    return tc.round(4), net.round(4), (nr*100).round(6), net.cumsum().round(4)

# ─────────────────────────────────────────────────────────────────────────────
# 1. MOMENTUM IS — MA(35,43) Lag-1
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1/4] MOMENTUM…")
print("  MA(35,43) IS…")

def _ma_pos(s, l, f1r, lag1=True):
    # NaN during warmup (no .fillna): callers that need a flat 0 apply it explicitly.
    sig = np.sign(f1r.rolling(s).mean() - f1r.rolling(l).mean())
    return sig.shift(1) if lag1 else sig

def _ma_tb(s, l, f1r, f1c, lag1=True):
    ma_s = f1r.rolling(s).mean()
    ma_l = f1r.rolling(l).mean()
    diff = ma_s - ma_l
    sig  = np.sign(diff)
    pos  = sig.shift(1) if lag1 else sig          # NaN warmup (blank), no fillna
    idx  = pos.index.intersection(f1c.index)
    pos  = pos.reindex(idx); fc = f1c.reindex(idx)
    gp   = pos * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp / fc.shift(1)).replace([np.inf,-np.inf], np.nan)
    tc, np_, nr, nc = _tc_cols(pos, f1c, gp)
    return pd.DataFrame({
        "Date": idx,
        f"F1_Raw": f1r.reindex(idx).round(2),
        f"MA_{s}": ma_s.reindex(idx).round(4),
        f"MA_{l}": ma_l.reindex(idx).round(4),
        "MA_Diff": diff.reindex(idx).round(4),
        "Raw_Signal": np.sign(diff.reindex(idx)),
        "Position": pos.values,
        "F1_Continuous": fc.round(2),
        "Price_Change": fc.diff().round(4),
        "Gross_PnL_USD": gp.round(4),
        "Gross_Return_pct": (gr*100).round(6),
        "Gross_CumPnL": gp.cumsum().round(4),
        "TC_Cost_5bps_USD": tc.values,
        "Net_PnL_5bps_USD": np_.values,
        "Net_Return_5bps_pct": nr.values,
        "Net_CumPnL_5bps": nc.values,
    }, index=idx)

ma_tb = _ma_tb(35, 43, f1r, f1c)
m_ma  = metrics_both(ma_tb["Position"].rename(index=dict(zip(ma_tb.index, ma_tb.index))).set_axis(ma_tb.index), f1c)
# Simpler: recompute from position series
pos_ma35 = _ma_pos(35, 43, f1r)
m_ma = metrics_both(pos_ma35, f1c)

def write_mom_ma_is():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,30); _cw(ws_d,2,90)
    _title(ws_d, 1, "MOMENTUM — MA(35,43) Lag-1  [IS Full History]")
    rows = [
        ("Strategy",    "MA Crossover MA(35,43) Lag-1"),
        ("Tab",         "Tab 7: Momentum Signals — Section 2 IS"),
        ("Signal",      "MA_Diff(t) = MA_35(t) − MA_43(t)   |   Position(t) = sign(MA_Diff(t−1))"),
        ("Why (35,43)?","Max-IS-Sharpe pair from a scan over 5≤m<n≤200 MA pairs on full history. "
                        "Realised Sharpe is on the Performance sheet."),
        ("Why Lag-1?",  "Trend signal is slow-moving; no intraday urgency. Lag-1 = realistic execution."),
        ("PnL",         "Gross_PnL(t) = Position(t) × [F1_Continuous(t) − F1_Continuous(t−1)] [USD/MT]"),
        ("Data source", "LME_Copper_Rolling_F1_v2.csv"),
        ("Flat zone",   "None — always ±1 once MA_43 initialised (day 43). 0 only in first 42-day warmup."),
        ("Assumptions", "Binary ±1 position. No leverage. F1_raw for signal, F1_continuous for PnL."),
    ]
    rows += TC_DESC_ROWS
    for r, (k, v) in enumerate(rows, start=2): _desc(ws_d, r, k, v)
    ws_t = wb.create_sheet("Tradebook")
    write_tb_sheet(ws_t, ma_tb, tc_start_col=13)
    ws_p = wb.create_sheet("Performance")
    write_perf_both(ws_p, m_ma, "MA(35,43) Lag-1 IS")
    wb.save(f"{TBDIR}/momentum/IS/MOM_IS_MA35_43_Lag1.xlsx")
    print("    Saved MOM_IS_MA35_43_Lag1.xlsx")

write_mom_ma_is()

# ─────────────────────────────────────────────────────────────────────────────
# 1b. MOMENTUM IS — CTA Baz-Granger
# ─────────────────────────────────────────────────────────────────────────────
print("  CTA Baz-Granger IS…")

def _ew(s, com): return s.ewm(com=com, adjust=False).mean()

def _cta_full(f1r, f1c):
    pv = f1r.rolling(63).std()
    frames = {}; us = []
    for i, (sk, lk) in enumerate([(8,24),(16,48),(32,96)], 1):
        ews = _ew(f1r, sk-1); ewl = _ew(f1r, lk-1)
        x = ews - ewl; y = x / pv; sy = y.rolling(252).std(); z = y / sy
        with np.errstate(invalid="ignore"):
            u = z * np.exp(-z**2/4) / 0.89
        frames[f"EWMA_S{sk}"]          = ews.round(4)
        frames[f"EWMA_L{lk}"]          = ewl.round(4)
        frames[f"x{i}_(EWMA_S-L)"]     = x.round(4)
        frames[f"Sigma63_F1raw"]        = pv.round(4) if i==1 else None
        frames[f"y{i}_(x/σ63)"]        = y.round(6)
        frames[f"Sigma252_y{i}"]        = sy.round(6)
        frames[f"z{i}_(y/σ252)"]        = z.round(6)
        frames[f"u{i}_(z·e^-z²/4/0.89)"] = pd.Series(u.values, index=f1r.index).round(6)
        frames[f"Sig{i}_(sign_u{i})"]  = pd.Series(np.sign(u.values), index=f1r.index)
        us.append(u.values)
    frames = {k:v for k,v in frames.items() if v is not None}
    u_mean = np.nanmean(np.stack(us, axis=1), axis=1)
    sig_avg = pd.Series(np.sign(u_mean), index=f1r.index)
    pos = sig_avg.shift(1)                         # NaN until σ63 then σ252 of y init (~314d)
    idx = pos.index.intersection(f1c.index)
    pos = pos.reindex(idx); fc = f1c.reindex(idx)
    gp  = pos * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp / fc.shift(1)).replace([np.inf,-np.inf], np.nan)
    tc, np_, nr, nc = _tc_cols(pos, f1c, gp)
    out = {"Date": idx, "F1_Raw": f1r.reindex(idx).round(2)}
    for k, v in frames.items():
        if isinstance(v, pd.Series): out[k] = v.reindex(idx).values
    out["u_mean"] = u_mean[f1r.index.get_indexer(idx)]
    out["Raw_Signal_(sign_u_mean)"] = sig_avg.reindex(idx).values
    out["Position"]            = pos.values
    out["F1_Continuous"]       = fc.round(2).values
    out["Price_Change"]        = fc.diff().round(4).values
    out["Gross_PnL_USD"]       = gp.round(4).values
    out["Gross_Return_pct"]    = (gr*100).round(6).values
    out["Gross_CumPnL"]        = gp.cumsum().round(4).values
    out["TC_Cost_5bps_USD"]    = tc.values
    out["Net_PnL_5bps_USD"]    = np_.values
    out["Net_Return_5bps_pct"] = nr.values
    out["Net_CumPnL_5bps"]     = nc.values
    return pd.DataFrame(out, index=idx)

cta_tb   = _cta_full(f1r, f1c)
pos_cta  = pd.Series(cta_tb["Position"].values, index=cta_tb.index)
m_cta    = metrics_both(pos_cta, f1c)
# tc_start_col: first 2 cols (Date, F1_Raw) + ~25 intermediate cols → find it
_cta_tc_col = list(cta_tb.columns).index("TC_Cost_5bps_USD") + 1

def write_mom_cta_is():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, "MOMENTUM — CTA Baz-Granger (3-timescale)  [IS Full History]")
    rows = [
        ("Strategy",    "CTA Baz-Granger 3-timescale Lag-1 (Baz et al. 2015 Eqs 29-33)"),
        ("Tab",         "Tab 7: Momentum Signals — Section 2 IS"),
        ("Timescales",  "(S,L) ∈ {(8,24),(16,48),(32,96)} — EWMA slow-fast pairs"),
        ("Step 1: x",   "x_i(t) = EWMA(F1_raw, com=S−1) − EWMA(F1_raw, com=L−1)"),
        ("Step 2: y",   "y_i(t) = x_i(t) / σ_63(t)   [σ_63 = 63-day rolling std of F1_raw]"),
        ("Step 3: z",   "z_i(t) = y_i(t) / σ_252_y(t)   [σ_252_y = 252-day rolling std of y_i]"),
        ("Step 4: u",   "u_i(t) = z_i(t) × exp(−z_i(t)²/4) / 0.89   [response function, saturates at extremes]"),
        ("Step 5",      "u_mean(t) = mean(u1,u2,u3). Position(t) = sign(u_mean(t−1)) [Lag-1]"),
        ("EWMA conv",   "com=n−1 → alpha=1/n. NOT span=n (which gives alpha=2/(n+1))."),
        ("0.89 norm",   "z·exp(−z²/4) peaks at z=√2 ≈ 0.8576. Dividing by 0.89 scales peak response ≈ 1."),
        ("Warmup",      "~314 days: σ_63 of price needs 63 days; then σ_252 of y needs a further 252 days "
                        "of the y series. x/y/z/u/u_mean and Position are BLANK until then; Date and price "
                        "columns are populated from day 1 (no rows dropped)."),
        ("Assumptions", "Binary ±1 once initialised; no threshold flat zone. Columns shown per timescale for full audit."),
    ]
    rows += TC_DESC_ROWS
    for r, (k, v) in enumerate(rows, start=2): _desc(ws_d, r, k, v)
    ws_t = wb.create_sheet("Tradebook")
    write_tb_sheet(ws_t, cta_tb, tc_start_col=_cta_tc_col)
    ws_p = wb.create_sheet("Performance")
    write_perf_both(ws_p, m_cta, "CTA Baz-Granger IS")
    wb.save(f"{TBDIR}/momentum/IS/MOM_IS_CTA_BazGranger_Lag1.xlsx")
    print("    Saved MOM_IS_CTA_BazGranger_Lag1.xlsx")

write_mom_cta_is()

# ─────────────────────────────────────────────────────────────────────────────
# 1c. MOMENTUM IS — Anchors EW (MA(10,25), MA(35,43), MA(63,100))
# ─────────────────────────────────────────────────────────────────────────────
print("  Anchors EW IS…")

def _anc_tb(f1r, f1c):
    p1 = _ma_pos(10, 25, f1r);  p2 = _ma_pos(35, 43, f1r);  p3 = _ma_pos(63, 100, f1r)
    port = (p1 + p2 + p3) / 3.0
    idx  = port.index.intersection(f1c.index)
    p1=p1.reindex(idx); p2=p2.reindex(idx); p3=p3.reindex(idx)
    port=port.reindex(idx); fc=f1c.reindex(idx)
    ma1s=f1r.rolling(10).mean().reindex(idx); ma1l=f1r.rolling(25).mean().reindex(idx)
    ma2s=f1r.rolling(35).mean().reindex(idx); ma2l=f1r.rolling(43).mean().reindex(idx)
    ma3s=f1r.rolling(63).mean().reindex(idx); ma3l=f1r.rolling(100).mean().reindex(idx)
    gp = port * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp / fc.shift(1)).replace([np.inf,-np.inf], np.nan)
    tc, np_, nr, nc = _tc_cols(port, f1c, gp)
    return pd.DataFrame({
        "Date": idx,
        "F1_Raw": f1r.reindex(idx).round(2),
        "MA_10": ma1s.round(4),"MA_25": ma1l.round(4),
        "MA_Diff_10_25": (ma1s-ma1l).round(4),"Signal_MA1025": p1.values,
        "MA_35": ma2s.round(4),"MA_43": ma2l.round(4),
        "MA_Diff_35_43": (ma2s-ma2l).round(4),"Signal_MA3543": p2.values,
        "MA_63": ma3s.round(4),"MA_100": ma3l.round(4),
        "MA_Diff_63_100": (ma3s-ma3l).round(4),"Signal_MA63100": p3.values,
        "Port_Position_EW": port.round(6),
        "F1_Continuous": fc.round(2),
        "Price_Change": fc.diff().round(4),
        "Gross_PnL_USD": gp.round(4),
        "Gross_Return_pct": (gr*100).round(6),
        "Gross_CumPnL": gp.cumsum().round(4),
        "TC_Cost_5bps_USD": tc.values,
        "Net_PnL_5bps_USD": np_.values,
        "Net_Return_5bps_pct": nr.values,
        "Net_CumPnL_5bps": nc.values,
    }, index=idx)

anc_tb = _anc_tb(f1r, f1c)
pos_anc_ew = pd.Series(anc_tb["Port_Position_EW"].values, index=anc_tb.index)
m_anc_ew   = metrics_both(pos_anc_ew, f1c)

def write_mom_anchors_is():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,30); _cw(ws_d,2,90)
    _title(ws_d, 1, "MOMENTUM — Anchors EW: MA(10,25)+MA(35,43)+MA(63,100) Lag-1  [IS Full History]")
    rows = [
        ("Strategy",    "3 Structural Anchor MAs — Equal-Weight Composite, Lag-1"),
        ("Tab",         "Tab 7: Momentum Signals — Anchors + IS-Opt Weights (IS comparison baseline)"),
        ("Anchors",     "MA(10,25) = Fast. MA(35,43) = Medium (also standalone best). MA(63,100) = Slow."),
        ("Why anchors?","These are structural, not data-fitted: fast/medium/slow timescales chosen a priori. "
                        "In the OOS walk-forward the weights are IS-optimised per window; "
                        "this IS tradebook shows the EW baseline (1/3 each anchor)."),
        ("Signal_i",    "Signal_i(t) = sign(MA_fast(t−1) − MA_slow(t−1))  [Lag-1 per anchor]"),
        ("Port_Position_EW","(Signal_MA1025 + Signal_MA3543 + Signal_MA63100) / 3  "
                        "[ranges −1 to +1; fractional when anchors disagree]"),
        ("PnL",         "Gross_PnL(t) = Port_Position_EW(t) × [F1_Continuous(t) − F1_Continuous(t−1)]"),
        ("OOS version", "See MOM_OOS_Anchors_ISopt_WalkForward.xlsx for the IS-opt weight walk-forward; "
                        "per-window OOS Sharpes are in that file's OOS_Summary sheet."),
        ("Note",        "The optimizer tends to concentrate weight on MA(35,43), so IS-opt ≈ standalone MA(35,43). "
                        "EW IS Sharpe shown here is the unweighted baseline."),
        ("Assumptions", "Binary ±1 per anchor; fractional composite. F1_raw signal, F1_continuous PnL."),
    ]
    rows += TC_DESC_ROWS
    for r, (k, v) in enumerate(rows, start=2): _desc(ws_d, r, k, v)
    ws_t = wb.create_sheet("Tradebook")
    tc_col = list(anc_tb.columns).index("TC_Cost_5bps_USD") + 1
    write_tb_sheet(ws_t, anc_tb, tc_start_col=tc_col)
    ws_p = wb.create_sheet("Performance")
    # Per-anchor breakdown too
    m1 = metrics_both(_ma_pos(10,25,f1r), f1c)
    m3 = metrics_both(_ma_pos(63,100,f1r), f1c)
    write_perf_both(ws_p, m_anc_ew, "Anchors EW IS")
    # Add per-anchor row to Performance
    for r2, (lbl, sh_g, sh_n) in enumerate([
        ("MA(10,25) standalone gross Sharpe", f"{m1['gsh']:.4f}", f"{m1['nsh']:.4f}"),
        ("MA(35,43) standalone gross Sharpe", f"{m_ma['gsh']:.4f}", f"{m_ma['nsh']:.4f}"),
        ("MA(63,100) standalone gross Sharpe", f"{m3['gsh']:.4f}", f"{m3['nsh']:.4f}"),
    ], start=22):
        for j2, v in enumerate([lbl, sh_g, sh_n], start=1):
            _dat(ws_p, r2, j2, v, alt=(r2%2==0))
    wb.save(f"{TBDIR}/momentum/IS/MOM_IS_Anchors_EW_Lag1.xlsx")
    print("    Saved MOM_IS_Anchors_EW_Lag1.xlsx")

write_mom_anchors_is()

# ─────────────────────────────────────────────────────────────────────────────
# 1d. MOMENTUM OOS — MA(35,43) walk-forward
# ─────────────────────────────────────────────────────────────────────────────
print("  MA(35,43) OOS walk-forward…")

def _oos_wf_tradebook(path, title, desc_extra, signal_fn, extra_pos_cols_fn=None):
    """
    Generic OOS walk-forward tradebook builder.
    signal_fn(oos_idx) -> (pos_series, extra_col_dict)
    """
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, title)
    for r, (k, v) in enumerate(desc_extra, start=2): _desc(ws_d, r, k, v)

    ws_s = wb.create_sheet("OOS_Summary")
    _sw(ws_s)
    s_hdrs = ["OOS_Year","IS_Period","N_Days",
              "Gross_Sharpe","Net_5bps_Sharpe",
              "Gross_AnnRet%","Net_AnnRet%","Max_DD_Gross","Max_DD_Net","Pct_Active%"]
    for j, h in enumerate(s_hdrs, start=1):
        _hdr(ws_s, 1, j, h); _cw(ws_s, j, 15)

    ws_t = wb.create_sheet("Tradebook_All_Windows")
    _sw(ws_t)
    t_base = ["OOS_Window","IS_Period","Date","F1_Continuous","Price_Change","Position",
              "Gross_PnL_USD","Gross_Return_pct","Gross_CumPnL_in_window",
              "TC_Cost_5bps_USD","Net_PnL_5bps_USD","Net_Return_5bps_pct","Net_CumPnL_5bps_in_window"]
    for j, h in enumerate(t_base, start=1):
        _hdr(ws_t, 1, j, h); _cw(ws_t, j, 17)

    idx_all = f1r.index.intersection(f1c.index); T = len(idx_all)
    oos_s = IS_W; r_tb = 2; r_sum = 2; summary_rows = []
    while oos_s < T:
        oos_e = min(oos_s + OOS_W, T)
        is_s  = oos_s - IS_W
        partial = (oos_e - oos_s) < OOS_W
        oos_yr  = str(idx_all[oos_s].year) + ("*" if partial else "")
        is_lbl  = f"{idx_all[is_s].year}–{idx_all[oos_s-1].year}"
        oos_idx = idx_all[oos_s:oos_e]

        pos_o, extra = signal_fn(oos_idx, idx_all, is_s, oos_s, oos_e)
        fc_o   = f1c.reindex(oos_idx)
        gp_o   = pos_o * fc_o.diff()
        with np.errstate(invalid="ignore", divide="ignore"):
            gr_o = (gp_o / fc_o.shift(1)).replace([np.inf,-np.inf], np.nan)
        gc_o   = gp_o.cumsum()
        tc_o, np_o, nr_o, nc_o = _tc_cols(pos_o, f1c, gp_o)

        def _sh(ret, pos):
            act = ret[pos.abs() > 0].dropna()
            return float(act.mean()/act.std(ddof=1)*np.sqrt(252)) if len(act)>5 else np.nan
        gsh_o = _sh(gr_o, pos_o); nsh_o = _sh(nr_o, pos_o)
        gar_o = float(gr_o.dropna().mean()*252*100)
        nar_o = float(nr_o.dropna().mean()*252*100)
        gdd_o = float((gc_o - gc_o.cummax()).min())
        ndd_o = float((nc_o - nc_o.cummax()).min())
        pct_a = 100.0*(pos_o.abs()>0).sum()/len(oos_idx)

        def _r(x): return round(x,4) if not np.isnan(x) else "N/A"
        summary_rows.append([oos_yr, is_lbl, len(oos_idx),
                              _r(gsh_o), _r(nsh_o),
                              round(gar_o,2), round(nar_o,2),
                              round(gdd_o,2), round(ndd_o,2), round(pct_a,1)])
        for ki, date in enumerate(oos_idx):
            alt = (r_tb % 2 == 0)
            pv  = pos_o.iloc[ki]
            hl  = "pos" if pv>0 else ("neg" if pv<0 else None)
            vals = [oos_yr, is_lbl, str(date.date()),
                    round(fc_o.iloc[ki],2),
                    round(fc_o.diff().iloc[ki],4) if ki>0 else None,
                    round(pv,6),
                    round(gp_o.iloc[ki],4),
                    round(gr_o.iloc[ki]*100,6) if not np.isnan(gr_o.iloc[ki]) else None,
                    round(gc_o.iloc[ki],4)]
            tc_vals = [round(tc_o.iloc[ki],4), round(np_o.iloc[ki],4),
                       round(nr_o.iloc[ki]*100,6) if not np.isnan(nr_o.iloc[ki]) else None,
                       round(nc_o.iloc[ki],4)]
            for j2, v in enumerate(vals, start=1):
                _dat(ws_t, r_tb, j2, v, alt=alt, hl=(hl if j2==6 else None))
            for j2, v in enumerate(tc_vals, start=len(vals)+1):
                _dat_tc(ws_t, r_tb, j2, v, alt=alt)
            r_tb += 1
        oos_s += OOS_W

    for r2, rd in enumerate(summary_rows, start=2):
        for j2, v in enumerate(rd, start=1):
            _dat(ws_s, r2, j2, v, alt=(r2%2==0))
    # Avg row
    gsh_all = [r[3] for r in summary_rows if r[3]!="N/A"]
    nsh_all = [r[4] for r in summary_rows if r[4]!="N/A"]
    r_avg = len(summary_rows)+2
    _dat(ws_s, r_avg, 1, "AVERAGE")
    _dat(ws_s, r_avg, 4, round(np.nanmean(gsh_all),4) if gsh_all else "N/A")
    _dat(ws_s, r_avg, 5, round(np.nanmean(nsh_all),4) if nsh_all else "N/A")
    r_pct = r_avg + 1
    _dat(ws_s, r_pct, 1, "% Positive Windows (Gross)")
    _dat(ws_s, r_pct, 4, f"{100*sum(1 for s in gsh_all if s>0)/len(gsh_all):.1f}%" if gsh_all else "N/A")
    _dat(ws_s, r_pct, 5, f"{100*sum(1 for s in nsh_all if s>0)/len(nsh_all):.1f}%" if nsh_all else "N/A")

    wb.save(path); print(f"    Saved: {os.path.basename(path)}")

# Signal fn for MA(35,43)
def _ma3543_sig(oos_idx, idx_all, is_s, oos_s, oos_e):
    pos = np.sign(f1r.rolling(35).mean() - f1r.rolling(43).mean()).shift(1).fillna(0)
    return pos.reindex(oos_idx).fillna(0), {}

_oos_wf_tradebook(
    f"{TBDIR}/momentum/OOS/MOM_OOS_MA35_43_WalkForward.xlsx",
    "MOMENTUM OOS — MA(35,43) Lag-1 Walk-Forward",
    [("Strategy","MA(35,43) Walk-Forward: IS=5yr, OOS=1yr, no re-optimisation"),
     ("Tab","Tab 7: Momentum Signals — Section 1 OOS"),
     ("MA state","Full-history MA carried into OOS (not restarted at window boundary)."),
     ("TC in summary","Gross_Sharpe = 0 bps. Net_5bps_Sharpe = 5 bps TC per unit of position change."),
     ("Partial window","Last window marked * if fewer than 252 OOS days. Included in averages."),
    ] + TC_DESC_ROWS,
    _ma3543_sig,
)

# ─────────────────────────────────────────────────────────────────────────────
# 1e. MOMENTUM OOS — Anchors + IS-Opt Weights walk-forward
# ─────────────────────────────────────────────────────────────────────────────
print("  Anchors + IS-Opt Weights OOS walk-forward…")

def _max_sharpe_w(ret_mat):
    def neg_sh(w):
        r = ret_mat @ w; act = r[r!=0]
        if len(act)<20: return 0.0
        sd = act.std(ddof=1)
        return -(act.mean()/sd*np.sqrt(252)) if sd>0 else 0.0
    best_val, best_w = np.inf, np.array([1/3,1/3,1/3])
    for w0 in [[1/3,1/3,1/3],[0.8,0.1,0.1],[0.1,0.8,0.1],[0.1,0.1,0.8],
               [1,0,0],[0,1,0],[0,0,1],[0.5,0.5,0],[0.5,0,0.5],[0,0.5,0.5]]:
        try:
            res = minimize(neg_sh,w0,method="SLSQP",bounds=[(0,1)]*3,
                           constraints=[{"type":"eq","fun":lambda w:w.sum()-1}],
                           options={"ftol":1e-9,"maxiter":500})
            if res.fun < best_val: best_val, best_w = res.fun, res.x
        except: pass
    w = np.clip(best_w,0,1); return w/w.sum()

def _anc_opt_sig(oos_idx, idx_all, is_s, oos_s, oos_e):
    # IS period: compute positions and IS returns for weight optimisation
    is_idx  = idx_all[is_s:oos_s]
    fc_is   = f1c.reindex(is_idx)
    ret_mat = []
    for ms, ml in [(10,25),(35,43),(63,100)]:
        p = _ma_pos(ms, ml, f1r).reindex(is_idx).fillna(0)
        with np.errstate(invalid="ignore", divide="ignore"):
            r = (p*fc_is.diff()/fc_is.shift(1)).replace([np.inf,-np.inf],np.nan).fillna(0)
        ret_mat.append(r.values)
    w = _max_sharpe_w(np.column_stack(ret_mat))
    # OOS composite position
    p1 = _ma_pos(10,25,f1r).reindex(oos_idx).fillna(0)
    p2 = _ma_pos(35,43,f1r).reindex(oos_idx).fillna(0)
    p3 = _ma_pos(63,100,f1r).reindex(oos_idx).fillna(0)
    port = w[0]*p1 + w[1]*p2 + w[2]*p3
    return port, {"w_fast":round(w[0],3),"w_med":round(w[1],3),"w_slow":round(w[2],3),
                  "p_MA1025":p1,"p_MA3543":p2,"p_MA63100":p3}

# Custom OOS walk-forward for Anchors (need extra columns)
def write_anchors_oos():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, "MOMENTUM OOS — Anchors + IS-Opt Weights Walk-Forward")
    desc_rows = [
        ("Strategy",    "3 Fixed Anchors [MA(10,25), MA(35,43), MA(63,100)] with IS-Optimised Weights"),
        ("Tab",         "Tab 7: Momentum Signals — Section 1 OOS, Anchors card"),
        ("IS Window",   "1260 days (~5yr). OOS Window: 252 days (~1yr). Non-overlapping OOS."),
        ("IS step",     "Per IS window: compute Lag-1 positions for each anchor on IS data. "
                        "Run max-Sharpe QP to find (w_fast, w_med, w_slow) summing to 1, all ≥0."),
        ("OOS step",    "Apply IS-optimised weights to OOS anchor positions → composite position."),
        ("Why anchors?","Structural choices: Fast=short-term trend, Medium=medium, Slow=long-term. "
                        "NOT data-fitted from IS scan → avoids pair-selection overfitting."),
        ("Result",      "Per-window OOS Sharpes are in the OOS_Summary sheet (authoritative). "
                        "The optimizer tends to concentrate weight on MA(35,43), making this near-equivalent to standalone."),
        ("QP problem",  "max_w  Sharpe(w1*r_MA1025 + w2*r_MA3543 + w3*r_MA63100)  s.t. sum(w)=1, w≥0. "
                        "Solved via SLSQP with 10 random starts to avoid local optima."),
        ("Columns",     "Per-window: w_fast, w_med, w_slow (IS-optimised weights). "
                        "Pos_MA1025, Pos_MA3543, Pos_MA63100 (individual anchor positions). "
                        "Port_Position = w·positions composite."),
        ("Assumptions", "Binary ±1 per anchor. Fractional composite. Lag-1. F1_continuous PnL. "
                        "Weights are CONSTRAINED: no shorting anchors (w≥0), no leverage (sum=1). "
                        "IS period data not restarted — full history MA state carries over."),
    ] + TC_DESC_ROWS
    for r, (k, v) in enumerate(desc_rows, start=2): _desc(ws_d, r, k, v)

    ws_s = wb.create_sheet("OOS_Summary")
    _sw(ws_s)
    s_h = ["OOS_Year","IS_Period","N_Days","w_fast_(MA1025)","w_med_(MA3543)","w_slow_(MA63100)",
           "Gross_Sharpe","Net_5bps_Sharpe","Gross_AnnRet%","Net_AnnRet%","Max_DD_Gross","Max_DD_Net"]
    for j, h in enumerate(s_h, start=1): _hdr(ws_s,1,j,h); _cw(ws_s,j,16)

    ws_t = wb.create_sheet("Tradebook_All_Windows")
    _sw(ws_t)
    t_h = ["OOS_Window","IS_Period","Date","F1_Continuous","Price_Change",
           "w_fast","w_med","w_slow",
           "Pos_MA1025","Pos_MA3543","Pos_MA63100","Port_Position",
           "Gross_PnL_USD","Gross_Return_pct","Gross_CumPnL",
           "TC_Cost_5bps_USD","Net_PnL_5bps_USD","Net_Return_5bps_pct","Net_CumPnL_5bps"]
    for j, h in enumerate(t_h, start=1): _hdr(ws_t,1,j,h); _cw(ws_t,j,16)
    TC_COL = t_h.index("TC_Cost_5bps_USD") + 1

    idx_all = f1r.index.intersection(f1c.index); T = len(idx_all)
    oos_s = IS_W; r_tb = 2; r_s = 2; summary_rows = []
    while oos_s < T:
        oos_e   = min(oos_s+OOS_W, T)
        is_s    = oos_s - IS_W
        partial = (oos_e-oos_s) < OOS_W
        oos_yr  = str(idx_all[oos_s].year) + ("*" if partial else "")
        is_lbl  = f"{idx_all[is_s].year}–{idx_all[oos_s-1].year}"
        oos_idx = idx_all[oos_s:oos_e]

        port, extra = _anc_opt_sig(oos_idx, idx_all, is_s, oos_s, oos_e)
        wf, wm, ws_w = extra["w_fast"], extra["w_med"], extra["w_slow"]
        p1o, p2o, p3o = extra["p_MA1025"], extra["p_MA3543"], extra["p_MA63100"]
        fc_o  = f1c.reindex(oos_idx)
        gp_o  = port * fc_o.diff()
        with np.errstate(invalid="ignore", divide="ignore"):
            gr_o = (gp_o/fc_o.shift(1)).replace([np.inf,-np.inf],np.nan)
        gc_o  = gp_o.cumsum()
        tc_o, np_o, nr_o, nc_o = _tc_cols(port, f1c, gp_o)

        def _sh(ret, pos):
            act = ret[pos.abs()>0].dropna()
            return float(act.mean()/act.std(ddof=1)*np.sqrt(252)) if len(act)>5 else np.nan
        gsh_o = _sh(gr_o, port); nsh_o = _sh(nr_o, port)
        gar_o = float(gr_o.dropna().mean()*252*100)
        nar_o = float(nr_o.dropna().mean()*252*100)
        gdd_o = float((gc_o-gc_o.cummax()).min())
        ndd_o = float((nc_o-nc_o.cummax()).min())
        def _r(x): return round(x,4) if not np.isnan(x) else "N/A"
        summary_rows.append([oos_yr, is_lbl, len(oos_idx), wf, wm, ws_w,
                              _r(gsh_o), _r(nsh_o), round(gar_o,2), round(nar_o,2),
                              round(gdd_o,2), round(ndd_o,2)])

        for ki, date in enumerate(oos_idx):
            alt = (r_tb%2==0)
            pv  = port.iloc[ki]
            vals = [oos_yr, is_lbl, str(date.date()),
                    round(fc_o.iloc[ki],2),
                    round(fc_o.diff().iloc[ki],4) if ki>0 else None,
                    wf, wm, ws_w,
                    round(p1o.iloc[ki],4), round(p2o.iloc[ki],4), round(p3o.iloc[ki],4),
                    round(pv,6),
                    round(gp_o.iloc[ki],4),
                    round(gr_o.iloc[ki]*100,6) if not np.isnan(gr_o.iloc[ki]) else None,
                    round(gc_o.iloc[ki],4)]
            tc_vals = [round(tc_o.iloc[ki],4), round(np_o.iloc[ki],4),
                       round(nr_o.iloc[ki]*100,6) if not np.isnan(nr_o.iloc[ki]) else None,
                       round(nc_o.iloc[ki],4)]
            for j2, v in enumerate(vals, start=1):
                _dat(ws_t, r_tb, j2, v, alt=alt, hl=("pos" if j2==12 and pv>0 else ("neg" if j2==12 and pv<0 else None)))
            for j2, v in enumerate(tc_vals, start=TC_COL):
                _dat_tc(ws_t, r_tb, j2, v, alt=alt)
            r_tb += 1
        oos_s += OOS_W

    for r2, rd in enumerate(summary_rows, start=2):
        for j2, v in enumerate(rd, start=1): _dat(ws_s, r2, j2, v, alt=(r2%2==0))
    gsh_all = [r[6] for r in summary_rows if r[6]!="N/A"]
    nsh_all = [r[7] for r in summary_rows if r[7]!="N/A"]
    r_avg = len(summary_rows)+2
    _dat(ws_s, r_avg, 1, "AVERAGE")
    _dat(ws_s, r_avg, 7, round(np.nanmean(gsh_all),4) if gsh_all else "N/A")
    _dat(ws_s, r_avg, 8, round(np.nanmean(nsh_all),4) if nsh_all else "N/A")

    wb.save(f"{TBDIR}/momentum/OOS/MOM_OOS_Anchors_ISopt_WalkForward.xlsx")
    print("    Saved MOM_OOS_Anchors_ISopt_WalkForward.xlsx")

write_anchors_oos()

# ─────────────────────────────────────────────────────────────────────────────
# 2. CARRY IS
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2/4] CARRY…")

def _carry_tb(raw_signal: pd.Series, f1c: pd.Series, same_day: bool,
              extra_input_cols: dict, formula_str: str):
    sig  = np.sign(raw_signal)
    # Timing convention (matches dashboard _carry_cum_pnl after the look-ahead fix):
    #   Same-Day = shift 1  -> trade at signal close(t), first return t->t+1.
    #   Lag-1    = shift 2  -> trade at next close(t+1), first return t+1->t+2.
    # Legacy shift-0 "same-day" earned a contemporaneous move = look-ahead (removed).
    pos  = sig.shift(1) if same_day else sig.shift(2)   # NaN warmup (blank), no fillna
    idx  = pos.index.intersection(f1c.index)
    pos  = pos.reindex(idx); fc = f1c.reindex(idx)
    gp   = pos * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp/fc.shift(1)).replace([np.inf,-np.inf], np.nan)
    tc_c, np_, nr, nc = _tc_cols(pos, f1c, gp)
    out = {"Date": idx}
    for k, v in extra_input_cols.items():
        out[k] = v.reindex(idx).round(4) if isinstance(v, pd.Series) else v
    out[f"Carry_Raw_({formula_str})"] = raw_signal.reindex(idx).round(6)
    out["Raw_Signal"]    = sig.reindex(idx).values
    out["Position"]      = pos.values
    out["F1_Continuous"] = fc.round(2)
    out["Price_Change"]  = fc.diff().round(4)
    out["Gross_PnL_USD"] = gp.round(4)
    out["Gross_Return_pct"] = (gr*100).round(6)
    out["Gross_CumPnL"]  = gp.cumsum().round(4)
    out["TC_Cost_5bps_USD"]    = tc_c.values
    out["Net_PnL_5bps_USD"]    = np_.values
    out["Net_Return_5bps_pct"] = nr.values
    out["Net_CumPnL_5bps"]     = nc.values
    return pd.DataFrame(out, index=idx)

CARRY_CASES = []

# V4 Carry Momentum 20d — 20-day change in roll yield (dashboard default carry leg)
if "F1" in crv.columns and "F2" in crv.columns:
    cm_base = ((crv["F1"]-crv["F2"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
    raw = (cm_base - cm_base.shift(20))                  # 20-day change in roll yield (NaN warmup kept)
    tb  = _carry_tb(raw, f1c, True,
                    {"F1_Price": crv["F1"], "F2_Price": crv["F2"],
                     "RollYield_(F1-F2)/F1": cm_base,
                     "RollYield_20d_ago": cm_base.shift(20)},
                    "Δ20d (F1-F2)/F1")
    CARRY_CASES.append(("V4_CarryMom20d_SameDay", tb,
        "V4 Carry Momentum 20d (Same-Day)", [
        ("Strategy",  "V4 Carry Momentum: sign of the 20-day change in the (F1-F2)/F1 roll yield. "
                      "Dashboard Tab 8 DEFAULT and portfolio carry leg."),
        ("Formula",   "RollYield(t) = (F1(t) − F2(t)) / F1(t);  Carry_Raw(t) = RollYield(t) − RollYield(t−20)"),
        ("Position",  "Position(t) = sign(Carry_Raw(t))  [Same-Day = shift 1, no look-ahead]"),
        ("Logic",     "A steepening curve (backwardation building) over ~20d signals physical tightening → go long."),
        ("Robustness","Examined against transaction-cost levels and an extra day's execution lag (shift-2); "
                      "1-day Δcarry and 60d horizons were compared. See the Performance sheet for realised figures."),
    ]))

# V1a (F1-F2)/F1 Same-Day — carry level signal (post look-ahead-fix)
if "F1" in crv.columns and "F2" in crv.columns:
    raw = ((crv["F1"]-crv["F2"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
    tb  = _carry_tb(raw, f1c, True, {"F1_Price": crv["F1"], "F2_Price": crv["F2"]}, "(F1-F2)/F1")
    CARRY_CASES.append(("V1a_(F1-F2)-F1_SameDay", tb,
        "V1 (F1-F2)/F1 Same-Day", [
        ("Strategy",  "V1 Roll Yield: (F1-F2)/F1 Same-Day — carry-level signal"),
        ("Formula",   "Carry_Raw(t) = (F1(t) − F2(t)) / F1(t)"),
        ("Position",  "Position(t) = sign(Carry_Raw(t))  [Same-Day = shift 1, no look-ahead]"),
        ("Timing",    "Same-Day uses shift 1 (trade at the signal close, first return t→t+1); a shift-0 "
                      "version would earn the contemporaneous move (look-ahead) and is not used."),
        ("No OOS?",   "Zero free parameters → IS=OOS by construction for the level signal."),
        ("TC note",   "Carry flips relatively often (every contango↔backwardation switch). "
                      "5 bps TC materially reduces Sharpe vs gross."),
    ]))

# V1b (F1-F3)/F1 Same-Day
if "F1" in crv.columns and "F3" in crv.columns:
    raw = ((crv["F1"]-crv["F3"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
    tb  = _carry_tb(raw, f1c, True, {"F1_Price": crv["F1"], "F3_Price": crv["F3"]}, "(F1-F3)/F1")
    CARRY_CASES.append(("V1b_(F1-F3)-F1_SameDay", tb, "V1 (F1-F3)/F1 Same-Day", [
        ("Formula",   "Carry_Raw(t) = (F1(t) − F3(t)) / F1(t)"),
        ("Position",  "Position(t) = sign(Carry_Raw(t))  [Same-Day]"),
        ("vs V1a",    "2-month spread; more stable, fewer flips, slightly lower TC drag."),
    ]))

# V1c Cash-3M Same-Day
if "cash_price" in cash_cu.columns and "3m_price" in cash_cu.columns:
    cp = cash_cu["cash_price"].dropna(); tm = cash_cu["3m_price"].dropna()
    idx_c = cp.index.intersection(tm.index)
    raw = ((cp.reindex(idx_c)-tm.reindex(idx_c))/cp.reindex(idx_c)).replace([np.inf,-np.inf],np.nan)
    tb  = _carry_tb(raw, f1c, True,
                    {"Cash_Price": cp, "3M_Price": tm}, "(Cash-3M)/Cash")
    CARRY_CASES.append(("V1c_(Cash-3M)-Cash_SameDay", tb, "V1 (Cash-3M)/Cash Same-Day", [
        ("Formula",   "Carry_Raw(t) = (Cash(t) − 3M(t)) / Cash(t)"),
        ("Data",      "LME Copper Cash and 3M prices from Metals Cash and 3M.xlsx."),
        ("PnL price", "F1_Continuous used for PnL (not Cash price) — consistent with all strategies."),
        ("Note",      "Cash vs F1 basis creates small PnL mismatch (conservative)."),
    ]))

# V2 Long Slope — best IS pair from 10
if "F3" in crv.columns:
    pairs = [(j,k) for j,k in [(3,15),(4,16),(5,17),(6,18),(7,19),(8,20),(9,21),(10,22),(11,23),(12,24)]
             if f"F{j}" in crv.columns and f"F{k}" in crv.columns]
    best_pair, best_sh = (3,15), -np.inf
    for j,k in pairs:
        raw_v2 = ((crv[f"F{j}"]-crv[f"F{k}"])/crv[f"F{k}"]).replace([np.inf,-np.inf],np.nan).dropna()
        pos_v2 = np.sign(raw_v2).reindex(f1c.index).fillna(0)
        fc_v2  = f1c.reindex(pos_v2.index)
        with np.errstate(invalid="ignore", divide="ignore"):
            r = (pos_v2*fc_v2.diff()/fc_v2.shift(1)).replace([np.inf,-np.inf],np.nan)
        act = r[pos_v2!=0].dropna()
        if len(act)>20:
            sh = float(act.mean()/act.std(ddof=1)*np.sqrt(252))
            if sh > best_sh: best_sh, best_pair = sh, (j,k)
    j,k = best_pair
    raw = ((crv[f"F{j}"]-crv[f"F{k}"])/crv[f"F{k}"]).replace([np.inf,-np.inf],np.nan)
    tb  = _carry_tb(raw, f1c, True,
                    {f"F{j}_Price": crv[f"F{j}"], f"F{k}_Price": crv[f"F{k}"]},
                    f"(F{j}-F{k})/F{k}")
    CARRY_CASES.append((f"V2_(F{j}-F{k})-F{k}_SameDay", tb,
        f"V2 Long Slope (F{j}-F{k})/F{k} Same-Day", [
        ("Formula",   f"Carry_Raw(t) = (F{j}(t) − F{k}(t)) / F{k}(t)"),
        ("IS scan",   f"10 pairs scanned: (3,15)…(12,24); the IS-highest-Sharpe pair (F{j}-F{k}) is saved. "
                      "Recompute all 10 to confirm the selection."),
        ("Logic",     "Mid-curve slope: F{j}>F{k} → backwardation at longer tenors → long."),
    ]))

# V3 Z-score Same-Day
if "F1" in crv.columns and "F2" in crv.columns:
    base   = ((crv["F1"]-crv["F2"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
    roll_m = base.rolling(252).mean(); roll_s = base.rolling(252).std()
    z_sc   = ((base-roll_m)/roll_s).replace([np.inf,-np.inf],np.nan)   # 251d warmup kept blank
    extra  = {"F1_Price": crv["F1"], "F2_Price": crv["F2"],
               "Roll_Yield_(F1-F2)/F1": base,
               "RY_252d_Mean": roll_m, "RY_252d_Std": roll_s, "Z_Score": z_sc}
    tb = _carry_tb(z_sc, f1c, True, extra, "Z_Score_of_(F1-F2)/F1")
    CARRY_CASES.append(("V3_Zscore_252d_SameDay", tb, "V3 Z-score (252d) Same-Day", [
        ("Formula",   "Z(t) = (Roll_Yield(t) − Mean_252(RY)) / Std_252(RY)"),
        ("Position",  "Position(t) = sign(Z(t))  [Same-Day]"),
        ("Warmup",    "First 251 days: Z undefined (rolling window). Position=0."),
        ("vs V1a",    "V1a = raw sign of carry level. V3 = Z-score of carry vs its own history. "
                      "Both binary but react differently to carry extremes vs trend."),
    ]))

for fname, tb, label, extra_desc in CARRY_CASES:
    pos_s = pd.Series(tb["Position"].values, index=tb.index)
    m     = metrics_both(pos_s, f1c)
    wb    = openpyxl.Workbook()
    ws_d  = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,32); _cw(ws_d,2,90)
    _title(ws_d, 1, f"CARRY SIGNAL — {label}  [IS Full History / No OOS]")
    all_rows = [
        ("Tab",     "Tab 8: Carry Signals — IS Full History"),
        ("No OOS",  "Zero free parameters → IS result = OOS result by construction. No walk-forward needed."),
        ("PnL",     "Gross_PnL(t) = Position(t) × [F1_Continuous(t) − F1_Continuous(t−1)]"),
    ] + extra_desc + TC_DESC_ROWS
    for r, (k, v) in enumerate(all_rows, start=2): _desc(ws_d, r, k, v)
    ws_t = wb.create_sheet("Tradebook")
    tc_col = list(tb.columns).index("TC_Cost_5bps_USD") + 1
    write_tb_sheet(ws_t, tb, tc_start_col=tc_col)
    ws_p = wb.create_sheet("Performance")
    write_perf_both(ws_p, m, label)
    wb.save(f"{TBDIR}/carry/IS/{fname}.xlsx")
    print(f"    Saved {fname}.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 3. VALUE IS
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3/4] VALUE…")

def _v1_tb(k, N, f1c, thr=0.10):
    col = f"F{k}"
    if col not in crv.columns: return None
    p   = crv[col].dropna()
    ma  = p.rolling(N, min_periods=N//2).mean()
    dev = ((p-ma)/ma).replace([np.inf,-np.inf],np.nan)          # warmup NaN kept (not dropped)
    sig = np.where(dev<-thr, 1.0, np.where(dev>thr, -1.0, 0.0))
    sig_s = pd.Series(sig, index=dev.index).where(dev.notna())  # warmup→blank; in-band flat stays 0
    pos   = sig_s.shift(1)                                       # NaN warmup (blank), no fillna
    idx   = pos.index.intersection(f1c.index)
    pos   = pos.reindex(idx); fc = f1c.reindex(idx)
    gp    = pos * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp/fc.shift(1)).replace([np.inf,-np.inf],np.nan)
    tc_c, np_, nr, nc = _tc_cols(pos, f1c, gp)
    return pd.DataFrame({
        "Date": idx,
        f"{col}_Price": crv[col].reindex(idx).round(2),
        f"MA_{N//252}yr_{col}_(N={N}d)": ma.reindex(idx).round(4),
        "Deviation_((Fk-MA)/MA)": dev.reindex(idx).round(6),
        "Threshold_Upper_(+10%)": 0.10,
        "Threshold_Lower_(-10%)": -0.10,
        "Raw_Signal_(+1_buy,-1_sell,0_flat)": sig_s.reindex(idx).values,
        "Position_(Lag-1)": pos.values,
        "F1_Continuous": fc.round(2),
        "Price_Change": fc.diff().round(4),
        "Gross_PnL_USD": gp.round(4),
        "Gross_Return_pct": (gr*100).round(6),
        "Gross_CumPnL": gp.cumsum().round(4),
        "TC_Cost_5bps_USD": tc_c.values,
        "Net_PnL_5bps_USD": np_.values,
        "Net_Return_5bps_pct": nr.values,
        "Net_CumPnL_5bps": nc.values,
    }, index=idx)

def _v2_tb(N, f1r, f1c):
    rev = f1r.shift(N) - f1r
    sig = np.sign(rev.replace([np.inf,-np.inf],np.nan))
    pos = sig.shift(1)                                    # NaN warmup (blank), no fillna
    idx = pos.index.intersection(f1c.index)               # full index, no row drop
    pos = pos.reindex(idx); fc = f1c.reindex(idx)
    gp  = pos * fc.diff()
    with np.errstate(invalid="ignore", divide="ignore"):
        gr = (gp/fc.shift(1)).replace([np.inf,-np.inf],np.nan)
    tc_c, np_, nr, nc = _tc_cols(pos, f1c, gp)
    yr  = N // 252
    return pd.DataFrame({
        "Date": idx,
        "F1_Raw_today": f1r.reindex(idx).round(2),
        f"F1_Raw_{yr}yr_ago_(t-{N}d)": f1r.shift(N).reindex(idx).round(2),
        f"Reversal_(F1[t-{N}]-F1[t])": rev.reindex(idx).round(4),
        "Raw_Signal_(sign_reversal)": np.sign(rev.reindex(idx)).values,
        "Position_(Lag-1)": pos.values,
        "F1_Continuous": fc.round(2),
        "Price_Change": fc.diff().round(4),
        "Gross_PnL_USD": gp.round(4),
        "Gross_Return_pct": (gr*100).round(6),
        "Gross_CumPnL": gp.cumsum().round(4),
        "TC_Cost_5bps_USD": tc_c.values,
        "Net_PnL_5bps_USD": np_.values,
        "Net_Return_5bps_pct": nr.values,
        "Net_CumPnL_5bps": nc.values,
    }, index=idx)

V1_SPECS = [
    (8, 1260,  "V1_F8_5yr_Lag1",
     "V1 F8·5yr·±10%·Lag-1",
     "Contract F8, 5yr (1260d) rolling mean. One of several F1–F15 × lookback configurations examined."),
    (12, 1260, "V1_F12_5yr_Lag1",
     "V1 F12·5yr·±10%·Lag-1",
     "Contract F12, 5yr (1260d) rolling mean — the tenor/lookback used by the NGL energy risk-premia paper."),
    (14, 1764, "V1_F14_7yr_Lag1",
     "V1 F14·7yr·±10%·Lag-1",
     "Contract F14, 7yr (1764d) rolling mean."),
    (14, 2520, "V1_F14_10yr_Lag1",
     "V1 F14·10yr·±10%·Lag-1",
     "Contract F14, 10yr (2520d) rolling mean. Longer lookback → fewer flat days, slower reversion signal."),
]

for k, N, fname, title, why_note in V1_SPECS:
    tb = _v1_tb(k, N, f1c)
    if tb is None: continue
    pos_s = pd.Series(tb["Position_(Lag-1)"].values, index=tb.index)
    m = metrics_both(pos_s, f1c)
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, f"VALUE SIGNAL — {title}  [IS Full History]")
    desc = [
        ("Tab",       "Tab 9: Value Signals — Section 2 IS"),
        ("Strategy",  f"V1 MA Reversion on F{k}. {N//252}yr ({N}d) rolling mean. ±10% threshold. Lag-1."),
        ("Formula",   f"Deviation(t) = (F{k}(t) − MA_{N}d(t)) / MA_{N}d(t)"),
        ("Signal",    "Signal(t) = +1 if Dev < −10%  |  −1 if Dev > +10%  |  0 if |Dev| ≤ 10%"),
        ("Position",  "Position(t) = Signal(t−1)  [Lag-1]"),
        (f"Why F{k}?", why_note),
        ("PnL price", f"F1_Continuous used for PnL (NOT F{k}). We trade the front-month."),
        ("Flat zone", "Inside the ±10% band Signal=0 (a genuine flat day, shown as 0). "
                      "Sharpe is computed on active days only."),
        ("Why ±10%?", "Threshold band avoids whipsaws near the mean; matches the ±10% band of the "
                      "NGL energy risk-premia paper."),
        ("Warmup",    f"min_periods={N//2}: Deviation, Signal and Position are BLANK for the first {N//2} "
                      f"rows (rolling mean undefined), then populated. Date and price columns run from day 1."),
        ("Same-Day?", "Same-Day entry (no lag) is empirically perverse for this signal; Lag-1 is used."),
    ] + TC_DESC_ROWS
    for r, (k2, v) in enumerate(desc, start=2): _desc(ws_d, r, k2, v)
    ws_t = wb.create_sheet("Tradebook")
    tc_col = list(tb.columns).index("TC_Cost_5bps_USD") + 1
    write_tb_sheet(ws_t, tb, tc_start_col=tc_col)
    ws_p = wb.create_sheet("Performance")
    write_perf_both(ws_p, m, title)
    wb.save(f"{TBDIR}/value/IS/{fname}.xlsx")
    print(f"    Saved {fname}.xlsx")

for N, tag in [(756,"3yr"),(2520,"10yr")]:
    tb = _v2_tb(N, f1r, f1c)
    pos_s = pd.Series(tb["Position_(Lag-1)"].values, index=tb.index)
    m = metrics_both(pos_s, f1c)
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, f"VALUE SIGNAL — V2 BG Reversal {tag}·Lag-1  [IS Full History]")
    desc = [
        ("Tab",      "Tab 9: Value Signals — Section 2 IS"),
        ("Strategy", f"V2 Baz-Granger Reversal. {tag} ({N}d) lookback. Lag-1. No flat zone."),
        ("Formula",  f"Reversal(t) = F1_raw(t−{N}) − F1_raw(t)"),
        ("Signal",   "Signal(t) = sign(Reversal(t)). +1 if price fell over N days → expect reversion."),
        ("Position", "Position(t) = Signal(t−1)  [Lag-1]"),
        ("No flat",  "Always ±1 when signal is defined. No threshold."),
        ("Lookback", "Performance across lookbacks (1/3/5/7/10yr) is non-monotonic; compare the IS "
                     "Performance sheets across the V2_BG_* files rather than assuming a ranking."),
        ("F1_raw",   "Unadjusted price for reversal signal. F1_continuous for PnL."),
        ("Warmup",   f"First {N} rows: F1_raw[t−{N}] undefined → Reversal, Signal and Position are BLANK. "
                     "Date and price columns run from day 1."),
    ] + TC_DESC_ROWS
    for r, (k2, v) in enumerate(desc, start=2): _desc(ws_d, r, k2, v)
    ws_t = wb.create_sheet("Tradebook")
    tc_col = list(tb.columns).index("TC_Cost_5bps_USD") + 1
    write_tb_sheet(ws_t, tb, tc_start_col=tc_col)
    ws_p = wb.create_sheet("Performance")
    write_perf_both(ws_p, m, f"V2 BG {tag} IS")
    wb.save(f"{TBDIR}/value/IS/V2_BG_{tag}_Lag1.xlsx")
    print(f"    Saved V2_BG_{tag}_Lag1.xlsx")

# ── VALUE OOS walk-forward ────────────────────────────────────────────────────
print("  Value OOS walk-forward (4 signals)…")

def _build_val_pos(variant, k, N):
    if variant=="v1":
        col=f"F{k}"
        if col not in crv.columns: return pd.Series(dtype=float)
        p=crv[col].dropna(); ma=p.rolling(N, min_periods=N//2).mean()
        dev=((p-ma)/ma).replace([np.inf,-np.inf],np.nan).dropna()
        sig=np.where(dev<-0.10,1.0,np.where(dev>0.10,-1.0,0.0))
        return pd.Series(sig,index=dev.index).shift(1).fillna(0)
    else:
        rev=f1r.shift(N)-f1r
        s=np.sign(rev.replace([np.inf,-np.inf],np.nan).dropna())
        return s.shift(1).fillna(0)

V_OOS = [
    ("V1_F8_5yr",  "v1",8, 1260,"V1 F8·5yr·Lag-1"),
    ("V1_F12_5yr", "v1",12,1260,"V1 F12·5yr·Lag-1"),
    ("V2_BG_3yr",  "v2",None,756,"V2 BG Reversal·3yr·Lag-1"),
    ("V2_BG_10yr", "v2",None,2520,"V2 BG Reversal·10yr·Lag-1"),
]

for fname, var, k_v, N_v, lbl in V_OOS:
    full_pos = _build_val_pos(var, k_v, N_v)
    if full_pos.empty: continue
    idx_all = full_pos.index.intersection(f1c.index); T = len(idx_all)
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,35); _cw(ws_d,2,90)
    _title(ws_d, 1, f"VALUE OOS — {lbl}  [Walk-Forward Validation]")
    desc = [
        ("Tab",    "Tab 9: Value Signals — Section 1 OOS"),
        ("IS=5yr", "1260 days. OOS=1yr, 252 days. Non-overlapping windows."),
        ("Fixed?", "YES — params fixed throughout. Tests IS-selected params on unseen data."),
        ("Params", f"variant={var}, k={k_v}, N={N_v}. No re-opt per window."),
        ("TC",     "Both Gross and Net 5bps Sharpe shown per window."),
    ] + TC_DESC_ROWS
    for r, (k2, v) in enumerate(desc, start=2): _desc(ws_d, r, k2, v)
    ws_s = wb.create_sheet("OOS_Summary")
    _sw(ws_s)
    for j, h in enumerate(["OOS_Year","IS_Period","N_Days",
                            "Gross_Sharpe","Net_5bps_Sharpe",
                            "Gross_AnnRet%","Net_AnnRet%","Pct_Active%"], start=1):
        _hdr(ws_s,1,j,h); _cw(ws_s,j,14)
    ws_t = wb.create_sheet("Tradebook_All_Windows")
    _sw(ws_t)
    t_h = ["OOS_Window","IS_Period","Date","F1_Continuous","Price_Change","Position",
           "Gross_PnL_USD","Gross_Return_pct","Gross_CumPnL",
           "TC_Cost_5bps_USD","Net_PnL_5bps_USD","Net_Return_5bps_pct","Net_CumPnL_5bps"]
    for j, h in enumerate(t_h, start=1): _hdr(ws_t,1,j,h); _cw(ws_t,j,17)
    TC_COL = t_h.index("TC_Cost_5bps_USD")+1
    oos_s=IS_W; r_tb=2; r_s=2; sum_rows=[]
    while oos_s<T:
        oos_e=min(oos_s+OOS_W,T); is_s_=oos_s-IS_W
        oos_yr=str(idx_all[oos_s].year)+("*" if (oos_e-oos_s)<OOS_W else "")
        is_lbl=f"{idx_all[is_s_].year}–{idx_all[oos_s-1].year}"
        oos_idx=idx_all[oos_s:oos_e]
        pos_o=full_pos.reindex(oos_idx).fillna(0); fc_o=f1c.reindex(oos_idx)
        gp_o=pos_o*fc_o.diff()
        with np.errstate(invalid="ignore", divide="ignore"):
            gr_o=(gp_o/fc_o.shift(1)).replace([np.inf,-np.inf],np.nan)
        gc_o=gp_o.cumsum()
        tc_o,np_o,nr_o,nc_o=_tc_cols(pos_o,f1c,gp_o)
        def _sh(ret,pos):
            act=ret[pos.abs()>0].dropna()
            return float(act.mean()/act.std(ddof=1)*np.sqrt(252)) if len(act)>5 else np.nan
        gsh_o=_sh(gr_o,pos_o); nsh_o=_sh(nr_o,pos_o)
        gar_o=float(gr_o.dropna().mean()*252*100)
        nar_o=float(nr_o.dropna().mean()*252*100)
        pct_a=100.0*(pos_o.abs()>0).sum()/len(oos_idx)
        def _r(x): return round(x,4) if not np.isnan(x) else "N/A"
        sum_rows.append([oos_yr,is_lbl,len(oos_idx),_r(gsh_o),_r(nsh_o),
                          round(gar_o,2),round(nar_o,2),round(pct_a,1)])
        for ki,date in enumerate(oos_idx):
            alt=(r_tb%2==0); pv=pos_o.iloc[ki]
            vals=[oos_yr,is_lbl,str(date.date()),round(fc_o.iloc[ki],2),
                  round(fc_o.diff().iloc[ki],4) if ki>0 else None,int(pv),
                  round(gp_o.iloc[ki],4),
                  round(gr_o.iloc[ki]*100,6) if not np.isnan(gr_o.iloc[ki]) else None,
                  round(gc_o.iloc[ki],4)]
            tc_vals=[round(tc_o.iloc[ki],4),round(np_o.iloc[ki],4),
                     round(nr_o.iloc[ki]*100,6) if not np.isnan(nr_o.iloc[ki]) else None,
                     round(nc_o.iloc[ki],4)]
            for j2,v in enumerate(vals,start=1):
                _dat(ws_t,r_tb,j2,v,alt=alt,hl=("pos" if j2==6 and pv>0 else ("neg" if j2==6 and pv<0 else None)))
            for j2,v in enumerate(tc_vals,start=TC_COL):
                _dat_tc(ws_t,r_tb,j2,v,alt=alt)
            r_tb+=1
        oos_s+=OOS_W
    for r2,rd in enumerate(sum_rows,start=2):
        for j2,v in enumerate(rd,start=1): _dat(ws_s,r2,j2,v,alt=(r2%2==0))
    gsh_all=[r[3] for r in sum_rows if r[3]!="N/A"]
    nsh_all=[r[4] for r in sum_rows if r[4]!="N/A"]
    r_avg=len(sum_rows)+2
    _dat(ws_s,r_avg,1,"AVERAGE")
    _dat(ws_s,r_avg,4,round(np.nanmean(gsh_all),4) if gsh_all else "N/A")
    _dat(ws_s,r_avg,5,round(np.nanmean(nsh_all),4) if nsh_all else "N/A")
    wb.save(f"{TBDIR}/value/OOS/{fname}_OOS_WalkForward.xlsx")
    print(f"    Saved {fname}_OOS_WalkForward.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# 4. PORTFOLIO
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4/4] PORTFOLIO…")

pos_mom   = _ma_pos(35, 43, f1r)                              # NaN warmup, full index
# Carry leg = Carry Momentum 20d, Same-Day = shift 1. Matches app.py portfolio carry leg.
_pf_cr_base = ((crv["F1"]-crv["F2"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
_pf_cr_mom  = (_pf_cr_base - _pf_cr_base.shift(20))           # NaN warmup kept
pos_carry = np.sign(_pf_cr_mom).shift(1).reindex(f1c.index)   # NaN warmup, full index
pos_val   = _build_val_pos("v1", 8, 1260).reindex(f1c.index)  # NaN warmup, full index

# IS display spans the FULL date axis (no rows dropped). The composite is BLANK until
# all three legs are initialised (slowest = Value, ~630d); Date/price populated from day 1.
pf_idx = f1c.index
pm = pos_mom.reindex(pf_idx); pc = pos_carry.reindex(pf_idx); pv = pos_val.reindex(pf_idx)
pp = (pm + pc + pv) / 3.0
warm_pf = pp.isna()
pf = f1c.reindex(pf_idx)
pmf, pcf, pvf, ppf = pm.fillna(0), pc.fillna(0), pv.fillna(0), pp.fillna(0)
gp_m=(pmf*pf.diff()).where(~warm_pf); gp_c=(pcf*pf.diff()).where(~warm_pf)
gp_v=(pvf*pf.diff()).where(~warm_pf); gp_p=(ppf*pf.diff()).where(~warm_pf)
with np.errstate(invalid="ignore", divide="ignore"):
    gr_p=(gp_p/pf.shift(1)).replace([np.inf,-np.inf],np.nan)
tc_p, np_p, nr_p, nc_p = _tc_cols(pp, f1c, gp_p)

pf_tb = pd.DataFrame({
    "Date": pf_idx,
    "F1_Continuous": pf.round(2),
    "Price_Change": pf.diff().round(4),
    "Mom_Position_(MA35_43_Lag1)": pm.where(~warm_pf).values,
    "Carry_Position_(CarryMom20d_SameDay)": pc.where(~warm_pf).values,
    "Val_Position_(V1_F8_5yr_Lag1)": pv.where(~warm_pf).values,
    "Port_Position_EW_((M+C+V)/3)": pp.round(6).values,
    "Mom_Gross_PnL": gp_m.round(4).values,
    "Carry_Gross_PnL": gp_c.round(4).values,
    "Val_Gross_PnL": gp_v.round(4).values,
    "Port_Gross_PnL": gp_p.round(4).values,
    "Port_Gross_Return_pct": (gr_p*100).round(6).values,
    "Mom_Gross_CumPnL": gp_m.cumsum().round(4).values,
    "Carry_Gross_CumPnL": gp_c.cumsum().round(4).values,
    "Val_Gross_CumPnL": gp_v.cumsum().round(4).values,
    "Port_Gross_CumPnL": gp_p.cumsum().round(4).values,
    "TC_Cost_5bps_USD": tc_p.values,
    "Net_PnL_5bps_USD": np_p.values,
    "Net_Return_5bps_pct": nr_p.values,
    "Net_CumPnL_5bps": nc_p.values,
}, index=pf_idx)

m_port = metrics_both(pp, f1c); m_mom_= metrics_both(pm, f1c)
m_car_ = metrics_both(pc, f1c); m_val_ = metrics_both(pv, f1c)

def write_port_is():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,40); _cw(ws_d,2,90)
    _title(ws_d, 1, "PORTFOLIO IS — Equal-Weight Fixed  (Mom+Carry+Value)  [Full History]")
    desc = [
        ("Tab",     "Tab 10: Portfolio — IS Full History"),
        ("Formula", "Port_Position(t) = (Mom(t) + Carry(t) + Value(t)) / 3"),
        ("Components","Mom: MA(35,43) Lag-1 | Carry: Carry Momentum 20d Same-Day | Value: V1 F8 5yr ±10% Lag-1"),
        ("Range",   "−1 (all short) to +1 (all long). Fractional when signals disagree."),
        ("PnL",     "Port_Gross_PnL(t) = Port_Position(t) × ΔF1_Continuous(t) [USD/MT]"),
        ("Corr",    "The three legs are weakly correlated; pairwise correlations of the position series "
                    "can be recomputed from the Mom/Carry/Val position columns. Low correlation is the "
                    "source of the diversification benefit."),
        ("TC eq.",  "TC on composite (M+C+V)/3 = (1/3)×TC_mom + (1/3)×TC_carry + (1/3)×TC_val. "
                    "Mathematically exact: a flip of one component changes the portfolio by ±2/3, "
                    "TC = |2/3|×(5/10000/2)×F1 = (1/3)×(one-signal flip TC)."),
        ("Headline #s", "Realised Sharpe/return are on the Performance sheet; per-window OOS figures are "
                    "in PORT_OOS_MethodA_EW_WalkForward.xlsx (OOS_Summary). No performance numbers are hardcoded here."),
    ] + TC_DESC_ROWS
    for r, (k, v) in enumerate(desc, start=2): _desc(ws_d, r, k, v)
    ws_t = wb.create_sheet("Tradebook")
    tc_col = list(pf_tb.columns).index("TC_Cost_5bps_USD")+1
    write_tb_sheet(ws_t, pf_tb, tc_start_col=tc_col)
    ws_p = wb.create_sheet("Performance")
    _sw(ws_p); _cw(ws_p,1,38)
    for j, h in enumerate(["Metric","Portfolio EW","Momentum","Carry","Value"], start=1):
        _hdr(ws_p,1,j,h); _cw(ws_p,j,20)
    perf_rows = [
        ("Gross Sharpe",      m_port["gsh"], m_mom_["gsh"], m_car_["gsh"], m_val_["gsh"]),
        ("Net 5bps Sharpe",   m_port["nsh"], m_mom_["nsh"], m_car_["nsh"], m_val_["nsh"]),
        ("Gross Ann Return%", m_port["gar"], m_mom_["gar"], m_car_["gar"], m_val_["gar"]),
        ("Net Ann Return%",   m_port["nar"], m_mom_["nar"], m_car_["nar"], m_val_["nar"]),
        ("Gross Max DD",      m_port["gdd"], m_mom_["gdd"], m_car_["gdd"], m_val_["gdd"]),
        ("Net Max DD",        m_port["ndd"], m_mom_["ndd"], m_car_["ndd"], m_val_["ndd"]),
        ("Gross Total PnL",   m_port["gtp"], m_mom_["gtp"], m_car_["gtp"], m_val_["gtp"]),
        ("N Active Days",     m_port["n_active"], m_mom_["n_active"], m_car_["n_active"], m_val_["n_active"]),
    ]
    for i, rd in enumerate(perf_rows, start=2):
        alt=(i%2==0)
        for j2, v in enumerate(rd, start=1):
            _dat(ws_p, i, j2, round(v,4) if isinstance(v,float) else v, alt=alt)
    wb.save(f"{TBDIR}/portfolio/IS/PORT_IS_EW_Fixed.xlsx")
    print("    Saved PORT_IS_EW_Fixed.xlsx")

write_port_is()

# ── Portfolio OOS (Method A) ──────────────────────────────────────────────────
print("  Portfolio OOS Method A walk-forward…")

def write_port_oos():
    wb = openpyxl.Workbook()
    ws_d = wb.active; ws_d.title = "Description"
    _sw(ws_d); _cw(ws_d,1,40); _cw(ws_d,2,90)
    _title(ws_d, 1, "PORTFOLIO OOS — Method A: EW Fixed Walk-Forward")
    desc = [
        ("Tab",     "Tab 10: Portfolio — OOS Walk-Forward"),
        ("Method",  "A: EW Fixed. All params fixed: MA(35,43), Carry Momentum 20d, V1 F8 5yr ±10%. Weights=1/3 each."),
        ("Results", "Per-window Gross/Net Sharpe are computed in OOS_Summary (authoritative). "
                    "No performance numbers are hardcoded in this description."),
        ("TC",      "Both Gross and Net 5bps Sharpe computed per window. TC on portfolio composite."),
        ("Signals", "Mom: MA(35,43) Lag-1 | Carry: Carry Momentum 20d Same-Day | Val: V1 F8 5yr ±10% Lag-1"),
    ] + TC_DESC_ROWS
    for r, (k, v) in enumerate(desc, start=2): _desc(ws_d, r, k, v)

    ws_s = wb.create_sheet("OOS_Summary")
    _sw(ws_s)
    s_h=["OOS_Year","IS_Period","N_Days","Mom_Sh","Carry_Sh","Val_Sh",
         "Port_Gross_Sh","Port_Net5bps_Sh","Gross_AnnRet%","Net_AnnRet%","Max_DD_Gross","Max_DD_Net"]
    for j, h in enumerate(s_h, start=1): _hdr(ws_s,1,j,h); _cw(ws_s,j,14)

    ws_t = wb.create_sheet("Tradebook_All_Windows")
    _sw(ws_t)
    t_h=["OOS_Window","IS_Period","Date","F1_Continuous","Price_Change",
         "Mom_Pos","Carry_Pos","Val_Pos","Port_Pos_EW",
         "Mom_Gross_PnL","Carry_Gross_PnL","Val_Gross_PnL","Port_Gross_PnL","Port_Gross_Return_pct","Port_Gross_CumPnL",
         "TC_Cost_5bps_USD","Net_PnL_5bps_USD","Net_Return_5bps_pct","Net_CumPnL_5bps"]
    for j, h in enumerate(t_h, start=1): _hdr(ws_t,1,j,h); _cw(ws_t,j,15)
    TC_COL=t_h.index("TC_Cost_5bps_USD")+1

    # Walk-forward runs on the defined (post-warmup) region so windowing is unaffected
    # by the blank IS warmup rows now shown in the Tradebook sheet.
    pidx = pp.dropna().index
    T2=len(pidx); oos_s=IS_W; r_tb=2; sum_rows=[]
    while oos_s<T2:
        oos_e=min(oos_s+OOS_W,T2); is_s=oos_s-IS_W
        oos_yr=str(pidx[oos_s].year)+("*" if (oos_e-oos_s)<OOS_W else "")
        is_lbl=f"{pidx[is_s].year}–{pidx[oos_s-1].year}"
        oos_i=pidx[oos_s:oos_e]
        mo=pm.reindex(oos_i); co=pc.reindex(oos_i); vo=pv.reindex(oos_i)
        po=(mo+co+vo)/3.0; fo=pf.reindex(oos_i)
        gpm_o=mo*fo.diff(); gpc_o=co*fo.diff(); gpv_o=vo*fo.diff(); gpp_o=po*fo.diff()
        with np.errstate(invalid="ignore", divide="ignore"):
            grp=(gpp_o/fo.shift(1)).replace([np.inf,-np.inf],np.nan)
        gcp=gpp_o.cumsum()
        tc_o,npo,nro,nco=_tc_cols(po,f1c,gpp_o)
        def _sh(ret,pos):
            act=ret[pos.abs()>0].dropna()
            return float(act.mean()/act.std(ddof=1)*np.sqrt(252)) if len(act)>5 else np.nan
        def _sh2(gp,pos,fc):
            with np.errstate(invalid="ignore", divide="ignore"):
                r=(gp/fc.shift(1)).replace([np.inf,-np.inf],np.nan)
            return _sh(r,pos)
        sm=_sh2(gpm_o,mo,fo); sc=_sh2(gpc_o,co,fo); sv=_sh2(gpv_o,vo,fo)
        sg=_sh(grp,po); sn=_sh(nro,po)
        gar=float(grp.dropna().mean()*252*100); nar=float(nro.dropna().mean()*252*100)
        gdd=float((gcp-gcp.cummax()).min()); ndd=float((nco-nco.cummax()).min())
        def _r(x): return round(x,4) if not np.isnan(x) else "N/A"
        sum_rows.append([oos_yr,is_lbl,len(oos_i),_r(sm),_r(sc),_r(sv),_r(sg),_r(sn),
                          round(gar,2),round(nar,2),round(gdd,2),round(ndd,2)])
        for ki,date in enumerate(oos_i):
            alt=(r_tb%2==0); pov=po.iloc[ki]
            vals=[oos_yr,is_lbl,str(date.date()),round(fo.iloc[ki],2),
                  round(fo.diff().iloc[ki],4) if ki>0 else None,
                  round(mo.iloc[ki],4),round(co.iloc[ki],4),round(vo.iloc[ki],4),round(pov,6),
                  round(gpm_o.iloc[ki],4),round(gpc_o.iloc[ki],4),round(gpv_o.iloc[ki],4),
                  round(gpp_o.iloc[ki],4),
                  round(grp.iloc[ki]*100,6) if not np.isnan(grp.iloc[ki]) else None,
                  round(gcp.iloc[ki],4)]
            tcv=[round(tc_o.iloc[ki],4),round(npo.iloc[ki],4),
                 round(nro.iloc[ki]*100,6) if not np.isnan(nro.iloc[ki]) else None,
                 round(nco.iloc[ki],4)]
            for j2,v in enumerate(vals,start=1):
                _dat(ws_t,r_tb,j2,v,alt=alt,hl=("pos" if j2==9 and pov>0 else ("neg" if j2==9 and pov<0 else None)))
            for j2,v in enumerate(tcv,start=TC_COL):
                _dat_tc(ws_t,r_tb,j2,v,alt=alt)
            r_tb+=1
        oos_s+=OOS_W
    for r2,rd in enumerate(sum_rows,start=2):
        for j2,v in enumerate(rd,start=1): _dat(ws_s,r2,j2,v,alt=(r2%2==0))
    gsh_a=[r[6] for r in sum_rows if r[6]!="N/A"]
    nsh_a=[r[7] for r in sum_rows if r[7]!="N/A"]
    r_avg=len(sum_rows)+2
    _dat(ws_s,r_avg,1,"AVERAGE")
    for j2,col_vals,label in [(7,gsh_a,"avg gross"),(8,nsh_a,"avg net")]:
        _dat(ws_s,r_avg,j2,round(np.nanmean(col_vals),4) if col_vals else "N/A")
    _dat(ws_s,r_avg+1,1,"% Positive (Gross)")
    _dat(ws_s,r_avg+1,7,f"{100*sum(1 for s in gsh_a if s>0)/len(gsh_a):.1f}%" if gsh_a else "N/A")
    _dat(ws_s,r_avg+1,8,f"{100*sum(1 for s in nsh_a if s>0)/len(nsh_a):.1f}%" if nsh_a else "N/A")
    wb.save(f"{TBDIR}/portfolio/OOS/PORT_OOS_MethodA_EW_WalkForward.xlsx")
    print("    Saved PORT_OOS_MethodA_EW_WalkForward.xlsx")

write_port_oos()

# ─────────────────────────────────────────────────────────────────────────────
print("\nAll tradebooks generated.")
for root, dirs, files in os.walk(TBDIR):
    lvl=root.replace(TBDIR,"").count(os.sep)
    print("  "*lvl + os.path.basename(root) + "/")
    for f in sorted(files):
        kb=os.path.getsize(os.path.join(root,f))//1024
        print("  "*(lvl+1) + f + f"  ({kb} KB)")
