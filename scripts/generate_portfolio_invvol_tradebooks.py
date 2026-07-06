"""
generate_portfolio_invvol_tradebooks.py
Adds the INVERSE-VOL portfolio tradebooks (IS + OOS walk-forward) alongside the
existing equal-weight ones. Inverse-vol = each sleeve weighted by 1 / its trailing
63d return-vol (renormalised daily, lagged 1 day; warmup weights = 1/3) — exactly
the dashboard Tab-10 method.

Writes into BOTH tradebook sets, each with its folder's TC convention:
  tradebooks/portfolio/        -> TC on F1_continuous
  Tradebooks 6-28/portfolio/   -> TC on F1_raw
Only portfolio/ files are written (other files / locked carry files untouched).
"""
import os, sys, io, warnings
warnings.filterwarnings("ignore")
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import numpy as np, pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
DATA_DIR = os.path.join(_REPO_ROOT, "data")
OUTPUTS_DIR = os.path.join(_REPO_ROOT, "outputs")
BASE = DATA_DIR
IS_W, OOS_W, VOLW = 1260, 252, 63

# ── Data ───────────────────────────────────────────────────────────────────────
f1_df = pd.read_csv(f"{BASE}/LME_Copper_Rolling_F1_v2.csv", parse_dates=["Date"]).set_index("Date")
f1_df.index = f1_df.index.normalize()
f1r = f1_df["F1_raw"].sort_index(); f1c = f1_df["F1_continuous"].sort_index()
xls = pd.ExcelFile(f"{BASE}/Metals Futures Curve.csv")
cu = next(s for s in xls.sheet_names if "copper" in s.lower() and "lme" in s.lower())
draw = pd.read_excel(xls, sheet_name=cu, header=None, nrows=5)
hr = [i for i in range(min(4, len(draw))) if any(k in " ".join(str(v).lower() for v in draw.iloc[i].values if pd.notna(v)) for k in ["date","f1","f2","price"])]
dfc = pd.read_excel(xls, sheet_name=cu, header=hr if hr else [0,1])
if isinstance(dfc.columns, pd.MultiIndex):
    dfc.columns = ["_".join(str(p) for p in t if pd.notna(p) and "Unnamed" not in str(p) and str(p).strip()) or str(t) for t in dfc.columns]
dfc.columns = [str(c).strip() for c in dfc.columns]
dcc = [c for c in dfc.columns if "date" in c.lower()]
dfc = dfc.rename(columns={dcc[0]: "Date"}); dfc["Date"] = pd.to_datetime(dfc["Date"], errors="coerce")
dfc = dfc.dropna(subset=["Date"]).set_index("Date").sort_index(); dfc.index = dfc.index.normalize()
crv = {}
for col in dfc.columns:
    cl = col.lower().replace(" ", "_")
    for i in range(1, 28):
        if f"f{i}_" in cl and "price" in cl: crv[f"F{i}"] = pd.to_numeric(dfc[col], errors="coerce"); break
crv = pd.DataFrame(crv, index=dfc.index)

# ── Style (matches generate_tradebooks.py) ─────────────────────────────────────
HDR_FILL=PatternFill("solid",fgColor="1A1A2E"); D1=PatternFill("solid",fgColor="0F0F1A")
D2=PatternFill("solid",fgColor="141428"); POS=PatternFill("solid",fgColor="003300")
NEG=PatternFill("solid",fgColor="330000"); DSC=PatternFill("solid",fgColor="0D0D1F")
HDR_FONT=Font(name="Calibri",bold=True,color="B87333",size=10); DAT=Font(name="Calibri",color="D4D4D4",size=9)
DSCF=Font(name="Calibri",color="C8B89A",size=10); TTL=Font(name="Calibri",bold=True,color="E8D5B0",size=13)
KEY=Font(name="Calibri",bold=True,color="B87333",size=10); TCF=Font(name="Calibri",bold=True,color="64B5F6",size=9)
def _sw(ws): ws.sheet_view.showGridLines=False
def _cw(ws,j,w): ws.column_dimensions[get_column_letter(j)].width=w
def _hdr(ws,r,c,v):
    x=ws.cell(row=r,column=c,value=v); x.font=HDR_FONT; x.fill=HDR_FILL; x.alignment=Alignment(horizontal="center",wrap_text=True)
def _dat(ws,r,c,v,alt=False,hl=None,tc=False):
    x=ws.cell(row=r,column=c,value=v); fill=D2 if alt else D1
    if hl=="pos": fill=POS
    elif hl=="neg": fill=NEG
    x.font=TCF if tc else DAT; x.fill=fill; x.alignment=Alignment(horizontal="center")
def _desc(ws,r,k,v=""):
    a=ws.cell(row=r,column=1,value=k); a.font=KEY; a.fill=DSC
    b=ws.cell(row=r,column=2,value=v); b.font=DSCF; b.fill=DSC
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=9)
def _title(ws,r,t):
    c=ws.cell(row=r,column=1,value=t); c.font=TTL; c.fill=DSC; ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=9)
def write_tb(ws, data, tc_start):
    _sw(ws); ws.freeze_panes=ws.cell(row=2,column=1)
    cols=list(data.columns)
    for j,col in enumerate(cols,1): _hdr(ws,1,j,col); _cw(ws,j,max(len(col)+2,12))
    _cw(ws,1,13)
    for i,(date,row) in enumerate(data.iterrows(),2):
        alt=(i%2==0)
        ws.cell(row=i,column=1,value=date.date() if hasattr(date,'date') else date).font=DAT
        ws.cell(row=i,column=1).fill=D2 if alt else D1
        for j,col in enumerate(cols[1:],2):
            v=row[col]
            if isinstance(v,float) and not np.isnan(v): v=round(v,6)
            hl=None
            if "Position" in col and isinstance(v,(int,float)) and not (isinstance(v,float) and np.isnan(v)):
                hl="pos" if v>0 else ("neg" if v<0 else None)
            blank = isinstance(v,float) and np.isnan(v)
            _dat(ws,i,j, None if blank else v, alt=alt, hl=hl, tc=(j>=tc_start))

# ── Legs (identical to dashboard Tab-10 / generate_tradebooks.py) ──────────────
pos_mom = np.sign(f1r.rolling(35).mean()-f1r.rolling(43).mean()).shift(1)            # shift-1
_crb = ((crv["F1"]-crv["F2"])/crv["F1"]).replace([np.inf,-np.inf],np.nan)
pos_carry = np.sign(_crb-_crb.shift(20)).shift(1).reindex(f1c.index)                  # carry-mom 20d
_f8=crv["F8"].dropna(); _ma8=_f8.rolling(1260,min_periods=630).mean()
_dev=((_f8-_ma8)/_ma8).replace([np.inf,-np.inf],np.nan)
_sig=pd.Series(np.where(_dev<-0.10,1.0,np.where(_dev>0.10,-1.0,0.0)),index=_dev.index).where(_dev.notna())
pos_val = _sig.shift(1).reindex(f1c.index)                                            # V1 F8 5yr

idx = f1c.index
pm = pos_mom.reindex(idx); pc = pos_carry.reindex(idx); pv = pos_val.reindex(idx)
fcc = f1c.reindex(idx); frr = f1r.reindex(idx)

# inverse-vol weights (63d trailing return-vol, lagged 1d, warmup 1/3)
def _sret(p): return (p.fillna(0)*fcc.diff()/fcc.shift(1)).replace([np.inf,-np.inf],np.nan).fillna(0)
_volw = pd.DataFrame({"m":_sret(pm),"c":_sret(pc),"v":_sret(pv)}).rolling(VOLW).std().shift(1)
_iw = (1.0/_volw).replace([np.inf,-np.inf],np.nan); _iw = _iw.div(_iw.sum(axis=1),axis=0)
wm=_iw["m"]; wc=_iw["c"]; wv=_iw["v"]
# composite defined only where all three legs are defined; weights shown where defined
defined = pm.notna() & pc.notna() & pv.notna()
wm_f=wm.fillna(1/3); wc_f=wc.fillna(1/3); wv_f=wv.fillna(1/3)
port_iv = (wm_f*pm.fillna(0) + wc_f*pc.fillna(0) + wv_f*pv.fillna(0)).where(defined)
warm = port_iv.isna()

def _metrics(pos, tc_price, bps=5):
    posf=pos.fillna(0.0)
    gp=(posf*fcc.diff()).where(~warm)
    chg=posf.diff().abs(); chg.iloc[0]=abs(posf.iloc[0])
    tc=(chg*(bps/10000.0/2.0)*tc_price).where(~warm)
    net=gp-tc
    def sh(pnl):
        r=(pnl/fcc.shift(1)).replace([np.inf,-np.inf],np.nan)
        a=r[(posf!=0)&(~warm)].dropna()
        return float(a.mean()/a.std(ddof=1)*np.sqrt(252)) if len(a)>20 and a.std(ddof=1)>0 else np.nan
    def ann(pnl):
        r=(pnl/fcc.shift(1)).replace([np.inf,-np.inf],np.nan); return float(r.dropna().mean()*252*100)
    def dd(pnl):
        c=pnl.cumsum(); return float((c-c.cummax()).min())
    return dict(gsh=sh(gp),nsh=sh(net),gar=ann(gp),nar=ann(net),gdd=dd(gp),ndd=dd(net),
                gtp=float(gp.sum()),ntp=float(net.sum()),
                nact=int(((posf!=0)&(~warm)).sum()),gp=gp,net=net,tc=tc)

# ── Writers ────────────────────────────────────────────────────────────────────
def write_set(folder, tc_price, tc_basis):
    pdir=os.path.join(OUTPUTS_DIR, folder, "portfolio")
    os.makedirs(os.path.join(pdir,"IS"),exist_ok=True); os.makedirs(os.path.join(pdir,"OOS"),exist_ok=True)
    m_iv=_metrics(port_iv, tc_price); m_ew_pos=(pm.fillna(0)+pc.fillna(0)+pv.fillna(0))/3.0
    m_ew_pos=m_ew_pos.where(defined); m_ew=_metrics(m_ew_pos, tc_price)
    gp=m_iv["gp"]; net=m_iv["net"]; tc=m_iv["tc"]
    with np.errstate(invalid="ignore",divide="ignore"):
        gr=(gp/fcc.shift(1)).replace([np.inf,-np.inf],np.nan); nr=(net/fcc.shift(1)).replace([np.inf,-np.inf],np.nan)
    DESC=[("Tab","Tab 10: Portfolio — Inverse-Vol weighting"),
        ("Method","Inverse-volatility: each sleeve weighted by 1 / its trailing 63d return-vol, "
                  "renormalised daily across the 3 sleeves, lagged 1 day (no look-ahead). Warmup weights = 1/3."),
        ("Legs","Mom: MA(35,43) shift-1 | Carry: CarryMom 20d shift-1 | Value: V1 F8 5yr ±10% shift-1"),
        ("Weights","w_i(t) = (1/vol_i(t)) / Σ_j (1/vol_j(t)), vol from 63d rolling std of sleeve return"),
        ("Composite","Port_Position_IV(t) = w_mom·Mom + w_carry·Carry + w_val·Value"),
        ("PnL","Gross_PnL(t) = Port_Position_IV(t) × ΔF1_Continuous(t) [USD/MT]"),
        ("TC basis",f"TC_Cost(t) = |ΔPort_Position| × (5/10000/2) × {tc_basis}(t). "
                    "PnL/returns on F1_Continuous; spread paid on the actual traded price."),
        ("vs EW","Equal-weight (1/3 each) is the dashboard default; inverse-vol is the alternative shown here. "
                 "For these similar-vol sleeves EW tends to edge it on net Sharpe with a smaller drawdown."),
        ("Warmup","Blank until all three legs are initialised (~630d, Value-driven). Date/price from day 1.")]
    # IS tradebook
    tb=pd.DataFrame({
        "Date":idx,"F1_Continuous":fcc.round(2),"Price_Change":fcc.diff().round(4),
        "Mom_Position":pm.where(~warm).values,"Carry_Position":pc.where(~warm).values,"Val_Position":pv.where(~warm).values,
        "w_mom":wm_f.where(~warm).round(4).values,"w_carry":wc_f.where(~warm).round(4).values,"w_val":wv_f.where(~warm).round(4).values,
        "Port_Position_IV":port_iv.round(6).values,
        "Port_Gross_PnL":gp.round(4).values,"Port_Gross_Return_pct":(gr*100).round(6).values,"Port_Gross_CumPnL":gp.cumsum().round(4).values,
        f"TC_Cost_5bps_on_{tc_basis}":tc.round(4).values,"Net_PnL_5bps":net.round(4).values,
        "Net_Return_5bps_pct":(nr*100).round(6).values,"Net_CumPnL_5bps":net.cumsum().round(4).values,
    },index=idx)
    wb=openpyxl.Workbook(); wd=wb.active; wd.title="Description"; _sw(wd); _cw(wd,1,22); _cw(wd,2,95)
    _title(wd,1,f"PORTFOLIO IS — Inverse-Vol (63d)  [TC on {tc_basis}]")
    for r,(k,v) in enumerate(DESC,2): _desc(wd,r,k,v)
    wt=wb.create_sheet("Tradebook"); tcc=list(tb.columns).index(f"TC_Cost_5bps_on_{tc_basis}")+1; write_tb(wt,tb,tcc)
    wp=wb.create_sheet("Performance"); _sw(wp); _cw(wp,1,30)
    for j,h in enumerate(["Metric","Inverse-Vol","Equal-Weight"],1): _hdr(wp,1,j,h); _cw(wp,j,18)
    rows=[("Gross Sharpe",m_iv["gsh"],m_ew["gsh"]),("Net 5bps Sharpe",m_iv["nsh"],m_ew["nsh"]),
          ("Gross Ann Ret %",m_iv["gar"],m_ew["gar"]),("Net Ann Ret %",m_iv["nar"],m_ew["nar"]),
          ("Gross Max DD",m_iv["gdd"],m_ew["gdd"]),("Net Max DD",m_iv["ndd"],m_ew["ndd"]),
          ("Gross Total PnL",m_iv["gtp"],m_ew["gtp"]),("Active Days",m_iv["nact"],m_ew["nact"])]
    for i,(lbl,a,b) in enumerate(rows,2):
        for j,val in enumerate([lbl,round(a,4) if isinstance(a,float) else a,round(b,4) if isinstance(b,float) else b],1):
            _dat(wp,i,j,val,alt=(i%2==0))
    p1=os.path.join(pdir,"IS","PORT_IS_InverseVol_63d.xlsx"); wb.save(p1); print("  saved",p1.replace(OUTPUTS_DIR+os.sep,""))
    # OOS walk-forward (inverse-vol method fixed; sliced into non-overlapping windows)
    wb2=openpyxl.Workbook(); wd2=wb2.active; wd2.title="Description"; _sw(wd2); _cw(wd2,1,22); _cw(wd2,2,95)
    _title(wd2,1,f"PORTFOLIO OOS — Inverse-Vol (63d) Walk-Forward  [TC on {tc_basis}]")
    for r,(k,v) in enumerate(DESC+[("Walk-forward","IS=1260d, OOS=252d, non-overlapping. Inverse-vol is a deterministic "
        "rolling rule (no per-window fit); windows just partition the fixed series.")],2): _desc(wd2,r,k,v)
    ws=wb2.create_sheet("OOS_Summary"); _sw(ws)
    for j,h in enumerate(["OOS_Year","IS_Period","N_Days","Gross_Sharpe","Net_5bps_Sharpe","Gross_AnnRet%","Net_AnnRet%","Max_DD_Gross"],1):
        _hdr(ws,1,j,h); _cw(ws,j,15)
    didx=port_iv.dropna().index; T=len(didx); oos_s=IS_W; r=2; gl=[]; nl=[]
    while oos_s<T:
        oos_e=min(oos_s+OOS_W,T); part=(oos_e-oos_s)<OOS_W
        oi=didx[oos_s:oos_e]; yr=str(oi[0].year)+("*" if part else "")
        lbl=f"{didx[oos_s-IS_W].year}–{didx[oos_s-1].year}"
        po=port_iv.reindex(oi); fo=fcc.reindex(oi); pr=tc_price.reindex(oi)
        gpo=po*fo.diff(); chg=po.diff().abs(); chg.iloc[0]=abs(po.iloc[0])
        tco=chg*(5/10000.0/2.0)*pr; no=gpo-tco
        def _s(pnl):
            rr=(pnl/fo.shift(1)).replace([np.inf,-np.inf],np.nan); a=rr[po.abs()>0].dropna()
            return float(a.mean()/a.std(ddof=1)*np.sqrt(252)) if len(a)>5 and a.std(ddof=1)>0 else np.nan
        gs=_s(gpo); ns=_s(no)
        gar=float((gpo/fo.shift(1)).replace([np.inf,-np.inf],np.nan).dropna().mean()*252*100)
        nar=float((no/fo.shift(1)).replace([np.inf,-np.inf],np.nan).dropna().mean()*252*100)
        gddw=float((gpo.cumsum()-gpo.cumsum().cummax()).min())
        def R(x): return round(x,4) if not np.isnan(x) else "N/A"
        for j,val in enumerate([yr,lbl,len(oi),R(gs),R(ns),round(gar,2),round(nar,2),round(gddw,2)],1):
            _dat(ws,r,j,val,alt=(r%2==0))
        if not np.isnan(gs): gl.append(gs)
        if not np.isnan(ns): nl.append(ns)
        r+=1; oos_s+=OOS_W
    _dat(ws,r+1,1,"AVERAGE"); _dat(ws,r+1,4,round(np.nanmean(gl),4) if gl else "N/A"); _dat(ws,r+1,5,round(np.nanmean(nl),4) if nl else "N/A")
    p2=os.path.join(pdir,"OOS","PORT_OOS_InverseVol_63d_WalkForward.xlsx"); wb2.save(p2); print("  saved",p2.replace(OUTPUTS_DIR+os.sep,""))
    return m_iv,m_ew

print("tradebooks/  (TC on F1_continuous):")
write_set("tradebooks", fcc, "F1_Continuous")
print("Tradebooks 6-28/  (TC on F1_raw):")
write_set("Tradebooks 6-28", frr, "F1_raw")
print("Done.")
