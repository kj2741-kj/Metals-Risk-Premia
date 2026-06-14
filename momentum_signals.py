"""
momentum_signals.py
===================
CTA momentum (Baz-Granger Eqs 29-33) and MA-crossover momentum signals.
Optimization + full tradebook generation for LME Copper.

Signal / PnL separation (no leakage):
  - Signal  : computed at close of day t using F1_RAW only (no future data)
  - Position (Lag-1)    : Signal[t-1]  -> trade entered at close t-1, live from t
  - Position (Same-Day) : Signal[t]    -> trade entered at close t
  - PnL     : Position[t] x delta_F1_continuous[t]  (roll cost captured in F1_cont)

Return / risk metrics:
  - All expressed as % of notional (daily return = PnL / F1_cont_prev * 100)
  - Ann Return (%), Std Dev (%), Max DD (%), Calmar, Sortino use % returns
  - Sharpe ratio is unitless (same formula, same value either way)
  - Total PnL and Avg Win/Loss kept in USD/MT for interpretability

Outputs (momentum_output_v3/):
  CTA_Paper_Tradebook.xlsx          2 sheets: Lag-1 + Same-Day
  CTA_Top5_Tradebook_{S}_{L}.xlsx   2 sheets: Lag-1 + Same-Day
  MA_Top5_Tradebook_{m}_{n}.xlsx    2 sheets: Lag-1 + Same-Day
  MA_Crossover_Optimization.csv     full lag-1 scan (7875 pairs)
  CTA_Optimization.csv              full lag-1 scan (7875 pairs)
"""

from __future__ import annotations

import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from rolling_continuous import get_metal_rolling_f1

OUT_DIR      = Path("momentum_output_v3")
OUT_DIR.mkdir(exist_ok=True)

MAX_LOOKBACK = 126
CTA_SHORT    = (8, 16, 32)
CTA_LONG     = (24, 48, 96)
CTA_PW       = 63
CTA_SW       = 252


# ══════════════════════════════════════════════════════════════════════════════
# EWMA helper  (Baz-Granger convention: lambda = (n-1)/n  ->  com = n-1)
# ══════════════════════════════════════════════════════════════════════════════

def _ewma(s: pd.Series, n: int) -> pd.Series:
    return s.ewm(com=n - 1, adjust=False).mean()


# ══════════════════════════════════════════════════════════════════════════════
# PERFORMANCE METRICS
# ══════════════════════════════════════════════════════════════════════════════

def _consecutive(arr: np.ndarray, val: int) -> int:
    """Longest consecutive run of `val` in a sign array."""
    best = cur = 0
    for x in arr:
        if x == val:
            cur += 1
            best = max(best, cur)
        else:
            cur = 0
    return best


def compute_performance(daily_pnl: pd.Series,
                        position:  pd.Series,
                        f1_cont:   pd.Series,
                        same_day:  bool = False) -> dict:
    """
    Full performance summary.

    Return / risk metrics expressed as % of notional:
      daily_ret[t] = position[t] * delta_F1_cont[t] / F1_cont[t-1]

    Only active days (position != 0) enter mean/std/win-loss stats.
    Max Drawdown uses the FULL equity curve (including zero-return days)
    so that the drawdown from peak during warmup or flat periods is captured.
    """
    f1_prev    = f1_cont.shift(1)
    daily_ret  = daily_pnl / f1_prev          # daily % return (fraction)

    active_ret = daily_ret[position != 0].dropna()
    active_pnl = daily_pnl[position != 0].dropna()
    n          = len(active_ret)

    # ── Annualised return & vol (% terms) ────────────────────────────────────
    ann_ret_pct   = float(active_ret.mean()  * 252 * 100) if n > 1 else np.nan
    ann_std_pct   = float(active_ret.std()   * np.sqrt(252) * 100) if n > 1 else np.nan
    sharpe        = ann_ret_pct / ann_std_pct if (ann_std_pct and ann_std_pct > 0) else np.nan

    # ── Sortino (% terms, downside only) ────────────────────────────────────
    down_ret      = active_ret[active_ret < 0]
    sortino_denom = float(down_ret.std() * np.sqrt(252) * 100) if len(down_ret) > 1 else np.nan
    sortino       = ann_ret_pct / sortino_denom if (sortino_denom and sortino_denom > 0) else np.nan

    # ── Max Drawdown (% terms, full equity curve) ───────────────────────────
    full_ret      = daily_ret.fillna(0)
    cum_ret_pct   = full_ret.cumsum() * 100
    running_max   = cum_ret_pct.cummax()
    max_dd_pct    = float((cum_ret_pct - running_max).min())
    calmar        = ann_ret_pct / abs(max_dd_pct) if max_dd_pct != 0 else np.nan

    # ── Win/loss in USD/MT ───────────────────────────────────────────────────
    wins          = active_pnl[active_pnl > 0]
    losses        = active_pnl[active_pnl < 0]
    total_pnl     = float(active_pnl.sum())
    avg_win       = float(wins.mean())          if len(wins)   > 0 else np.nan
    avg_loss      = float(losses.mean())        if len(losses) > 0 else np.nan
    pf_num        = float(wins.sum())
    pf_den        = float(abs(losses.sum()))
    profit_factor = pf_num / pf_den             if pf_den > 0  else np.nan
    hit_rate      = float((active_ret > 0).mean()) if n > 0    else np.nan

    sign_arr      = np.where(active_pnl > 0, 1, -1)
    max_con_w     = _consecutive(sign_arr,  1)
    max_con_l     = _consecutive(sign_arr, -1)

    pos_note = ("Position[t] = Signal[t]  (same-day entry, no lag)"
                if same_day else
                "Position[t] = Signal[t-1]  (1-day lag, no leakage)")

    return {
        "Entry Convention"        : "Same-Day" if same_day else "Lag-1 (Next-Day)",
        "Start Date"              : str(daily_pnl.index[0].date()),
        "End Date"                : str(daily_pnl.index[-1].date()),
        "Total Calendar Days"     : len(daily_pnl),
        "Active Trading Days"     : n,
        "Warmup Days"             : len(daily_pnl) - n,
        "Total PnL (USD/MT)"      : round(total_pnl,    2),
        "Annualized Return (%)"   : round(ann_ret_pct,  4),
        "Annualized Std Dev (%)"  : round(ann_std_pct,  4),
        "Sharpe Ratio"            : round(sharpe,        4),
        "Sortino Ratio"           : round(sortino,       4),
        "Max Drawdown (%)"        : round(max_dd_pct,   4),
        "Calmar Ratio"            : round(calmar,        4),
        "Hit Rate"                : f"{hit_rate*100:.2f}%",
        "Avg Win (USD/MT)"        : round(avg_win,       2),
        "Avg Loss (USD/MT)"       : round(avg_loss,      2),
        "Profit Factor"           : round(profit_factor, 4),
        "Max Consecutive Wins"    : max_con_w,
        "Max Consecutive Losses"  : max_con_l,
        "POSITION NOTE"           : pos_note,
        "PnL NOTE"                : "Daily_PnL = Position x delta_F1_continuous  (roll cost in F1_cont)",
        "Cum_PnL NOTE"            : "Cum_PnL = cumsum(Daily_PnL)  [signed position, not abs]",
        "Return NOTE"             : "Daily_Return(%) = Daily_PnL / F1_cont[t-1] * 100",
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL WRITER — 2 sheets per file
# ══════════════════════════════════════════════════════════════════════════════

_HEADER_FILL   = PatternFill("solid", fgColor="2B3A47")   # dark slate
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
_SECTION_FILL  = PatternFill("solid", fgColor="4A4A4A")   # charcoal
_SECTION_FONT  = Font(bold=True, color="F5C842", size=10) # copper/gold
_METRIC_FILL   = PatternFill("solid", fgColor="F2F2F2")
_METRIC_FONT   = Font(size=9)
_TB_HEADER_FILL = PatternFill("solid", fgColor="B87333")  # copper
_TB_HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
_EVEN_FILL     = PatternFill("solid", fgColor="FAFAFA")
_ODD_FILL      = PatternFill("solid", fgColor="FFFFFF")
_THIN_BORDER   = Border(
    bottom=Side(style="thin", color="D0D0D0")
)


def _style_cell(cell, fill=None, font=None, align=None):
    if fill:  cell.fill      = fill
    if font:  cell.font      = font
    if align: cell.alignment = align


def _write_xl_sheet(wb: Workbook,
                    tb: pd.DataFrame,
                    metrics: dict,
                    sheet_name: str) -> None:
    """Write one Excel sheet: performance summary block then tradebook rows."""
    ws = wb.create_sheet(sheet_name)

    # ── Section 1: Performance Summary ──────────────────────────────────────
    ws.append(["PERFORMANCE SUMMARY", ""])
    _style_cell(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT,
                align=Alignment(horizontal="left"))
    _style_cell(ws.cell(ws.max_row, 2), fill=_SECTION_FILL)

    ws.append(["Metric", "Value"])
    for col in (1, 2):
        _style_cell(ws.cell(ws.max_row, col), fill=_HEADER_FILL, font=_HEADER_FONT)

    for i, (k, v) in enumerate(metrics.items()):
        ws.append([k, v])
        fill = _METRIC_FILL if i % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        _style_cell(ws.cell(ws.max_row, 1), fill=fill, font=_METRIC_FONT)
        _style_cell(ws.cell(ws.max_row, 2), fill=fill, font=_METRIC_FONT)

    ws.append(["", ""])   # spacer

    # ── Section 2: Tradebook ─────────────────────────────────────────────────
    ws.append(["TRADEBOOK"] + [""] * (len(tb.columns) - 1))
    _style_cell(ws.cell(ws.max_row, 1), fill=_SECTION_FILL, font=_SECTION_FONT)

    # Column headers
    ws.append(list(tb.columns))
    hdr_row = ws.max_row
    for col_idx in range(1, len(tb.columns) + 1):
        _style_cell(ws.cell(hdr_row, col_idx), fill=_TB_HEADER_FILL,
                    font=_TB_HEADER_FONT, align=Alignment(horizontal="center"))

    # Data rows
    for row_idx, (_, row) in enumerate(tb.iterrows()):
        vals = []
        for v in row:
            if isinstance(v, float) and np.isnan(v):
                vals.append(None)
            elif hasattr(v, 'item'):
                vals.append(v.item())
            else:
                vals.append(v)
        ws.append(vals)
        fill = _EVEN_FILL if row_idx % 2 == 0 else _ODD_FILL
        for col_idx in range(1, len(tb.columns) + 1):
            _style_cell(ws.cell(ws.max_row, col_idx), fill=fill,
                        font=Font(size=8))

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20   # Date
    ws.column_dimensions["B"].width = 30   # Strategy / Value
    for i in range(3, len(tb.columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    # Wider metric column
    ws.column_dimensions["A"].width = max(
        20, max((len(str(k)) for k in metrics.keys()), default=20) + 2
    )


def save_tradebook_excel(tb_lag:  pd.DataFrame, met_lag:  dict,
                          tb_same: pd.DataFrame, met_same: dict,
                          filepath: Path) -> None:
    """Save one .xlsx with two sheets: Lag-1 and Same-Day."""
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet
    _write_xl_sheet(wb, tb_lag,  met_lag,  "Lag-1 (Next-Day Entry)")
    _write_xl_sheet(wb, tb_same, met_same, "Same-Day Entry")
    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
# POSITION BUILDER (shared)
# ══════════════════════════════════════════════════════════════════════════════

def _build_position(signal: pd.Series, same_day: bool) -> pd.Series:
    sig_np = signal.values.astype(float)
    T      = len(sig_np)
    pos_np = np.empty(T)
    if same_day:
        pos_np[:] = np.where(np.isfinite(sig_np), sig_np, 0.0)
    else:
        pos_np[0] = 0.0
        pos_np[1:] = np.where(np.isfinite(sig_np[:-1]), sig_np[:-1], 0.0)
    return pd.Series(pos_np, index=signal.index)


# ══════════════════════════════════════════════════════════════════════════════
# MA CROSSOVER TRADEBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_ma_tradebook(f1_raw: pd.Series,
                       f1_cont: pd.Series,
                       m: int, n: int,
                       same_day: bool = False) -> pd.DataFrame:
    """
    Full tradebook for a single MA(m,n) crossover strategy.

    Signal = sign(SMA(m) - SMA(n)) computed from F1_raw only.
    Position enters 1 day after signal (same_day=False) or same day (same_day=True).
    PnL = Position[t] x delta_F1_continuous[t]  (F1_cont never used for signal).
    """
    ma_m      = f1_raw.rolling(m).mean()
    ma_n      = f1_raw.rolling(n).mean()
    crossover = ma_m - ma_n
    signal    = np.sign(crossover)
    position  = _build_position(signal, same_day)

    delta     = f1_cont.diff()
    daily_pnl = position * delta
    cum_pnl   = daily_pnl.cumsum()
    mtm       = position * f1_cont
    label     = f"MA_Crossover({m},{n})" + (" [Same-Day]" if same_day else " [Lag-1]")

    return pd.DataFrame({
        "Date"                  : f1_raw.index,
        "Strategy"              : label,
        "F1_raw"                : f1_raw.round(4).values,
        "F1_continuous"         : f1_cont.round(4).values,
        f"MA_{m}"               : ma_m.round(4).values,
        f"MA_{n}"               : ma_n.round(4).values,
        "Crossover"             : crossover.round(4).values,
        "Signal"                : signal.values,
        "Position"              : position.values,
        "F1_cont_daily_change"  : delta.round(4).values,
        "Daily_PnL"             : daily_pnl.round(4).values,
        "MTM"                   : mtm.round(4).values,
        "Cum_PnL"               : cum_pnl.round(4).values,
    })


# ══════════════════════════════════════════════════════════════════════════════
# CTA SINGLE-PAIR TRADEBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_cta_single_tradebook(f1_raw: pd.Series,
                                f1_cont: pd.Series,
                                s: int, l: int,
                                pw: int = CTA_PW,
                                sw: int = CTA_SW,
                                same_day: bool = False) -> pd.DataFrame:
    """
    Tradebook for one EWMA crossover pair with full CTA normalisation pipeline.

    Pipeline (Baz-Granger Eqs 29-33):
      x = EWMA(S) - EWMA(L)                       [Eq 29]
      y = x / RollingStd(F1_raw, pw=63)           [Eq 30]
      z = y / RollingStd(y, sw=252)               [Eq 31]
      u = z * exp(-z^2/4) / 0.89  [response fn]  [Eq 32]
      Signal = sign(u)

    EWMA convention: com = n-1  (lambda = (n-1)/n, alpha = 1/n).
    All signal computation uses F1_raw only; PnL uses F1_continuous only.
    """
    ewma_s    = _ewma(f1_raw, s)
    ewma_l    = _ewma(f1_raw, l)
    x         = ewma_s - ewma_l
    y         = x / f1_raw.rolling(pw).std()
    with np.errstate(invalid="ignore"):
        z     = y / y.rolling(sw).std()
        u     = z * np.exp(-z.values**2 / 4) / 0.89
    signal    = np.sign(u)
    position  = _build_position(signal, same_day)

    delta     = f1_cont.diff()
    daily_pnl = position * delta
    cum_pnl   = daily_pnl.cumsum()
    mtm       = position * f1_cont
    label     = f"CTA({s},{l})" + (" [Same-Day]" if same_day else " [Lag-1]")

    return pd.DataFrame({
        "Date"                  : f1_raw.index,
        "Strategy"              : label,
        "F1_raw"                : f1_raw.round(4).values,
        "F1_continuous"         : f1_cont.round(4).values,
        f"EWMA_{s}"             : ewma_s.round(4).values,
        f"EWMA_{l}"             : ewma_l.round(4).values,
        "x_EWMA_diff"           : x.round(6).values,
        "y_vol_norm"            : y.round(6).values,
        "z_sig_norm"            : z.round(6).values,
        "u_response"            : u.round(6).values,
        "Signal"                : signal.values,
        "Position"              : position.values,
        "F1_cont_daily_change"  : delta.round(4).values,
        "Daily_PnL"             : daily_pnl.round(4).values,
        "MTM"                   : mtm.round(4).values,
        "Cum_PnL"               : cum_pnl.round(4).values,
    })


# ══════════════════════════════════════════════════════════════════════════════
# CTA PAPER (3-PAIR) TRADEBOOK
# ══════════════════════════════════════════════════════════════════════════════

def build_cta_paper_tradebook(f1_raw: pd.Series,
                               f1_cont: pd.Series,
                               short_params: tuple = CTA_SHORT,
                               long_params:  tuple = CTA_LONG,
                               pw: int = CTA_PW,
                               sw: int = CTA_SW,
                               same_day: bool = False) -> pd.DataFrame:
    """
    Tradebook for the full Baz-Granger 3-pair CTA signal (Eq 33).

    S_CTA = (1/3) * sum(u_k)  for k in 3 timescales.
    Signal = sign(S_CTA).
    """
    price_vol = f1_raw.rolling(pw).std()
    u_cols    = {}
    for k, (sk, lk) in enumerate(zip(short_params, long_params), start=1):
        ewma_s  = _ewma(f1_raw, sk)
        ewma_l  = _ewma(f1_raw, lk)
        x       = ewma_s - ewma_l
        y       = x / price_vol
        with np.errstate(invalid="ignore"):
            z   = y / y.rolling(sw).std()
            u   = z * np.exp(-z.values**2 / 4) / 0.89
        u_cols[f"u_{k}(S={sk},L={lk})"] = u

    u_df      = pd.DataFrame(u_cols, index=f1_raw.index)
    s_cta     = u_df.mean(axis=1)
    signal    = np.sign(s_cta)
    position  = _build_position(signal, same_day)

    delta     = f1_cont.diff()
    daily_pnl = position * delta
    cum_pnl   = daily_pnl.cumsum()
    mtm       = position * f1_cont

    sp_str = "/".join(str(s) for s in short_params)
    lp_str = "/".join(str(l) for l in long_params)
    label  = f"CTA_Paper(S={sp_str},L={lp_str})" + (" [Same-Day]" if same_day else " [Lag-1]")

    base = {
        "Date"          : f1_raw.index,
        "Strategy"      : label,
        "F1_raw"        : f1_raw.round(4).values,
        "F1_continuous" : f1_cont.round(4).values,
    }
    for col, series in u_cols.items():
        base[col] = series.round(6).values

    base.update({
        "S_CTA"               : s_cta.round(6).values,
        "Signal"              : signal.values,
        "Position"            : position.values,
        "F1_cont_daily_change": delta.round(4).values,
        "Daily_PnL"           : daily_pnl.round(4).values,
        "MTM"                 : mtm.round(4).values,
        "Cum_PnL"             : cum_pnl.round(4).values,
    })
    return pd.DataFrame(base)


# ══════════════════════════════════════════════════════════════════════════════
# OPTIMIZATION HELPERS  (vectorised; no tradebook overhead)
# ══════════════════════════════════════════════════════════════════════════════

def _metrics_from_pnl(daily_pnl_np: np.ndarray,
                      pos_np: np.ndarray) -> dict:
    """Fast metrics from numpy arrays (used in tight optimization loops).
    Returns $/MT-based Sharpe for ranking; full % metrics are in compute_performance."""
    active_mask = (pos_np != 0) & np.isfinite(daily_pnl_np)
    active      = daily_pnl_np[active_mask]
    if len(active) < 20:
        return dict(sharpe=np.nan, ann_return=np.nan, max_drawdown=np.nan,
                    total_pnl=np.nan, hit_rate=np.nan)

    mu    = active.mean()
    sigma = active.std()
    ann_r = mu * 252
    shrp  = (ann_r / (sigma * np.sqrt(252))) if sigma > 0 else np.nan

    cum   = np.cumsum(daily_pnl_np[np.isfinite(daily_pnl_np)])
    mdd   = float((cum - np.maximum.accumulate(cum)).min())

    return dict(
        sharpe       = round(float(shrp),  4),
        ann_return   = round(float(ann_r), 4),
        max_drawdown = round(mdd,          4),
        total_pnl    = round(float(active.sum()), 4),
        hit_rate     = round(float((active > 0).mean()), 4),
    )


def optimize_ma_crossover(f1_raw: pd.Series,
                           f1_cont: pd.Series,
                           max_n: int = MAX_LOOKBACK) -> pd.DataFrame:
    """All (m,n) pairs with 1 <= m < n <= max_n. Returns df sorted by Sharpe (lag-1)."""
    print(f"  Pre-computing SMA windows 1..{max_n} ...")
    T = len(f1_raw)
    sma = np.full((T, max_n), np.nan)
    for k in range(max_n):
        sma[:, k] = f1_raw.rolling(k + 1).mean().values

    delta_np = f1_cont.diff().values.astype(float)

    print(f"  Scanning {max_n*(max_n-1)//2} MA pairs ...")
    records = []
    for m_idx in range(max_n - 1):
        m_sma = sma[:, m_idx]
        for n_idx in range(m_idx + 1, max_n):
            n_sma   = sma[:, n_idx]
            raw_sig = np.sign(m_sma - n_sma)
            pos     = np.empty(T)
            pos[0]  = 0.0
            pos[1:] = np.where(np.isfinite(raw_sig[:-1]), raw_sig[:-1], 0.0)
            dpnl    = pos * delta_np
            met     = _metrics_from_pnl(dpnl, pos)
            met.update({"m": m_idx + 1, "n": n_idx + 1})
            records.append(met)

    return (pd.DataFrame(records)
              [["m", "n", "sharpe", "ann_return", "max_drawdown", "total_pnl", "hit_rate"]]
              .sort_values("sharpe", ascending=False)
              .reset_index(drop=True))


def optimize_cta(f1_raw: pd.Series,
                 f1_cont: pd.Series,
                 max_n: int = MAX_LOOKBACK,
                 pw: int = CTA_PW,
                 sw: int = CTA_SW) -> pd.DataFrame:
    """Single-pair CTA scan over all (S,L) with 1 <= S < L <= max_n. Lag-1 only."""
    price_vol = f1_raw.rolling(pw).std()
    delta_np  = f1_cont.diff().values.astype(float)

    print(f"  Pre-computing EWMA spans 1..{max_n} ...")
    T        = len(f1_raw)
    ewma_arr = np.full((T, max_n), np.nan)
    for k in range(max_n):
        ewma_arr[:, k] = f1_raw.ewm(com=k, adjust=False).mean().values

    print(f"  Scanning {max_n*(max_n-1)//2} CTA pairs ...")
    records = []
    for s_idx in range(max_n - 1):
        s      = s_idx + 1
        ewma_s = pd.Series(ewma_arr[:, s_idx], index=f1_raw.index)
        for l_idx in range(s_idx + 1, max_n):
            l      = l_idx + 1
            ewma_l = pd.Series(ewma_arr[:, l_idx], index=f1_raw.index)
            x      = ewma_s - ewma_l
            y      = x / price_vol
            with np.errstate(invalid="ignore", divide="ignore"):
                y_std = y.rolling(sw).std()
                z     = (y / y_std).values
                u     = z * np.exp(-z**2 / 4) / 0.89
            raw_sig = np.sign(u)
            pos     = np.empty(T)
            pos[0]  = 0.0
            pos[1:] = np.where(np.isfinite(raw_sig[:-1]), raw_sig[:-1], 0.0)
            dpnl    = pos * delta_np
            met     = _metrics_from_pnl(dpnl, pos)
            met.update({"S": s, "L": l})
            records.append(met)

    return (pd.DataFrame(records)
              [["S", "L", "sharpe", "ann_return", "max_drawdown", "total_pnl", "hit_rate"]]
              .sort_values("sharpe", ascending=False)
              .reset_index(drop=True))


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _print_summary(label: str, met: dict) -> None:
    print(f"  [{label}]")
    for k in ["Sharpe Ratio", "Sortino Ratio", "Annualized Return (%)",
              "Max Drawdown (%)", "Calmar Ratio", "Total PnL (USD/MT)",
              "Hit Rate", "Profit Factor"]:
        print(f"    {k:<30} {met.get(k, 'n/a')}")


def _print_opt_table(df: pd.DataFrame, cols: list) -> None:
    display = cols + ["sharpe", "ann_return", "max_drawdown", "total_pnl", "hit_rate"]
    print(df[display].to_string(index=False))


def _make_tradebook_pair(build_fn, f1_raw, f1_cont, **kwargs):
    """Build both lag-1 and same-day tradebooks; return (tb_lag, met_lag, tb_same, met_same)."""
    tb_lag  = build_fn(f1_raw, f1_cont, same_day=False, **kwargs)
    pos_lag = pd.Series(tb_lag["Position"].values,   index=f1_raw.index)
    pnl_lag = pd.Series(tb_lag["Daily_PnL"].values,  index=f1_raw.index)
    met_lag = compute_performance(pnl_lag, pos_lag, f1_cont, same_day=False)

    tb_same  = build_fn(f1_raw, f1_cont, same_day=True, **kwargs)
    pos_same = pd.Series(tb_same["Position"].values,  index=f1_raw.index)
    pnl_same = pd.Series(tb_same["Daily_PnL"].values, index=f1_raw.index)
    met_same = compute_performance(pnl_same, pos_same, f1_cont, same_day=True)

    return tb_lag, met_lag, tb_same, met_same


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── Load data ─────────────────────────────────────────────────────────────
    print("=" * 70)
    print("Loading LME Copper continuous F1 series ...")
    print("=" * 70)
    rolling = get_metal_rolling_f1("LP", verbose=True)
    f1_raw  = rolling["F1_raw"]
    f1_cont = rolling["F1_continuous"]

    # ── CTA Paper signal (3 pairs, Baz-Granger fixed params) ─────────────────
    print("\n" + "=" * 70)
    print("CTA Paper signal  S=(8,16,32)  L=(24,48,96) ...")
    print("=" * 70)
    tb_lag, met_lag, tb_same, met_same = _make_tradebook_pair(
        build_cta_paper_tradebook, f1_raw, f1_cont
    )
    print("  Lag-1:"); _print_summary("CTA Paper [Lag-1]",    met_lag)
    print("  Same-Day:"); _print_summary("CTA Paper [Same-Day]", met_same)
    fpath = OUT_DIR / "CTA_Paper_Tradebook.xlsx"
    save_tradebook_excel(tb_lag, met_lag, tb_same, met_same, fpath)
    print(f"  Saved -> {fpath.name}")

    # ── MA Crossover Optimization ─────────────────────────────────────────────
    print("\n" + "=" * 70)
    print(f"MA Crossover optimization  n <= {MAX_LOOKBACK} ...")
    print("=" * 70)
    ma_opt = optimize_ma_crossover(f1_raw, f1_cont)
    ma_opt.to_csv(OUT_DIR / "MA_Crossover_Optimization.csv", index=False)
    print(f"  Saved MA_Crossover_Optimization.csv  ({len(ma_opt)} pairs, Lag-1 Sharpe ranking)")
    print("\n  TOP 5 by Sharpe (Lag-1):")
    _print_opt_table(ma_opt.head(5), cols=["m", "n"])

    top5_ma = ma_opt.head(5)
    print("\n  Saving top-5 MA tradebooks (both sheets) ...")
    for _, row in top5_ma.iterrows():
        m, n = int(row["m"]), int(row["n"])
        tb_lag, met_lag, tb_same, met_same = _make_tradebook_pair(
            build_ma_tradebook, f1_raw, f1_cont, m=m, n=n
        )
        fpath = OUT_DIR / f"MA_Top5_Tradebook_{m}_{n}.xlsx"
        save_tradebook_excel(tb_lag, met_lag, tb_same, met_same, fpath)
        print(f"    MA({m},{n})  Lag-1 Sharpe={met_lag['Sharpe Ratio']}  "
              f"SameDay Sharpe={met_same['Sharpe Ratio']}  -> {fpath.name}")

    # ── CTA Optimization ──────────────────────────────────────────────────────
    print("\n" + "=" * 70)
    print(f"CTA single-pair optimization  L <= {MAX_LOOKBACK} ...")
    print("=" * 70)
    cta_opt = optimize_cta(f1_raw, f1_cont)
    cta_opt.to_csv(OUT_DIR / "CTA_Optimization.csv", index=False)
    print(f"  Saved CTA_Optimization.csv  ({len(cta_opt)} pairs, Lag-1 Sharpe ranking)")
    print("\n  TOP 5 by Sharpe (Lag-1):")
    _print_opt_table(cta_opt.head(5), cols=["S", "L"])

    top5_cta = cta_opt.head(5)
    print("\n  Saving top-5 CTA tradebooks (both sheets) ...")
    for _, row in top5_cta.iterrows():
        s, l = int(row["S"]), int(row["L"])
        tb_lag, met_lag, tb_same, met_same = _make_tradebook_pair(
            build_cta_single_tradebook, f1_raw, f1_cont, s=s, l=l
        )
        fpath = OUT_DIR / f"CTA_Top5_Tradebook_{s}_{l}.xlsx"
        save_tradebook_excel(tb_lag, met_lag, tb_same, met_same, fpath)
        print(f"    CTA({s},{l})  Lag-1 Sharpe={met_lag['Sharpe Ratio']}  "
              f"SameDay Sharpe={met_same['Sharpe Ratio']}  -> {fpath.name}")

    print(f"\nAll outputs -> {OUT_DIR.resolve()}")
    print(f"  11 .xlsx tradebook files (2 sheets each: Lag-1 + Same-Day)")
    print(f"   2 .csv  optimization files (Lag-1 ranking, 7875 pairs each)")


if __name__ == "__main__":
    main()
